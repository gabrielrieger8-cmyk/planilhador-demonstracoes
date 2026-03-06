"""Geração de arquivos Excel/CSV a partir dos dados parseados."""

from __future__ import annotations

import csv
import logging
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

VBA_TEMPLATE = Path(__file__).parent.parent / "vba" / "template.xlsm"

logger = logging.getLogger("planilhador")

# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

AGRUPADORA_FONT = Font(name="Calibri", bold=True, size=11)
AGRUPADORA_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

NORMAL_FONT = Font(name="Calibri", size=11)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

BR_NUMBER_FORMAT = '#,##0.00'

TIPO_LABELS = {
    "balancete": "Balancete",
    "dre": "DRE",
    "balanco_patrimonial": "Balanço Patrimonial",
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _sanitize_tab_name(name: str) -> str:
    """Sanitiza nome de aba Excel (max 31 chars, sem caracteres proibidos)."""
    sanitized = re.sub(r'[/\\*?\[\]:]', '.', name)
    return sanitized[:31]


def _build_title(empresa: str, tipo: str, periodo: str) -> str:
    """Constrói título padronizado: Empresa - Tipo - Período."""
    label = TIPO_LABELS.get(tipo, tipo)
    parts = [p for p in [empresa, label, periodo] if p]
    return " - ".join(parts)


# ---------------------------------------------------------------------------
# API pública
# ---------------------------------------------------------------------------


def _unique_tab_name(name: str, used: set[str]) -> str:
    """Garante nome de aba único, adicionando sufixo se necessário."""
    base = _sanitize_tab_name(name)
    if base not in used:
        used.add(base)
        return base
    for n in range(2, 100):
        suffix = f" ({n})"
        candidate = _sanitize_tab_name(name[:31 - len(suffix)] + suffix)
        if candidate not in used:
            used.add(candidate)
            return candidate
    return base


_MESES_ABREV = [
    "", "jan", "fev", "mar", "abr", "mai", "jun",
    "jul", "ago", "set", "out", "nov", "dez",
]


def _periodo_to_short(periodo: str) -> str:
    """Converte período como '01.12.2025 A 31.12.2025' em 'dez/25'."""
    if not periodo:
        return ""
    # Tenta extrair mês/ano do último date-like token (dd.mm.yyyy ou mm.yyyy ou mm/yyyy)
    matches = re.findall(r'(\d{1,2})[./](\d{4})', periodo)
    if matches:
        mes_str, ano = matches[-1]
        mes = int(mes_str)
        if 1 <= mes <= 12:
            return f"{_MESES_ABREV[mes]}/{ano[2:]}"
    return periodo


def _short_tab_name(tipo: str, periodo: str) -> str:
    """Nome curto para aba: 'Balancete dez/25'."""
    label = TIPO_LABELS.get(tipo, tipo)
    short = _periodo_to_short(periodo)
    if short:
        return f"{label} {short}"
    return label


def export_excel_multi(
    demonstracoes: list[dict],
    empresa: str,
    output_path: Path,
    formula_opts: dict[str, bool] | None = None,
    append_to: Path | None = None,
    include_vba: bool = False,
) -> Path:
    """Gera Excel com abas agrupadas por tipo de demonstração.

    Para DRE e Balanço com múltiplos períodos, coloca todos os períodos
    lado a lado em uma única aba (formato comparativo).

    Args:
        demonstracoes: Lista de dicts com {tipo, periodo, dados}.
        empresa: Nome da empresa.
        output_path: Caminho para salvar o arquivo.
        formula_opts: Dict com chaves 'dre', 'balanco', 'balancete' (bool).
        append_to: Se fornecido, abre este workbook existente e adiciona abas.
        include_vba: Se True, embute macros VBA do consolidador (salva como .xlsm).

    Returns:
        Path do arquivo gerado.
    """
    if formula_opts is None:
        formula_opts = {"dre": False, "balanco": True, "balancete": False}

    # Decide se abre workbook existente ou cria novo
    if append_to and append_to.exists():
        wb = load_workbook(str(append_to), keep_vba=True)
        default_ws = None
        used_names: set[str] = {ws.title for ws in wb.worksheets}
    elif include_vba and VBA_TEMPLATE.exists():
        wb = load_workbook(str(VBA_TEMPLATE), keep_vba=True)
        # Remove a sheet padrão do template
        default_ws = wb.active
        used_names: set[str] = set()
    else:
        wb = Workbook()
        default_ws = wb.active
        used_names: set[str] = set()

    # Agrupa demonstrações por tipo, preservando ordem de aparição
    groups: list[tuple[str, list[dict]]] = []
    seen_tipos: dict[str, int] = {}
    for demo in demonstracoes:
        tipo = demo["tipo"]
        if tipo not in seen_tipos:
            seen_tipos[tipo] = len(groups)
            groups.append((tipo, []))
        groups[seen_tipos[tipo]][1].append(demo)

    def _get_ws(tab_name: str) -> object:
        """Cria nova sheet ou reutiliza default_ws na primeira aba."""
        nonlocal default_ws
        if default_ws is not None:
            default_ws.title = tab_name
            ws = default_ws
            default_ws = None  # Consumed
            return ws
        return wb.create_sheet(title=tab_name)

    tab_idx = 0
    for tipo, demos in groups:
        # DRE/Balanço multi-período: comparativo lado a lado
        if len(demos) > 1 and tipo in ("dre", "balanco_patrimonial"):
            label = TIPO_LABELS.get(tipo, tipo)
            tab_name = _unique_tab_name(label, used_names)
            ws = _get_ws(tab_name)

            titulo = f"{empresa} - {label}" if empresa else label

            if tipo == "dre":
                _write_dre_comparativo(ws, demos, titulo, use_formulas=formula_opts.get("dre", False))
            else:
                _write_balanco_comparativo(ws, demos, titulo, use_formulas=formula_opts.get("balanco", True))

            tab_idx += 1
        else:
            # Single-period ou balancete: 1 aba por demonstração
            for demo in demos:
                periodo = demo.get("periodo", "")
                dados = demo.get("dados", {})

                if len(demos) > 1 or len(groups) > 1:
                    raw_name = _short_tab_name(tipo, periodo)
                else:
                    raw_name = _build_title(empresa, tipo, periodo)
                tab_name = _unique_tab_name(raw_name, used_names)
                ws = _get_ws(tab_name)

                titulo = _build_title(empresa, tipo, periodo)

                if tipo == "balancete":
                    _write_balancete(ws, dados, titulo, use_formulas=formula_opts.get("balancete", False))
                elif tipo == "dre":
                    _write_dre(ws, dados, titulo, use_formulas=formula_opts.get("dre", False))
                elif tipo == "balanco_patrimonial":
                    _write_balanco(ws, dados, titulo, use_formulas=formula_opts.get("balanco", True))

                tab_idx += 1

    # Remove sheet padrão vazia do template VBA (se não foi usada)
    if default_ws is not None and len(wb.worksheets) > 1:
        wb.remove(default_ws)

    # Ajusta extensão para .xlsm se tem VBA
    if wb.vba_archive and output_path.suffix.lower() == ".xlsx":
        output_path = output_path.with_suffix(".xlsm")

    wb.save(str(output_path))
    logger.info("Excel gerado: %s (%d aba(s))", output_path, tab_idx)
    return output_path


def export_excel(dados: dict, tipo: str, output_path: Path) -> Path:
    """Exporta dados de uma única demonstração para Excel."""
    empresa = dados.get("empresa", "")
    periodo = dados.get("periodo", dados.get("data_referencia", ""))
    demonstracoes = [{"tipo": tipo, "periodo": periodo, "dados": dados}]
    return export_excel_multi(demonstracoes, empresa, output_path)


def export_csv(dados: dict, tipo: str, output_path: Path) -> Path:
    """Exporta dados como CSV (delimitador ;, UTF-8 BOM)."""
    if tipo == "balancete":
        return _export_balancete_csv(dados, output_path)
    elif tipo == "dre":
        return _export_dre_csv(dados, output_path)
    elif tipo == "balanco_patrimonial":
        return _export_balanco_csv(dados, output_path)
    else:
        raise ValueError(f"Tipo não suportado para exportação: {tipo}")


# ---------------------------------------------------------------------------
# Balancete
# ---------------------------------------------------------------------------

BALANCETE_COLUMNS = [
    "Código", "Classificação", "Descrição", "Nível", "Natureza",
    "Saldo Anterior", "Débitos", "Créditos", "Saldo Atual",
]
BALANCETE_NUMERIC_COLS = {5, 6, 7, 8}
BALANCETE_COL_WIDTHS = {0: 14, 1: 18, 2: 42, 3: 8, 4: 10, 5: 18, 6: 18, 7: 18, 8: 18}


def _write_balancete(ws, dados: dict, titulo: str, use_formulas: bool = False) -> None:
    """Escreve conteúdo de balancete em um worksheet, opcionalmente com fórmulas SUM."""
    contas = dados.get("contas", [])

    ws.append([titulo])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(BALANCETE_COLUMNS))
    ws["A1"].font = Font(name="Calibri", bold=True, size=14)

    cnpj = dados.get("cnpj", "")
    periodo = dados.get("periodo", "")
    if cnpj:
        ws.append([f"CNPJ: {cnpj}  |  Período: {periodo}"])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(BALANCETE_COLUMNS))
        ws["A2"].font = Font(name="Calibri", size=11, color="666666")
    ws.append([])

    ws.append(BALANCETE_COLUMNS)
    header_row = ws.max_row
    for col_idx in range(len(BALANCETE_COLUMNS)):
        cell = ws.cell(row=header_row, column=col_idx + 1)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Mapeia classificação → row number para fórmulas de totalizadoras
    classif_to_row = {}

    for conta in contas:
        row = [
            conta.get("codigo_conta", ""),
            conta.get("classificacao", ""),
            conta.get("descricao", ""),
            conta.get("nivel", ""),
            conta.get("natureza", ""),
            conta.get("saldo_anterior", 0) or 0,
            conta.get("debitos", 0) or 0,
            conta.get("creditos", 0) or 0,
            conta.get("saldo_atual", 0) or 0,
        ]
        ws.append(row)

        current_row = ws.max_row
        classif = conta.get("classificacao", "")
        classif_to_row[classif] = current_row
        is_totalizador = conta.get("is_totalizador", False)

        for col_idx in range(len(BALANCETE_COLUMNS)):
            cell = ws.cell(row=current_row, column=col_idx + 1)
            cell.border = THIN_BORDER
            cell.font = AGRUPADORA_FONT if is_totalizador else NORMAL_FONT

            if is_totalizador:
                cell.fill = AGRUPADORA_FILL

            if col_idx in BALANCETE_NUMERIC_COLS:
                cell.number_format = BR_NUMBER_FORMAT
                cell.alignment = RIGHT_ALIGN
            elif col_idx in (3, 4):
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = LEFT_ALIGN

    # Segunda passada: adiciona fórmulas SUM nas totalizadoras (se habilitado)
    if use_formulas:
        for conta in contas:
            if not conta.get("is_totalizador"):
                continue
            classif = conta.get("classificacao", "")
            parent_row = classif_to_row.get(classif)
            if not parent_row:
                continue

            # Encontra filhas diretas (classificação = parent + ".XX")
            child_rows = []
            for other in contas:
                other_classif = other.get("classificacao", "")
                if (other_classif != classif
                        and other_classif.startswith(classif + ".")
                        and other_classif[len(classif) + 1:].count(".") == 0):
                    child_row = classif_to_row.get(other_classif)
                    if child_row:
                        child_rows.append(child_row)

            if child_rows:
                # Colunas numéricas: F (Saldo Ant), G (Déb), H (Créd), I (Saldo Atual)
                for col_letter in ("F", "G", "H", "I"):
                    refs = "+".join(f"{col_letter}{r}" for r in sorted(child_rows))
                    ws.cell(row=parent_row, column=_col_idx(col_letter)).value = f"={refs}"

    for col_idx, width in BALANCETE_COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = width

    ws.freeze_panes = f"A{header_row + 1}"

    if contas:
        last_row = ws.max_row
        last_col = get_column_letter(len(BALANCETE_COLUMNS))
        ws.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"


def _col_idx(letter: str) -> int:
    """Converte letra de coluna em índice (A=1, B=2, etc.)."""
    return ord(letter.upper()) - ord("A") + 1


def _export_balancete_csv(dados: dict, output_path: Path) -> Path:
    contas = dados.get("contas", [])

    with open(str(output_path), "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(BALANCETE_COLUMNS)
        for conta in contas:
            writer.writerow([
                conta.get("codigo_conta", ""),
                conta.get("classificacao", ""),
                conta.get("descricao", ""),
                conta.get("nivel", ""),
                conta.get("natureza", ""),
                conta.get("saldo_anterior", 0),
                conta.get("debitos", 0),
                conta.get("creditos", 0),
                conta.get("saldo_atual", 0),
            ])

    logger.info("CSV balancete gerado: %s (%d contas)", output_path, len(contas))
    return output_path


# ---------------------------------------------------------------------------
# DRE
# ---------------------------------------------------------------------------

DRE_COLUMNS = ["Classificação", "Descrição", "Valor"]
DRE_COL_WIDTHS = {0: 14, 1: 50, 2: 20}


def _write_dre(ws, dados: dict, titulo: str, use_formulas: bool = True) -> None:
    linhas = dados.get("linhas", [])

    ws.append([titulo])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(DRE_COLUMNS))
    ws["A1"].font = Font(name="Calibri", bold=True, size=14)

    periodo = dados.get("periodo", "")
    if periodo:
        ws.append([f"Período: {periodo}"])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(DRE_COLUMNS))
        ws["A2"].font = Font(name="Calibri", size=11, color="666666")
    ws.append([])

    ws.append(DRE_COLUMNS)
    header_row = ws.max_row  # Row onde os headers foram escritos
    for col_idx in range(len(DRE_COLUMNS)):
        cell = ws.cell(row=header_row, column=col_idx + 1)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    VAL_COL = 3   # coluna C (Valor)

    # Rastreia linhas para fórmulas hierárquicas
    current_grouper_row = None
    current_grouper_children = []
    prev_subtotal_row = None
    subtotal_refs = []  # top-level rows entre subtotais

    def _close_grouper():
        nonlocal current_grouper_row, current_grouper_children
        if current_grouper_row and current_grouper_children and use_formulas:
            refs = ",".join(f"C{r}" for r in current_grouper_children)
            ws.cell(row=current_grouper_row, column=VAL_COL).value = f"=SUM({refs})"
        current_grouper_row = None
        current_grouper_children = []

    for linha in linhas:
        valor = linha.get("valor", 0) or 0
        nivel = linha.get("nivel", 1)
        is_subtotal = linha.get("is_subtotal", False)
        is_agrupadora = linha.get("is_agrupadora", False)
        classificacao = linha.get("classificacao", "")

        indent = "  " * (nivel - 1)
        descricao = f"{indent}{linha.get('descricao', '')}"

        ws.append([classificacao, descricao, valor])
        current_row = ws.max_row

        if use_formulas:
            if is_subtotal:
                _close_grouper()
                # Subtotal = prev_subtotal + top-level rows entre eles
                all_refs = []
                if prev_subtotal_row:
                    all_refs.append(prev_subtotal_row)
                all_refs.extend(subtotal_refs)
                if all_refs:
                    refs_str = ",".join(f"C{r}" for r in all_refs)
                    ws.cell(row=current_row, column=VAL_COL).value = f"=SUM({refs_str})"
                prev_subtotal_row = current_row
                subtotal_refs = []
            elif is_agrupadora:
                _close_grouper()
                current_grouper_row = current_row
                current_grouper_children = []
                subtotal_refs.append(current_row)
            elif nivel == 1:
                # nivel=1 não-subtotal, não-agrupadora (ex: RECEITA BRUTA)
                _close_grouper()
                subtotal_refs.append(current_row)
            else:
                # nivel=2 detail (folha) — mantém valor hardcoded do PDF
                if current_grouper_row:
                    current_grouper_children.append(current_row)
                else:
                    subtotal_refs.append(current_row)

        # Estilo
        for col_idx in range(len(DRE_COLUMNS)):
            cell = ws.cell(row=current_row, column=col_idx + 1)
            cell.border = THIN_BORDER

            if is_subtotal or nivel <= 1:
                cell.font = AGRUPADORA_FONT
                cell.fill = AGRUPADORA_FILL
            else:
                cell.font = NORMAL_FONT

            if col_idx == 0:  # Classificação
                cell.alignment = LEFT_ALIGN
            elif col_idx == 1:  # Descrição
                cell.alignment = LEFT_ALIGN
            elif col_idx == 2:  # Valor
                cell.number_format = BR_NUMBER_FORMAT
                cell.alignment = RIGHT_ALIGN

    _close_grouper()  # finaliza última agrupadora pendente

    for col_idx, width in DRE_COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = width

    ws.freeze_panes = f"A{header_row + 1}"


def _write_dre_comparativo(ws, demos_list: list[dict], titulo: str, use_formulas: bool = True) -> None:
    """Escreve DRE comparativa com períodos lado a lado + AV% + Variação."""
    periodos = []
    all_linhas = []
    for demo in demos_list:
        dados = demo.get("dados", {})
        periodo = demo.get("periodo", dados.get("periodo", ""))
        periodos.append(periodo)
        all_linhas.append(dados.get("linhas", []))

    num_periods = len(periodos)

    # Layout de colunas: Descrição | Per1 | Per2 | ...
    # Coluna de valor do período p: 2 + p (B, C, D, ...)
    def val_col(p: int) -> int:
        return 2 + p

    num_cols = 1 + num_periods

    # Headers
    headers = [""] + list(periodos)

    # Título
    ws.append([titulo])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    ws["A1"].font = Font(name="Calibri", bold=True, size=14)
    ws.append([])

    # Headers
    ws.append(headers)
    header_row = ws.max_row
    for col_idx in range(num_cols):
        cell = ws.cell(row=header_row, column=col_idx + 1)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Usa primeiro período como template para descrições
    template = all_linhas[0] if all_linhas else []

    # Rastreia rows para fórmulas hierárquicas
    current_grouper_row = None
    current_grouper_children = []
    prev_subtotal_row = None
    subtotal_refs = []  # top-level rows entre subtotais

    def _close_current_grouper():
        nonlocal current_grouper_row, current_grouper_children
        if current_grouper_row and current_grouper_children and use_formulas:
            for p in range(num_periods):
                col = val_col(p)
                cl = get_column_letter(col)
                refs = ",".join(f"{cl}{r}" for r in current_grouper_children)
                ws.cell(row=current_grouper_row, column=col).value = f"=SUM({refs})"
        current_grouper_row = None
        current_grouper_children = []

    for row_idx, linha in enumerate(template):
        nivel = linha.get("nivel", 1)
        is_subtotal = linha.get("is_subtotal", False)
        is_agrupadora = linha.get("is_agrupadora", False)
        indent = "  " * (nivel - 1)
        descricao = f"{indent}{linha.get('descricao', '')}"

        # Monta row: [desc, val1, val2, ...]
        row_data = [descricao]
        for p in range(num_periods):
            linhas_p = all_linhas[p]
            val = linhas_p[row_idx].get("valor", 0) or 0 if row_idx < len(linhas_p) else 0
            row_data.append(val)

        ws.append(row_data)
        current_row = ws.max_row

        if use_formulas:
            if is_subtotal:
                _close_current_grouper()
                # Subtotal = prev_subtotal + top-level rows entre eles
                all_refs = []
                if prev_subtotal_row:
                    all_refs.append(prev_subtotal_row)
                all_refs.extend(subtotal_refs)
                if all_refs:
                    for p in range(num_periods):
                        col = val_col(p)
                        cl = get_column_letter(col)
                        refs_str = ",".join(f"{cl}{r}" for r in all_refs)
                        ws.cell(row=current_row, column=col).value = f"=SUM({refs_str})"
                prev_subtotal_row = current_row
                subtotal_refs = []
            elif is_agrupadora:
                _close_current_grouper()
                current_grouper_row = current_row
                current_grouper_children = []
                subtotal_refs.append(current_row)
            elif nivel == 1:
                _close_current_grouper()
                subtotal_refs.append(current_row)
            else:
                # nivel=2 detail (folha) — mantém valor hardcoded do PDF
                if current_grouper_row:
                    current_grouper_children.append(current_row)
                else:
                    subtotal_refs.append(current_row)

        # Estilo
        for col_idx in range(num_cols):
            cell = ws.cell(row=current_row, column=col_idx + 1)
            cell.border = THIN_BORDER

            if is_subtotal or nivel <= 1:
                cell.font = AGRUPADORA_FONT
                cell.fill = AGRUPADORA_FILL
            else:
                cell.font = NORMAL_FONT

            if col_idx == 0:
                cell.alignment = LEFT_ALIGN
            else:
                cell.alignment = RIGHT_ALIGN
                cell.number_format = BR_NUMBER_FORMAT

    _close_current_grouper()  # finaliza última agrupadora pendente

    # Larguras
    ws.column_dimensions["A"].width = 50
    for p in range(num_periods):
        ws.column_dimensions[get_column_letter(val_col(p))].width = 20

    ws.freeze_panes = f"A{header_row + 1}"


def _export_dre_csv(dados: dict, output_path: Path) -> Path:
    linhas = dados.get("linhas", [])

    with open(str(output_path), "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(["Classificação", "Descrição", "Valor", "Nível", "Subtotal"])
        for linha in linhas:
            writer.writerow([
                linha.get("classificacao", ""),
                linha.get("descricao", ""),
                linha.get("valor", 0),
                linha.get("nivel", 1),
                "Sim" if linha.get("is_subtotal") else "Não",
            ])

    logger.info("CSV DRE gerado: %s (%d linhas)", output_path, len(linhas))
    return output_path


# ---------------------------------------------------------------------------
# Balanço Patrimonial
# ---------------------------------------------------------------------------

BALANCO_COLUMNS = ["Classificação", "Descrição", "Valor"]
BALANCO_COL_WIDTHS = {0: 14, 1: 50, 2: 20}


def _write_balanco(ws, dados: dict, titulo: str, use_formulas: bool = True) -> None:
    VAL_COL = 3  # coluna C (Valor)

    ws.append([titulo])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(BALANCO_COLUMNS))
    ws["A1"].font = Font(name="Calibri", bold=True, size=14)

    data_ref = dados.get("data_referencia", "")
    if data_ref:
        ws.append([f"Data de Referência: {data_ref}"])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(BALANCO_COLUMNS))
        ws["A2"].font = Font(name="Calibri", size=11, color="666666")
    ws.append([])

    # Rastreia rows para fórmulas SUM
    section_total_rows = {}  # "ativo", "passivo", "pl" → row number
    sub_total_rows = {}  # ("ativo","circulante") → row number

    section_font = Font(name="Calibri", bold=True, size=13, color="2F5496")

    def _style_section_row(row_num):
        for c in range(1, len(BALANCO_COLUMNS) + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.font = section_font
            cell.border = THIN_BORDER
        ws.cell(row=row_num, column=VAL_COL).number_format = BR_NUMBER_FORMAT
        ws.cell(row=row_num, column=VAL_COL).alignment = RIGHT_ALIGN

    def _write_section(title: str, section: dict, section_key: str):
        # Header da seção: fórmula ou valor estático
        section_total = section.get("total", 0) or 0
        ws.append(["", title, 0 if use_formulas else section_total])
        section_row = ws.max_row
        section_total_rows[section_key] = section_row
        _style_section_row(section_row)

        sub_rows_for_section = []

        for sub_key in ("circulante", "nao_circulante"):
            sub = section.get(sub_key, {})
            if not sub:
                continue

            sub_title = "Circulante" if sub_key == "circulante" else "Não Circulante"
            sub_total = sub.get("total", 0) or 0
            ws.append(["", f"  {sub_title}", 0 if use_formulas else sub_total])
            sub_header_row = ws.max_row
            sub_total_rows[(section_key, sub_key)] = sub_header_row
            sub_rows_for_section.append(sub_header_row)
            for c in range(1, len(BALANCO_COLUMNS) + 1):
                cell = ws.cell(row=sub_header_row, column=c)
                cell.font = AGRUPADORA_FONT
                cell.fill = AGRUPADORA_FILL
                cell.border = THIN_BORDER
            ws.cell(row=sub_header_row, column=VAL_COL).number_format = BR_NUMBER_FORMAT
            ws.cell(row=sub_header_row, column=VAL_COL).alignment = RIGHT_ALIGN

            conta_rows = []
            for conta in sub.get("contas", []):
                nivel = conta.get("nivel", 3)
                is_sub = conta.get("is_subtotal", False)
                indent = "    " * max(1, nivel - 2)
                classif = conta.get("classificacao", "")
                ws.append([classif, f"{indent}{conta.get('descricao', '')}", conta.get("valor", 0)])
                current_row = ws.max_row
                conta_rows.append(current_row)
                for c in range(1, len(BALANCO_COLUMNS) + 1):
                    cell = ws.cell(row=current_row, column=c)
                    cell.border = THIN_BORDER
                    if is_sub:
                        cell.font = AGRUPADORA_FONT
                        cell.fill = AGRUPADORA_FILL
                    if c == VAL_COL:
                        cell.number_format = BR_NUMBER_FORMAT
                        cell.alignment = RIGHT_ALIGN

            # Fórmula SUM no header da subsecção
            if use_formulas and conta_rows:
                first_r, last_r = conta_rows[0], conta_rows[-1]
                ws.cell(row=sub_header_row, column=VAL_COL).value = f"=SUM(C{first_r}:C{last_r})"

        # Fórmula SUM no header da seção (soma das subsecções)
        if use_formulas and sub_rows_for_section:
            refs = "+".join(f"C{r}" for r in sub_rows_for_section)
            ws.cell(row=section_row, column=VAL_COL).value = f"={refs}"

        ws.append([])

    _write_section("ATIVO", dados.get("ativo", {}), "ativo")
    # ---- PASSIVO (inclui Patrimônio Líquido) ----
    pl = dados.get("patrimonio_liquido", {})
    passivo_data = dados.get("passivo", {})
    if use_formulas:
        passivo_header_val = 0
    else:
        passivo_header_val = (passivo_data.get("total", 0) or 0) + (pl.get("total", 0) or 0)
    ws.append(["", "PASSIVO", passivo_header_val])
    passivo_row = ws.max_row
    section_total_rows["passivo"] = passivo_row
    _style_section_row(passivo_row)

    passivo_sub_rows = []  # rows das subsecções para fórmula do total

    for sub_key in ("circulante", "nao_circulante"):
        sub = passivo_data.get(sub_key, {})
        if not sub:
            continue

        sub_title = "Circulante" if sub_key == "circulante" else "Não Circulante"
        sub_total = sub.get("total", 0) or 0
        ws.append(["", f"  {sub_title}", 0 if use_formulas else sub_total])
        sub_header_row = ws.max_row
        passivo_sub_rows.append(sub_header_row)
        for c in range(1, len(BALANCO_COLUMNS) + 1):
            cell = ws.cell(row=sub_header_row, column=c)
            cell.font = AGRUPADORA_FONT
            cell.fill = AGRUPADORA_FILL
            cell.border = THIN_BORDER
        ws.cell(row=sub_header_row, column=VAL_COL).number_format = BR_NUMBER_FORMAT
        ws.cell(row=sub_header_row, column=VAL_COL).alignment = RIGHT_ALIGN

        conta_rows = []
        for conta in sub.get("contas", []):
            nivel = conta.get("nivel", 3)
            is_sub = conta.get("is_subtotal", False)
            indent = "    " * max(1, nivel - 2)
            classif = conta.get("classificacao", "")
            ws.append([classif, f"{indent}{conta.get('descricao', '')}", conta.get("valor", 0)])
            current_row = ws.max_row
            conta_rows.append(current_row)
            for c in range(1, len(BALANCO_COLUMNS) + 1):
                cell = ws.cell(row=current_row, column=c)
                cell.border = THIN_BORDER
                if is_sub:
                    cell.font = AGRUPADORA_FONT
                    cell.fill = AGRUPADORA_FILL
                if c == VAL_COL:
                    cell.number_format = BR_NUMBER_FORMAT
                    cell.alignment = RIGHT_ALIGN

        if use_formulas and conta_rows:
            first_r, last_r = conta_rows[0], conta_rows[-1]
            ws.cell(row=sub_header_row, column=VAL_COL).value = f"=SUM(C{first_r}:C{last_r})"

    # Patrimônio Líquido (subsecção do PASSIVO)
    if pl.get("contas"):
        pl_total = pl.get("total", 0) or 0
        ws.append(["", "  Patrimônio Líquido", 0 if use_formulas else pl_total])
        pl_sub_row = ws.max_row
        passivo_sub_rows.append(pl_sub_row)
        for c in range(1, len(BALANCO_COLUMNS) + 1):
            cell = ws.cell(row=pl_sub_row, column=c)
            cell.font = AGRUPADORA_FONT
            cell.fill = AGRUPADORA_FILL
            cell.border = THIN_BORDER
        ws.cell(row=pl_sub_row, column=VAL_COL).number_format = BR_NUMBER_FORMAT
        ws.cell(row=pl_sub_row, column=VAL_COL).alignment = RIGHT_ALIGN

        pl_conta_rows = []
        for conta in pl.get("contas", []):
            is_sub = conta.get("is_subtotal", False)
            classif = conta.get("classificacao", "")
            ws.append([classif, f"    {conta.get('descricao', '')}", conta.get("valor", 0)])
            current_row = ws.max_row
            pl_conta_rows.append(current_row)
            for c in range(1, len(BALANCO_COLUMNS) + 1):
                cell = ws.cell(row=current_row, column=c)
                cell.border = THIN_BORDER
                if is_sub:
                    cell.font = AGRUPADORA_FONT
                    cell.fill = AGRUPADORA_FILL
                if c == VAL_COL:
                    cell.number_format = BR_NUMBER_FORMAT
                    cell.alignment = RIGHT_ALIGN

        if use_formulas and pl_conta_rows:
            first_r, last_r = pl_conta_rows[0], pl_conta_rows[-1]
            ws.cell(row=pl_sub_row, column=VAL_COL).value = f"=SUM(C{first_r}:C{last_r})"

    # Fórmula PASSIVO total = PC + PNC + PL
    if use_formulas and passivo_sub_rows:
        refs = "+".join(f"C{r}" for r in passivo_sub_rows)
        ws.cell(row=passivo_row, column=VAL_COL).value = f"={refs}"

    ws.append([])

    # Validação: Ativo = Passivo (Passivo já inclui PL)
    ativo_row = section_total_rows.get("ativo")

    ws.append(["", "VALIDAÇÃO: Ativo = Passivo"])
    ws.cell(row=ws.max_row, column=2).font = Font(name="Calibri", bold=True, size=11)

    if use_formulas and ativo_row and passivo_row:
        ws.append([
            "",
            "Diferença (Ativo - Passivo):",
            None,
        ])
        diff_row = ws.max_row
        ws.cell(row=diff_row, column=VAL_COL).value = f"=C{ativo_row}-C{passivo_row}"
        ws.cell(row=diff_row, column=VAL_COL).number_format = BR_NUMBER_FORMAT
        ws.cell(row=diff_row, column=VAL_COL).alignment = RIGHT_ALIGN

        ws.append(["", "Status:"])
        status_row = ws.max_row
        ws.cell(row=status_row, column=VAL_COL).value = (
            f'=IF(ABS(C{diff_row})<0.01,"OK","DIVERGENTE")'
        )
        status_cell = ws.cell(row=status_row, column=VAL_COL)
        status_cell.font = Font(name="Calibri", bold=True, size=11)
    else:
        total_ativo = dados.get("ativo", {}).get("total", 0) or 0
        total_passivo = passivo_data.get("total", 0) or 0
        total_pl = pl.get("total", 0) or 0
        passivo_pl = total_passivo + total_pl
        valido = abs(total_ativo - passivo_pl) < max(abs(total_ativo), 0.01) * 0.01
        ws.append([
            "",
            f"Ativo Total: {total_ativo:,.2f}  |  Passivo: {passivo_pl:,.2f}  |  "
            f"{'OK' if valido else 'DIVERGENTE'}",
        ])
        status_cell = ws.cell(row=ws.max_row, column=2)
        status_cell.font = Font(
            name="Calibri", bold=True, size=11,
            color="006100" if valido else "9C0006",
        )

    for col_idx, width in BALANCO_COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = width


def _write_balanco_comparativo(ws, demos_list: list[dict], titulo: str, use_formulas: bool = True) -> None:
    """Escreve Balanço comparativo com períodos lado a lado em uma única aba."""
    periodos = []
    all_dados = []
    for demo in demos_list:
        dados = demo.get("dados", {})
        periodo = demo.get("periodo", dados.get("data_referencia", ""))
        periodos.append(periodo)
        all_dados.append(dados)

    num_periods = len(periodos)
    headers = [""] + periodos
    num_cols = len(headers)

    # Título
    ws.append([titulo])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    ws["A1"].font = Font(name="Calibri", bold=True, size=14)
    ws.append([])

    # Headers
    ws.append(headers)
    header_row = ws.max_row
    for col_idx in range(num_cols):
        cell = ws.cell(row=header_row, column=col_idx + 1)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    section_title_font = Font(name="Calibri", bold=True, size=13, color="2F5496")

    def _style_row(row_num, font=NORMAL_FONT, fill=None):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.font = font
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
            if c > 1:
                cell.number_format = BR_NUMBER_FORMAT
                cell.alignment = RIGHT_ALIGN
            else:
                cell.alignment = LEFT_ALIGN

    # Rastreia rows para fórmulas SUM
    section_rows = {}  # "ativo" / "passivo" → row number

    def _write_section(title: str, section_key: str):
        # Header da seção (ATIVO) com placeholder ou valor estático
        row_data = [title]
        for dados in all_dados:
            section = dados.get(section_key, {})
            row_data.append(0 if use_formulas else (section.get("total", 0) or 0))
        ws.append(row_data)
        section_row = ws.max_row
        section_rows[section_key] = section_row
        _style_row(section_row, font=section_title_font)

        sub_header_rows = []  # rows das subsecções para fórmula do total

        for sub_key, sub_title in [("circulante", "CIRCULANTE"), ("nao_circulante", "NÃO CIRCULANTE")]:
            any_contas = any(
                d.get(section_key, {}).get(sub_key, {}).get("contas", [])
                for d in all_dados
            )
            if not any_contas:
                continue

            # Header da subsecção com placeholder ou valor estático
            row_data = [sub_title]
            for dados in all_dados:
                sub = dados.get(section_key, {}).get(sub_key, {})
                row_data.append(0 if use_formulas else (sub.get("total", 0) or 0))
            ws.append(row_data)
            sub_header_row = ws.max_row
            sub_header_rows.append(sub_header_row)
            _style_row(sub_header_row, font=AGRUPADORA_FONT, fill=AGRUPADORA_FILL)

            # Contas de detalhe
            conta_rows = []
            max_contas = max(
                len(d.get(section_key, {}).get(sub_key, {}).get("contas", []))
                for d in all_dados
            )
            for i in range(max_contas):
                desc = ""
                for dados in all_dados:
                    contas = dados.get(section_key, {}).get(sub_key, {}).get("contas", [])
                    if i < len(contas):
                        desc = contas[i].get("descricao", "")
                        break

                row_data = [f"    {desc}"]
                for dados in all_dados:
                    contas = dados.get(section_key, {}).get(sub_key, {}).get("contas", [])
                    if i < len(contas):
                        row_data.append(contas[i].get("valor", 0) or 0)
                    else:
                        row_data.append(0)
                ws.append(row_data)
                conta_rows.append(ws.max_row)
                _style_row(ws.max_row)

            # Fórmula SUM no header da subsecção
            if use_formulas and conta_rows:
                first_r, last_r = conta_rows[0], conta_rows[-1]
                for p in range(num_periods):
                    col = p + 2
                    col_letter = get_column_letter(col)
                    ws.cell(row=sub_header_row, column=col).value = f"=SUM({col_letter}{first_r}:{col_letter}{last_r})"

        # Fórmula SUM no header da seção (soma das subsecções)
        if use_formulas and sub_header_rows:
            for p in range(num_periods):
                col = p + 2
                col_letter = get_column_letter(col)
                refs = "+".join(f"{col_letter}{r}" for r in sub_header_rows)
                ws.cell(row=section_row, column=col).value = f"={refs}"

        ws.append([])

    _write_section("ATIVO", "ativo")

    # ---- PASSIVO (inclui Patrimônio Líquido) ----
    row_data = ["PASSIVO"]
    for dados in all_dados:
        if use_formulas:
            row_data.append(0)
        else:
            p_total = dados.get("passivo", {}).get("total", 0) or 0
            pl_total = dados.get("patrimonio_liquido", {}).get("total", 0) or 0
            row_data.append(p_total + pl_total)
    ws.append(row_data)
    passivo_row = ws.max_row
    section_rows["passivo"] = passivo_row
    _style_row(passivo_row, font=section_title_font)

    passivo_sub_rows = []  # rows das subsecções para fórmula do total

    for sub_key, sub_title in [("circulante", "CIRCULANTE"), ("nao_circulante", "NÃO CIRCULANTE")]:
        any_contas = any(
            d.get("passivo", {}).get(sub_key, {}).get("contas", [])
            for d in all_dados
        )
        if not any_contas:
            continue

        row_data = [sub_title]
        for dados in all_dados:
            sub = dados.get("passivo", {}).get(sub_key, {})
            row_data.append(0 if use_formulas else (sub.get("total", 0) or 0))
        ws.append(row_data)
        sub_header_row = ws.max_row
        passivo_sub_rows.append(sub_header_row)
        _style_row(sub_header_row, font=AGRUPADORA_FONT, fill=AGRUPADORA_FILL)

        conta_rows = []
        max_contas = max(
            len(d.get("passivo", {}).get(sub_key, {}).get("contas", []))
            for d in all_dados
        )
        for i in range(max_contas):
            desc = ""
            for dados in all_dados:
                contas = dados.get("passivo", {}).get(sub_key, {}).get("contas", [])
                if i < len(contas):
                    desc = contas[i].get("descricao", "")
                    break

            row_data = [f"    {desc}"]
            for dados in all_dados:
                contas = dados.get("passivo", {}).get(sub_key, {}).get("contas", [])
                if i < len(contas):
                    row_data.append(contas[i].get("valor", 0) or 0)
                else:
                    row_data.append(0)
            ws.append(row_data)
            conta_rows.append(ws.max_row)
            _style_row(ws.max_row)

        if use_formulas and conta_rows:
            first_r, last_r = conta_rows[0], conta_rows[-1]
            for p in range(num_periods):
                col = p + 2
                col_letter = get_column_letter(col)
                ws.cell(row=sub_header_row, column=col).value = f"=SUM({col_letter}{first_r}:{col_letter}{last_r})"

    # Patrimônio Líquido (subsecção do PASSIVO)
    any_pl = any(d.get("patrimonio_liquido", {}).get("contas", []) for d in all_dados)
    if any_pl:
        row_data = ["PATRIMÔNIO LÍQUIDO"]
        for dados in all_dados:
            pl = dados.get("patrimonio_liquido", {})
            row_data.append(0 if use_formulas else (pl.get("total", 0) or 0))
        ws.append(row_data)
        pl_sub_row = ws.max_row
        passivo_sub_rows.append(pl_sub_row)
        _style_row(pl_sub_row, font=AGRUPADORA_FONT, fill=AGRUPADORA_FILL)

        pl_conta_rows = []
        max_pl_contas = max(
            len(d.get("patrimonio_liquido", {}).get("contas", []))
            for d in all_dados
        )
        for i in range(max_pl_contas):
            desc = ""
            for dados in all_dados:
                contas = dados.get("patrimonio_liquido", {}).get("contas", [])
                if i < len(contas):
                    desc = contas[i].get("descricao", "")
                    break

            row_data = [f"    {desc}"]
            for dados in all_dados:
                contas = dados.get("patrimonio_liquido", {}).get("contas", [])
                if i < len(contas):
                    row_data.append(contas[i].get("valor", 0) or 0)
                else:
                    row_data.append(0)
            ws.append(row_data)
            pl_conta_rows.append(ws.max_row)
            _style_row(ws.max_row)

        if use_formulas and pl_conta_rows:
            first_r, last_r = pl_conta_rows[0], pl_conta_rows[-1]
            for p in range(num_periods):
                col = p + 2
                col_letter = get_column_letter(col)
                ws.cell(row=pl_sub_row, column=col).value = f"=SUM({col_letter}{first_r}:{col_letter}{last_r})"

    # Fórmula PASSIVO total = PC + PNC + PL
    if use_formulas and passivo_sub_rows:
        for p in range(num_periods):
            col = p + 2
            col_letter = get_column_letter(col)
            refs = "+".join(f"{col_letter}{r}" for r in passivo_sub_rows)
            ws.cell(row=passivo_row, column=col).value = f"={refs}"

    ws.append([])

    # Validação: Ativo = Passivo
    ativo_row = section_rows.get("ativo")
    ws.append(["", "VALIDAÇÃO: Ativo = Passivo"])
    ws.cell(row=ws.max_row, column=1).font = Font(name="Calibri", bold=True, size=11)

    if use_formulas and ativo_row and passivo_row:
        row_data = ["Diferença (Ativo - Passivo):"]
        for p in range(num_periods):
            row_data.append(None)
        ws.append(row_data)
        diff_row = ws.max_row
        for p in range(num_periods):
            col = p + 2
            col_letter = get_column_letter(col)
            ws.cell(row=diff_row, column=col).value = f"={col_letter}{ativo_row}-{col_letter}{passivo_row}"
            ws.cell(row=diff_row, column=col).number_format = BR_NUMBER_FORMAT
            ws.cell(row=diff_row, column=col).alignment = RIGHT_ALIGN

        row_data = ["Status:"]
        for p in range(num_periods):
            row_data.append(None)
        ws.append(row_data)
        status_row = ws.max_row
        for p in range(num_periods):
            col = p + 2
            col_letter = get_column_letter(col)
            ws.cell(row=status_row, column=col).value = (
                f'=IF(ABS({col_letter}{diff_row})<0.01,"OK","DIVERGENTE")'
            )
            ws.cell(row=status_row, column=col).font = Font(name="Calibri", bold=True, size=11)

    # Larguras
    ws.column_dimensions["A"].width = 50
    for i in range(num_periods):
        ws.column_dimensions[get_column_letter(i + 2)].width = 20

    ws.freeze_panes = f"A{header_row + 1}"


def _export_balanco_csv(dados: dict, output_path: Path) -> Path:
    with open(str(output_path), "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(["Seção", "Classificação", "Descrição", "Valor"])

        for secao, dados_secao in [
            ("Ativo", dados.get("ativo", {})),
            ("Passivo", dados.get("passivo", {})),
        ]:
            for sub_key in ("circulante", "nao_circulante"):
                sub = dados_secao.get(sub_key, {})
                for conta in sub.get("contas", []):
                    writer.writerow([
                        secao,
                        conta.get("classificacao", ""),
                        conta.get("descricao", ""),
                        conta.get("valor", 0),
                    ])

        pl = dados.get("patrimonio_liquido", {})
        for conta in pl.get("contas", []):
            writer.writerow([
                "PL",
                conta.get("classificacao", ""),
                conta.get("descricao", ""),
                conta.get("valor", 0),
            ])

    logger.info("CSV balanço gerado: %s", output_path)
    return output_path


# ---------------------------------------------------------------------------
# Exportação bruta (sem formatação IA)
# ---------------------------------------------------------------------------

def _parse_pipe_table(raw_text: str) -> list[list[str]]:
    """Parseia texto pipe-separated da extração em lista de linhas/colunas."""
    rows = []
    for line in raw_text.splitlines():
        line = line.strip()
        if not line:
            continue
        # Pula linhas separadoras (---|---|---)
        if re.match(r'^[\s|:-]+$', line):
            continue
        if '|' in line:
            cells = [c.strip() for c in line.split('|')]
            # Remove células vazias das bordas (|col1|col2| → ['', 'col1', 'col2', ''])
            if cells and cells[0] == '':
                cells = cells[1:]
            if cells and cells[-1] == '':
                cells = cells[:-1]
            rows.append(cells)
        else:
            rows.append([line])
    return rows


def export_raw_csv(raw_text: str, output_path: Path) -> Path:
    """Exporta texto bruto pipe-separated como CSV."""
    rows = _parse_pipe_table(raw_text)

    with open(str(output_path), "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        for row in rows:
            writer.writerow(row)

    logger.info("CSV bruto gerado: %s (%d linhas)", output_path, len(rows))
    return output_path


def export_raw_excel(
    demonstracoes: list[dict],
    empresa: str,
    output_path: Path,
) -> Path:
    """Exporta textos brutos pipe-separated como Excel multi-aba."""
    wb = Workbook()
    default_ws = wb.active

    for i, demo in enumerate(demonstracoes):
        tipo = demo["tipo"]
        periodo = demo.get("periodo", "")
        raw_text = demo.get("raw_text", "")

        tab_name = _sanitize_tab_name(_build_title(empresa, tipo, periodo))

        if i == 0:
            default_ws.title = tab_name
            ws = default_ws
        else:
            ws = wb.create_sheet(title=tab_name)

        rows = _parse_pipe_table(raw_text)
        if not rows:
            continue

        # Header
        titulo = _build_title(empresa, tipo, periodo)
        ws.append([titulo])
        num_cols = max(len(r) for r in rows)
        if num_cols > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        ws["A1"].font = Font(name="Calibri", bold=True, size=14)
        ws.append([])

        # Primeira linha da tabela como cabeçalho
        header_row_num = ws.max_row + 1
        ws.append(rows[0])
        for col_idx in range(len(rows[0])):
            cell = ws.cell(row=header_row_num, column=col_idx + 1)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER

        # Dados
        for row in rows[1:]:
            ws.append(row)
            current_row = ws.max_row
            for col_idx in range(len(row)):
                cell = ws.cell(row=current_row, column=col_idx + 1)
                cell.border = THIN_BORDER
                cell.font = NORMAL_FONT

        # Auto-width
        for col_idx in range(num_cols):
            max_len = 0
            col_letter = get_column_letter(col_idx + 1)
            for row in rows:
                if col_idx < len(row):
                    max_len = max(max_len, len(str(row[col_idx])))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

        ws.freeze_panes = f"A{header_row_num + 1}"

    wb.save(str(output_path))
    logger.info("Excel bruto gerado: %s (%d aba(s))", output_path, len(demonstracoes))
    return output_path
