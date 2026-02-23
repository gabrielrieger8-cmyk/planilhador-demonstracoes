#!/usr/bin/env python
"""Controladoria Plus CLI — Conversor de Balancetes via terminal com Rich.

Uso:
    python cli.py                                    # modo interativo
    python cli.py balancete.pdf                      # modo direto
    python cli.py data/input/                        # pasta inteira
    python cli.py *.pdf --model gemini-2.5-flash     # com opções
    python cli.py balancete.pdf --xlsx --workers 4   # XLSX profissional
"""

from __future__ import annotations

import argparse
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

# --- Bootstrap (DEVE ser primeiro) ---
from controladoria_core.utils.config import configure

configure(project_root=Path(__file__).parent)

# --- Core imports ---
from controladoria_core.orchestrator import Orchestrator, OutputFormat, ProcessingResult
from controladoria_core.exporters.sign_logic import SignConfig
from controladoria_core.exporters.xlsx_builder import BalanceteXlsxBuilder, detect_periodo
from controladoria_core.exporters.reference_extractor import (
    extract_reference_from_xlsx,
    list_references,
    load_reference_for_prompt,
    save_reference,
)
from controladoria_core.agents.gemini_agent import GeminiAgent
from controladoria_core.parsers.csv_parser import save_as_csv
from controladoria_core.utils.config import (
    INPUT_DIR,
    OUTPUT_DIR,
    KNOWLEDGE_DIR,
    PROJECT_ROOT,
    MODELOS_DISPONIVEIS,
    GEMINI_API_KEY,
    config,
    logger,
)

# --- Rich imports ---
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.progress import (
    BarColumn,
    MofNCompleteColumn,
    Progress,
    SpinnerColumn,
    TextColumn,
    TimeElapsedColumn,
)
from rich.prompt import Confirm, IntPrompt, Prompt
from rich.text import Text
from rich.theme import Theme

# --- Questionary (menus com setas) ---
import questionary
from prompt_toolkit.key_binding import KeyBindings, merge_key_bindings


# ===================================================================
# THEME SYSTEM
# ===================================================================

_STYLES = {
    "escuro": {
        "heading": "bold blue",
        "accent": "bold cyan",
        "num": "cyan",
        "border": "blue",
        "ok": "green",
        "ok_bold": "bold green",
        "warn": "yellow",
        "err": "red",
        "err_bold": "bold red",
        "agrup": "bold blue",
    },
    "claro": {
        "heading": "bold dark_blue",
        "accent": "bold dark_cyan",
        "num": "dark_cyan",
        "border": "dark_blue",
        "ok": "dark_green",
        "ok_bold": "bold dark_green",
        "warn": "dark_orange3",
        "err": "red",
        "err_bold": "bold red",
        "agrup": "bold dark_blue",
    },
}

current_theme: str = "escuro"
console = Console(theme=Theme(_STYLES[current_theme]))

_QSTYLES = {
    "escuro": questionary.Style([
        ("qmark", "fg:cyan bold"),
        ("question", "bold"),
        ("pointer", "fg:cyan bold"),
        ("highlighted", "fg:cyan bold"),
        ("selected", "fg:green bold"),
        ("separator", "fg:#6c6c6c"),
        ("instruction", "fg:#6c6c6c"),
    ]),
    "claro": questionary.Style([
        ("qmark", "fg:darkblue bold"),
        ("question", "bold"),
        ("pointer", "fg:darkblue bold"),
        ("highlighted", "fg:darkblue bold"),
        ("selected", "fg:darkgreen bold"),
        ("separator", "fg:#6c6c6c"),
        ("instruction", "fg:#6c6c6c"),
    ]),
}


def ts(name: str) -> str:
    """Retorna string de estilo do tema atual (para kwargs como border_style)."""
    return _STYLES[current_theme][name]


def qstyle() -> questionary.Style:
    """Retorna estilo questionary do tema atual."""
    return _QSTYLES[current_theme]


def switch_theme() -> None:
    """Alterna entre tema escuro e claro."""
    global console, current_theme
    current_theme = "claro" if current_theme == "escuro" else "escuro"
    console = Console(theme=Theme(_STYLES[current_theme]))


def open_folder(path: Path) -> None:
    """Abre pasta no explorador de arquivos."""
    path.mkdir(parents=True, exist_ok=True)
    os.startfile(str(path))
    console.print(f"  [ok]Pasta aberta:[/ok] {path}")


def _ask(question: questionary.Question):
    """Executa questionary Question com suporte a ESC para voltar."""
    kb = KeyBindings()

    @kb.add("escape")
    def _(event):
        event.app.exit(exception=KeyboardInterrupt())

    app = question.application
    app.key_bindings = merge_key_bindings([app.key_bindings, kb])

    try:
        return app.run()
    except KeyboardInterrupt:
        return None


# ===================================================================
# ARGUMENT PARSING
# ===================================================================


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="cli",
        description="Controladoria Plus — Conversor de Balancetes via terminal",
    )
    parser.add_argument(
        "files", nargs="*", help="PDFs ou diretório para processar",
    )
    parser.add_argument(
        "--model", "-m",
        choices=list(MODELOS_DISPONIVEIS.keys()),
        help="Modelo Gemini a usar",
    )
    parser.add_argument(
        "--xlsx", action="store_true",
        help="Gerar XLSX Profissional após conversão",
    )
    parser.add_argument(
        "--workers", "-w", type=int, default=1,
        help="Número de workers paralelos (default: 1)",
    )
    parser.add_argument(
        "--sign-mode", choices=["auto", "skip"], default="auto",
        help="Modo de sinais D/C (default: auto)",
    )
    parser.add_argument(
        "--reference", "-r",
        help="Nome da referência RAG a usar",
    )
    parser.add_argument(
        "--output-dir", "-o", type=Path,
        help="Diretório de saída (default: data/output/)",
    )
    parser.add_argument(
        "--detail-level",
        choices=["completo", "agrupadoras"],
        default="completo",
        help="Nível de detalhe do XLSX (default: completo)",
    )
    return parser


# ===================================================================
# DIRECT MODE
# ===================================================================


def direct_mode(args: argparse.Namespace) -> None:
    """Processa arquivos diretamente com Rich progress bars."""
    if args.model:
        config.gemini_model = args.model

    pdfs = resolve_files(args.files)
    if not pdfs:
        console.print("[err_bold]Nenhum PDF válido encontrado.[/err_bold]")
        sys.exit(1)

    out_dir = Path(args.output_dir) if args.output_dir else OUTPUT_DIR

    show_header_panel(pdfs)
    results = process_with_progress(
        pdfs, out_dir,
        workers=args.workers,
        reference_name=args.reference,
    )
    show_results_table(results)

    if args.xlsx:
        generate_xlsx_from_results(
            results, out_dir,
            sign_mode=args.sign_mode,
            detail_level=args.detail_level,
        )


# ===================================================================
# INTERACTIVE MODE
# ===================================================================


def interactive_mode() -> None:
    """Interface interativa com menus e navegação por setas."""
    console.print()
    console.print(Panel(
        "[heading]Controladoria Plus CLI[/heading]\n"
        "Conversor de Balancetes com interface Rich",
        title="Bem-vindo",
        border_style=ts("border"),
    ))

    while True:
        console.print()
        tema_atual = "Escuro" if current_theme == "escuro" else "Claro"
        tema_novo = "Claro" if current_theme == "escuro" else "Escuro"

        choice = _ask(questionary.select(
            "Menu Principal",
            choices=[
                "Processar PDFs",
                "Gerar XLSX Profissional (de CSV existente)",
                f"Selecionar modelo (atual: {config.gemini_model})",
                "Gerenciar referências",
                "Listar arquivos em input/",
                questionary.Separator("─" * 40),
                "Abrir pasta input/",
                "Abrir pasta output/",
                "Abrir pasta referências",
                "Abrir raiz do projeto",
                questionary.Separator("─" * 40),
                f"Tema: {tema_atual} → {tema_novo}",
                "Sair",
            ],
            style=qstyle(),
            instruction="(↑↓ navegar, Enter selecionar, Esc voltar)",
        ))

        if choice is None or choice == "Sair":
            console.print("\n  [dim]Até logo![/dim]\n")
            break
        elif choice == "Processar PDFs":
            menu_process_pdfs()
        elif choice.startswith("Gerar XLSX"):
            menu_generate_xlsx()
        elif choice.startswith("Selecionar modelo"):
            menu_select_model()
        elif choice == "Gerenciar referências":
            menu_references()
        elif choice.startswith("Listar"):
            menu_list_input()
        elif choice == "Abrir pasta input/":
            open_folder(INPUT_DIR)
        elif choice == "Abrir pasta output/":
            open_folder(OUTPUT_DIR)
        elif choice == "Abrir pasta referências":
            open_folder(KNOWLEDGE_DIR)
        elif choice == "Abrir raiz do projeto":
            open_folder(PROJECT_ROOT)
        elif choice.startswith("Tema:"):
            switch_theme()
            console.print(f"\n  [ok]Tema alterado para: {current_theme.title()}[/ok]")


def menu_process_pdfs() -> None:
    """Submenu interativo de processamento de PDFs."""
    pdfs = list(INPUT_DIR.glob("*.pdf"))
    if not pdfs:
        console.print(f"\n  [warn]Nenhum PDF encontrado em {INPUT_DIR}[/warn]")
        console.print("  Coloque PDFs nessa pasta e tente novamente.\n")
        return

    show_header_panel(pdfs)

    if not Confirm.ask("  Processar todos?", default=True):
        return

    workers = IntPrompt.ask("  Workers paralelos", default=1)
    ref = select_reference_interactive()

    results = process_with_progress(
        pdfs, OUTPUT_DIR, workers=workers, reference_name=ref,
    )
    show_results_table(results)

    # Preview do primeiro resultado + loop de correção
    if results and results[0].success:
        preview_rows = results[0].details.get("preview_rows")
        if preview_rows:
            base_name = Path(results[0].file_path).stem
            pdf_path = Path(results[0].file_path)
            show_preview(preview_rows, base_name)

            # Correção pós-conversão
            corrected_rows, final_version = menu_correction(
                pdf_path=pdf_path,
                base_name=base_name,
                preview_rows=preview_rows,
                reference_name=ref,
            )
            if final_version > 1:
                results[0].details["preview_rows"] = corrected_rows

    if Confirm.ask("\n  Gerar XLSX Profissional?", default=False):
        sign_mode = Prompt.ask(
            "  Modo de sinais",
            choices=["auto", "skip"],
            default="auto",
        )
        generate_xlsx_from_results(results, OUTPUT_DIR, sign_mode=sign_mode)

        # Oferecer salvar como referência
        xlsx_path = OUTPUT_DIR / "Balancetes_Profissional.xlsx"
        if xlsx_path.exists():
            menu_save_reference(xlsx_path)


def menu_correction(
    pdf_path: Path,
    base_name: str,
    preview_rows: list[list[str]],
    reference_name: str | None,
    version: int = 1,
) -> tuple[list[list[str]], int]:
    """Loop de correção pós-conversão. Reprocessa o PDF com prompt de correção.

    Returns:
        Tupla (preview_rows atualizados, versão final).
    """
    current_rows = preview_rows
    current_version = version

    while True:
        wants = _ask(questionary.confirm(
            "Deseja corrigir o resultado?",
            default=False,
            style=qstyle(),
        ))

        if not wants:
            break

        correction = Prompt.ask("  Descreva a correção necessária").strip()
        if not correction:
            console.print("  [warn]Correção vazia, cancelando.[/warn]")
            continue

        current_version += 1

        correction_prompt = (
            f"ATENÇÃO — CORREÇÃO SOLICITADA PELO USUÁRIO:\n"
            f"{correction}\n\n"
            f"Reprocesse o PDF aplicando a correção acima. "
            f"Mantenha todas as outras contas como estavam."
        )

        console.print()
        with console.status(
            f"[bold]Reprocessando v{current_version}...[/bold]",
            spinner="dots",
        ):
            try:
                agent = GeminiAgent()
                result = agent.process(
                    str(pdf_path),
                    prompt=correction_prompt,
                    financial=True,
                    reference_name=reference_name,
                )
            except Exception as exc:
                console.print(f"  [err]Erro na correção: {exc}[/err]")
                current_version -= 1
                continue

        if not result.success or not result.text:
            console.print(f"  [err]Erro: {result.error or 'sem resposta'}[/err]")
            current_version -= 1
            continue

        versioned_name = f"{base_name}_v{current_version}"
        csv_paths, unified_rows = save_as_csv(
            result.text, versioned_name, output_dir=OUTPUT_DIR,
        )

        if not unified_rows or len(unified_rows) < 2:
            console.print("  [err]Resultado vazio após correção.[/err]")
            current_version -= 1
            continue

        current_rows = unified_rows

        console.print(
            f"\n  [ok_bold]Correção v{current_version} aplicada![/ok_bold] "
            f"({len(unified_rows) - 1} linhas)"
        )
        show_preview(unified_rows, f"{base_name} v{current_version}")

    return current_rows, current_version


def menu_generate_xlsx() -> None:
    """Gera XLSX Profissional a partir de CSVs existentes no output."""
    csvs = list(OUTPUT_DIR.glob("*.csv"))
    # Filtra csvs que não sejam _sintetico ou _sinal
    csvs = [c for c in csvs if "_sintetico" not in c.name and "_sinal" not in c.name]

    if not csvs:
        console.print(f"\n  [warn]Nenhum CSV encontrado em {OUTPUT_DIR}[/warn]")
        return

    # Seleção com setas
    csv_names = [c.name for c in csvs]
    all_choices = csv_names + [questionary.Separator(), "Todos"]

    result = _ask(questionary.select(
        "Selecione CSV para gerar XLSX",
        choices=all_choices,
        style=qstyle(),
    ))

    if result is None:
        return
    elif result == "Todos":
        selected = csvs
    else:
        idx = csv_names.index(result)
        selected = [csvs[idx]]

    sign_mode = Prompt.ask(
        "  Modo de sinais",
        choices=["auto", "skip"],
        default="auto",
    )

    import csv as csv_mod

    existing_wb = None
    for csv_file in selected:
        # Lê CSV
        with open(csv_file, "r", encoding="utf-8-sig") as f:
            reader = csv_mod.reader(f, delimiter=";")
            rows = [row for row in reader]

        if len(rows) < 2:
            console.print(f"  [warn]CSV vazio: {csv_file.name}[/warn]")
            continue

        periodo = detect_periodo(csv_file.name)
        sign_cfg = SignConfig(mode=sign_mode)

        builder = BalanceteXlsxBuilder(
            unified_rows=rows,
            periodo=periodo,
            filename=csv_file.stem,
            sign_config=sign_cfg,
        )

        if sign_mode == "auto":
            sign_result = builder.detect_signs()
            if sign_result.has_dc and sign_result.matches_convention:
                builder.apply_signs(SignConfig(mode="auto"))

        out_path = OUTPUT_DIR / "Balancetes_Profissional.xlsx"
        builder.build(
            output_path=out_path,
            existing_workbook=existing_wb,
        )
        existing_wb = out_path
        console.print(f"  [ok]✓[/ok] {csv_file.name} → aba {periodo}")

    if existing_wb:
        console.print(f"\n  [ok_bold]XLSX gerado: {existing_wb}[/ok_bold]")


def menu_select_model() -> None:
    """Seleção interativa de modelo com setas."""
    models = list(MODELOS_DISPONIVEIS.items())
    choices = []

    for model_id, info in models:
        label = (
            f"{info['label']} — {model_id}  "
            f"(in: ${info['input_price']:.2f}, out: ${info['output_price']:.2f}/1M)"
        )
        if model_id == config.gemini_model:
            label += "  ◀ atual"
        choices.append(questionary.Choice(label, value=model_id))

    result = _ask(questionary.select(
        "Selecionar modelo",
        choices=choices,
        style=qstyle(),
    ))

    if result:
        config.gemini_model = result
        console.print(f"\n  [ok]Modelo alterado para: {result}[/ok]")


def menu_references() -> None:
    """Submenu de gerenciamento de referências."""
    while True:
        console.print()
        choice = _ask(questionary.select(
            "Gerenciar Referências",
            choices=[
                "Listar referências",
                "Importar referência de XLSX",
                "Chat com referência",
                "Excluir referência",
                questionary.Separator("─" * 30),
                "Voltar",
            ],
            style=qstyle(),
            instruction="(↑↓ navegar, Enter selecionar, Esc voltar)",
        ))

        if choice is None or choice == "Voltar":
            break
        elif choice == "Listar referências":
            _show_references_table()
        elif choice == "Importar referência de XLSX":
            menu_upload_reference()
        elif choice == "Chat com referência":
            menu_chat_reference()
        elif choice == "Excluir referência":
            menu_delete_reference()


def _show_references_table() -> None:
    """Lista referências em tabela Rich."""
    refs = list_references()
    if not refs:
        console.print("\n  [warn]Nenhuma referência disponível.[/warn]")
        return

    console.print()
    table = Table(title="Referências Disponíveis", title_style=ts("heading"))
    table.add_column("#", style=ts("num"), width=3)
    table.add_column("Nome")
    table.add_column("Empresa")
    table.add_column("Período")
    table.add_column("Contas", justify="right")
    table.add_column("Criada em")

    for i, ref in enumerate(refs, 1):
        table.add_row(
            str(i),
            ref.get("display_name", ref.get("filename", "")),
            ref.get("empresa", ""),
            ref.get("periodo", ""),
            str(ref.get("total_contas", 0)),
            ref.get("created_at", "")[:10],
        )
    console.print(table)


def _extract_and_save_reference(xlsx_path: Path) -> None:
    """Fluxo compartilhado: extrai padrão de um XLSX e salva como referência."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(str(xlsx_path), read_only=True)
        sheets = [s for s in wb.sheetnames if not s.startswith("_")]
        wb.close()
    except Exception as exc:
        console.print(f"  [err]Erro ao ler XLSX: {exc}[/err]")
        return

    if not sheets:
        console.print("  [warn]Nenhuma aba encontrada no XLSX.[/warn]")
        return

    if len(sheets) == 1:
        sheet_name = sheets[0]
        console.print(f"  Aba selecionada: [accent]{sheet_name}[/accent]")
    else:
        sheet_name = _ask(questionary.select(
            "Selecione a aba de referência",
            choices=sheets,
            style=qstyle(),
        ))
        if not sheet_name:
            return

    ref_name = Prompt.ask("  Nome da referência", default="").strip()

    console.print("  [dim]Instruções: correções ou observações para a IA (opcional).[/dim]")
    instructions = Prompt.ask("  Instruções", default="").strip()

    with console.status("[bold]Extraindo referência...[/bold]", spinner="dots"):
        try:
            ref = extract_reference_from_xlsx(xlsx_path, sheet_name=sheet_name)
            txt_path, json_path = save_reference(
                ref, user_instructions=instructions, name=ref_name,
            )
        except Exception as exc:
            console.print(f"\n  [err]Erro ao extrair referência: {exc}[/err]")
            return

    summary = Text()
    summary.append("Nome:        ", style="dim")
    summary.append(f"{ref_name or ref.empresa}\n", style="bold")
    summary.append("Empresa:     ", style="dim")
    summary.append(f"{ref.empresa}\n")
    summary.append("Período:     ", style="dim")
    summary.append(f"{ref.periodo}\n")
    summary.append("Contas:      ", style="dim")
    summary.append(f"{ref.total_contas}\n")
    summary.append("Grupos:      ", style="dim")
    summary.append(f"{len(ref.grupos)}\n")
    summary.append("Hierarquia:  ", style="dim")
    summary.append(f"{len(ref.hierarchy_tree)} nós\n")
    summary.append("Sinais:      ", style="dim")
    summary.append(f"{len(ref.sign_examples)} exemplos\n")
    if instructions:
        summary.append("Instruções:  ", style="dim")
        summary.append("Sim\n", style="ok")
    summary.append(f"\nArquivos: {txt_path.name}, {json_path.name}", style="dim")

    console.print()
    console.print(Panel(summary, title="Referência Salva", border_style=ts("ok")))


def menu_upload_reference() -> None:
    """Importa referência de um XLSX externo."""
    xlsx_input = Prompt.ask("  Caminho do arquivo XLSX").strip().strip('"').strip("'")
    if not xlsx_input:
        return

    xlsx_path = Path(xlsx_input)
    if not xlsx_path.exists():
        console.print(f"  [err]Arquivo não encontrado: {xlsx_path}[/err]")
        return

    if xlsx_path.suffix.lower() not in (".xlsx", ".xls"):
        console.print("  [err]Apenas arquivos .xlsx são aceitos.[/err]")
        return

    _extract_and_save_reference(xlsx_path)


def menu_save_reference(xlsx_path: Path) -> None:
    """Oferece salvar XLSX gerado como referência."""
    wants = _ask(questionary.confirm(
        "Salvar como referência para futuras conversões?",
        default=False,
        style=qstyle(),
    ))
    if not wants:
        return
    _extract_and_save_reference(xlsx_path)


def menu_delete_reference() -> None:
    """Exclui uma referência do knowledge/."""
    refs = list_references()
    if not refs:
        console.print("\n  [warn]Nenhuma referência disponível.[/warn]")
        return

    choices = []
    for ref in refs:
        name = ref.get("display_name", ref.get("filename", ""))
        choices.append(questionary.Choice(name, value=ref.get("filename", "")))
    choices.append(questionary.Separator())
    choices.append(questionary.Choice("Cancelar", value=""))

    filename = _ask(questionary.select(
        "Selecione referência para excluir",
        choices=choices,
        style=qstyle(),
    ))

    if not filename:
        return

    if not Confirm.ask(f"  Excluir '{filename}'?", default=False):
        return

    for ext in (".txt", ".json"):
        fpath = KNOWLEDGE_DIR / f"{filename}{ext}"
        if fpath.exists():
            fpath.unlink()

    console.print(f"  [ok]Referência '{filename}' excluída.[/ok]")


def menu_chat_reference() -> None:
    """Chat conversacional com IA para ajustar referências."""
    refs = list_references()
    if not refs:
        console.print("\n  [warn]Nenhuma referência disponível.[/warn]")
        return

    # Selecionar referência
    choices = []
    for ref in refs:
        name = ref.get("display_name", ref.get("filename", ""))
        empresa = ref.get("empresa", "")
        label = f"{name} ({empresa})" if empresa else name
        choices.append(questionary.Choice(label, value=ref.get("filename", "")))

    ref_filename = _ask(questionary.select(
        "Selecione a referência",
        choices=choices,
        style=qstyle(),
    ))
    if not ref_filename:
        return

    ref_text = load_reference_for_prompt(reference_name=ref_filename)
    if not ref_text:
        console.print(f"  [err]Referência '{ref_filename}' não encontrada.[/err]")
        return

    # Selecionar modelo
    model_choices = []
    for model_id, info in MODELOS_DISPONIVEIS.items():
        label = (
            f"{info['label']} — "
            f"(in: ${info['input_price']:.2f}, out: ${info['output_price']:.2f}/1M)"
        )
        model_choices.append(questionary.Choice(label, value=model_id))

    model_id = _ask(questionary.select(
        "Modelo para o chat",
        choices=model_choices,
        style=qstyle(),
    ))
    if not model_id:
        return

    # System prompt (mesmo do web /chat-reference)
    system_prompt = (
        "Você é um assistente especializado em contabilidade brasileira. "
        "O usuário tem uma referência de balancete (plano de contas, hierarquia, sinais D/C). "
        "Sua tarefa é ajudar a ajustar essa referência com base nas instruções do usuário.\n\n"
        "REFERÊNCIA ATUAL:\n"
        f"{ref_text}\n\n"
        "INSTRUÇÕES:\n"
        "- Analise a referência acima e responda a pergunta/instrução do usuário.\n"
        "- Se o usuário pedir uma alteração, descreva EXATAMENTE o que deve mudar.\n"
        "- Se precisar atualizar a referência, responda com a seção modificada.\n"
        "- Responda em português.\n"
    )

    console.print()
    console.print(Panel(
        f"Chat com referência: [bold]{ref_filename}[/bold]\n"
        f"Modelo: {model_id}\n"
        "[dim]Digite 'sair' para encerrar o chat.[/dim]",
        border_style=ts("border"),
    ))

    history: list[dict[str, str]] = []

    from google import genai
    client = genai.Client(api_key=GEMINI_API_KEY)

    while True:
        console.print()
        message = Prompt.ask("[bold]Você[/bold]")

        if not message or message.strip().lower() in ("sair", "exit", "quit"):
            console.print("  [dim]Chat encerrado.[/dim]")
            break

        history.append({"role": "user", "content": message})

        conversation_parts = [system_prompt]
        for msg in history[:-1]:
            role_label = "Usuário" if msg["role"] == "user" else "Assistente"
            conversation_parts.append(f"{role_label}: {msg['content']}")
        conversation_parts.append(f"Usuário: {message}")

        full_prompt = "\n\n".join(conversation_parts)

        with console.status("[bold]Pensando...[/bold]", spinner="dots"):
            try:
                response = client.models.generate_content(
                    model=model_id,
                    contents=[full_prompt],
                    config={
                        "temperature": 0.3,
                        "max_output_tokens": 8192,
                    },
                )
                ai_response = response.text or "Sem resposta."
            except Exception as exc:
                console.print(f"  [err]Erro: {exc}[/err]")
                history.pop()
                continue

        history.append({"role": "assistant", "content": ai_response})

        console.print()
        console.print(Panel(
            ai_response,
            title="Assistente",
            border_style=ts("accent"),
            padding=(1, 2),
        ))


def menu_list_input() -> None:
    """Lista PDFs no diretório de input."""
    pdfs = list(INPUT_DIR.glob("*.pdf"))
    if not pdfs:
        console.print(f"\n  [warn]Nenhum PDF em {INPUT_DIR}[/warn]")
        return

    console.print(f"\n  [accent]PDFs em {INPUT_DIR}:[/accent]")
    for i, pdf in enumerate(pdfs, 1):
        size_mb = pdf.stat().st_size / (1024 * 1024)
        console.print(f"    [num]{i}[/num]. {pdf.name} [dim]({size_mb:.1f} MB)[/dim]")
    console.print(f"\n  Total: {len(pdfs)} arquivo(s)")


# ===================================================================
# HELPERS
# ===================================================================


def resolve_files(file_args: list[str]) -> list[Path]:
    """Resolve argumentos de arquivo/diretório para lista de PDFs."""
    pdfs: list[Path] = []
    for arg in file_args:
        p = Path(arg)
        if not p.exists() and not p.is_absolute():
            p = INPUT_DIR / p
        if p.is_dir():
            pdfs.extend(sorted(p.glob("*.pdf")))
        elif p.exists() and p.suffix.lower() == ".pdf":
            pdfs.append(p)
        else:
            console.print(f"  [warn]Ignorando: {arg}[/warn]")
    return pdfs


def select_reference_interactive() -> str | None:
    """Pede ao usuário para selecionar referência com setas."""
    refs = list_references()
    if not refs:
        return None

    choices = [questionary.Choice("Nenhuma", value="")]
    for ref in refs:
        name = ref.get("display_name", ref.get("filename", ""))
        choices.append(questionary.Choice(name, value=ref.get("filename", "")))

    result = _ask(questionary.select(
        "Referência RAG",
        choices=choices,
        style=qstyle(),
    ))

    if result:
        console.print(f"  [ok]Usando referência: {result}[/ok]")
        return result
    return None


def show_header_panel(pdfs: list[Path]) -> None:
    """Exibe painel com info do job."""
    import fitz

    total_pages = 0
    for pdf in pdfs:
        try:
            doc = fitz.open(str(pdf))
            total_pages += len(doc)
            doc.close()
        except Exception:
            pass

    info = Text()
    info.append("Modelo:      ", style="dim")
    info.append(f"{config.gemini_model}\n", style="bold")
    info.append("Arquivos:    ", style="dim")
    info.append(f"{len(pdfs)} PDF(s)", style="bold")
    if total_pages:
        info.append(f" ({total_pages} páginas)", style="dim")

    console.print()
    console.print(Panel(info, title="Controladoria Plus CLI", border_style=ts("border")))


def process_with_progress(
    pdfs: list[Path],
    output_dir: Path,
    workers: int = 1,
    reference_name: str | None = None,
) -> list[ProcessingResult]:
    """Processa PDFs com Rich progress bars."""
    results: list[ProcessingResult] = []
    orch = Orchestrator()

    console.print()

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        MofNCompleteColumn(),
        TimeElapsedColumn(),
        console=console,
    ) as progress:
        overall = progress.add_task(
            "[bold]Total", total=len(pdfs),
        )

        if workers <= 1:
            # Processamento sequencial
            for pdf in pdfs:
                task = progress.add_task(
                    f"  {pdf.name}", total=1,
                )
                result = orch.process(
                    pdf,
                    output_format=OutputFormat.CSV,
                    output_dir=output_dir,
                    reference_name=reference_name,
                )
                results.append(result)
                progress.update(task, completed=1)
                progress.update(overall, advance=1)
        else:
            # Processamento paralelo
            tasks_map: dict[str, int] = {}
            for pdf in pdfs:
                tid = progress.add_task(f"  {pdf.name}", total=1)
                tasks_map[str(pdf)] = tid

            with ThreadPoolExecutor(max_workers=workers) as executor:
                futures = {}
                for i, pdf in enumerate(pdfs):
                    if i > 0:
                        time.sleep(2)  # stagger delay
                    future = executor.submit(
                        orch.process,
                        pdf,
                        OutputFormat.CSV,
                        output_dir,
                        reference_name=reference_name,
                    )
                    futures[future] = pdf

                for future in as_completed(futures):
                    pdf = futures[future]
                    result = future.result()
                    results.append(result)
                    tid = tasks_map[str(pdf)]
                    progress.update(tid, completed=1)
                    progress.update(overall, advance=1)

    return results


def show_results_table(results: list[ProcessingResult]) -> None:
    """Exibe resultados em tabela Rich."""
    console.print()
    table = Table(title="Resultados", title_style=ts("heading"))
    table.add_column("Arquivo")
    table.add_column("Status")
    table.add_column("Tempo", justify="right")
    table.add_column("Custo", justify="right")
    table.add_column("Saída")

    total_time = 0.0
    total_cost = 0.0
    ok_count = 0

    for r in results:
        name = Path(r.file_path).name
        total_time += r.processing_time
        total_cost += r.estimated_cost

        if r.success:
            ok_count += 1
            status = "[ok]OK[/ok]"
            saida = ", ".join(Path(f).name for f in r.output_files) if r.output_files else "-"
        else:
            status = "[err]ERRO[/err]"
            saida = r.error or "-"

        table.add_row(
            name,
            status,
            f"{r.processing_time:.1f}s",
            f"${r.estimated_cost:.4f}",
            saida,
        )

    console.print(table)

    # Resumo
    err_count = len(results) - ok_count
    summary = (
        f"  Total: {len(results)} arquivo(s) | "
        f"[ok]OK: {ok_count}[/ok]"
    )
    if err_count:
        summary += f" | [err]Erros: {err_count}[/err]"
    summary += f" | Tempo: {total_time:.1f}s | Custo: ${total_cost:.4f}"
    console.print(summary)


def show_preview(
    rows: list[list[str]], title: str, max_rows: int = 15,
) -> None:
    """Exibe preview dos dados em tabela Rich."""
    if not rows or len(rows) < 2:
        return

    console.print()
    table = Table(
        title=f"Preview: {title}",
        show_lines=True,
        title_style=ts("heading"),
    )

    header = rows[0]
    for col in header:
        table.add_column(col, overflow="fold")

    for row in rows[1 : max_rows + 1]:
        # Detecta agrupadora para estilizar
        tipo_idx = None
        for i, h in enumerate(header):
            if h.strip().lower() == "tipo":
                tipo_idx = i
                break

        is_agrup = (
            tipo_idx is not None
            and tipo_idx < len(row)
            and row[tipo_idx].strip().upper() == "A"
        )

        style = ts("agrup") if is_agrup else None
        table.add_row(*[str(c) for c in row], style=style)

    remaining = len(rows) - max_rows - 1
    if remaining > 0:
        filler = [f"... +{remaining} linhas" if i == 0 else "" for i in range(len(header))]
        table.add_row(*filler, style="dim")

    console.print(table)


def generate_xlsx_from_results(
    results: list[ProcessingResult],
    output_dir: Path,
    sign_mode: str = "auto",
    detail_level: str = "completo",
) -> None:
    """Gera XLSX Profissional consolidado a partir dos resultados."""
    sign_cfg = SignConfig(mode=sign_mode)
    existing_wb = None
    out_path = output_dir / "Balancetes_Profissional.xlsx"
    generated = 0

    console.print()
    for r in results:
        if not r.success:
            continue

        preview_rows = r.details.get("preview_rows")
        if not preview_rows or len(preview_rows) < 2:
            continue

        filename = Path(r.file_path).stem
        periodo = detect_periodo(filename)

        builder = BalanceteXlsxBuilder(
            unified_rows=preview_rows,
            periodo=periodo,
            filename=filename,
            sign_config=sign_cfg,
        )

        if sign_mode == "auto":
            sign_result = builder.detect_signs()
            if sign_result.has_dc and sign_result.matches_convention:
                builder.apply_signs(SignConfig(mode="auto"))

        if detail_level != "completo":
            builder.filter_rows(detail_level=detail_level)

        builder.build(
            output_path=out_path,
            existing_workbook=existing_wb,
        )
        existing_wb = out_path
        generated += 1
        console.print(f"  [ok]✓[/ok] {filename} → aba {periodo}")

    if generated:
        console.print(
            f"\n  [ok_bold]XLSX Profissional gerado: {out_path}[/ok_bold]"
        )
    else:
        console.print("  [warn]Nenhum resultado válido para gerar XLSX.[/warn]")


# ===================================================================
# MAIN
# ===================================================================


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    if args.files:
        direct_mode(args)
    else:
        interactive_mode()


if __name__ == "__main__":
    main()
