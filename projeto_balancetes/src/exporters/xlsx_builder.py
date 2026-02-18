"""Gerador de Excel profissional para balancetes.

Cria planilha Excel com:
- Colunas reordenadas (sem Tipo visível, formatação condicional)
- Excel Tables nomeadas Tab_MM_AAAA
- Números em formato brasileiro (float + custom format)
- Fórmulas SUM de validação para agrupadoras
- Integração com sign_logic para conversão D/C → +/-
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.comments import Comment

from src.exporters.hierarchy import build_hierarchy, get_account_group
from src.exporters.sign_logic import (
    SignConfig,
    SignDetectionResult,
    apply_sign_convention,
    detect_sign_mode,
)
from src.utils.config import logger


# ---------------------------------------------------------------------------
# Constantes de layout
# ---------------------------------------------------------------------------

# Ordem das colunas visíveis no Excel de saída
# (o Tipo fica como última coluna, oculta)
VISIBLE_COLUMNS = [
    "Código",
    "Classificação",
    "Descrição",
    "Saldo Anterior",
    "Natureza SA",
    "Débito",
    "Crédito",
    "Saldo Atual",
    "Natureza SAT",
]

# Coluna oculta (última)
HIDDEN_COL_TIPO = "Tipo"

# Todas as colunas na planilha (visíveis + oculta)
ALL_COLUMNS = VISIBLE_COLUMNS + [HIDDEN_COL_TIPO]

# Índices (0-based) das colunas numéricas
COL_IDX_SA = 3       # Saldo Anterior
COL_IDX_NAT_SA = 4   # Natureza SA
COL_IDX_DEB = 5      # Débito
COL_IDX_CRED = 6     # Crédito
COL_IDX_SAT = 7      # Saldo Atual
COL_IDX_NAT_SAT = 8  # Natureza SAT
COL_IDX_TIPO = 9     # Tipo (oculta)

NUMERIC_COL_INDICES = {COL_IDX_SA, COL_IDX_DEB, COL_IDX_CRED, COL_IDX_SAT}

# Larguras
COL_WIDTHS = {
    0: 12,   # Código
    1: 16,   # Classificação
    2: 42,   # Descrição
    3: 18,   # Saldo Anterior
    4: 4,    # Natureza SA
    5: 18,   # Débito
    6: 18,   # Crédito
    7: 18,   # Saldo Atual
    8: 4,    # Natureza SAT
    9: 6,    # Tipo (oculta)
}

# Custom number format brasileiro: #.##0,00
BR_NUMBER_FORMAT = '#,##0.00'

# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

AGRUPADORA_FONT = Font(name="Calibri", bold=True, size=11)
AGRUPADORA_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

NORMAL_FONT = Font(name="Calibri", size=11)

RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


# ---------------------------------------------------------------------------
# Mapeamento de colunas: CSV → Excel
# ---------------------------------------------------------------------------

def _detect_csv_layout(header: list[str]) -> dict[str, int | None]:
    """Detecta posição das colunas no CSV de entrada.

    Suporta layouts com ou sem Tipo, com ou sem Natureza SA/SAT.

    Returns:
        Dict com índices: codigo, classificacao, descricao, tipo,
        saldo_anterior, nat_sa, debito, credito, saldo_atual, nat_sat.
        Valores None = coluna não presente.
    """
    mapping: dict[str, int | None] = {
        "codigo": None,
        "classificacao": None,
        "descricao": None,
        "tipo": None,
        "saldo_anterior": None,
        "nat_sa": None,
        "debito": None,
        "credito": None,
        "saldo_atual": None,
        "nat_sat": None,
    }

    for i, h in enumerate(header):
        hl = h.strip().lower()
        if "c" == hl[:1] and ("dig" in hl or "ódigo" in hl or "odigo" in hl):
            mapping["codigo"] = i
        elif "classifica" in hl:
            mapping["classificacao"] = i
        elif "descri" in hl:
            mapping["descricao"] = i
        elif hl == "tipo" or hl == "type":
            mapping["tipo"] = i
        elif "saldo" in hl and "ant" in hl:
            mapping["saldo_anterior"] = i
        elif "natureza" in hl and "sa" in hl and "sat" not in hl:
            mapping["nat_sa"] = i
        elif "natureza" in hl and "sat" in hl:
            mapping["nat_sat"] = i
        elif "d" in hl and ("bit" in hl or "ébit" in hl or "ebito" in hl):
            mapping["debito"] = i
        elif "cr" in hl and ("dit" in hl or "édit" in hl or "edito" in hl):
            mapping["credito"] = i
        elif "saldo" in hl and "at" in hl:
            mapping["saldo_atual"] = i

    # Fallback posicional se nomes não detectados
    if mapping["codigo"] is None and len(header) >= 7:
        mapping["codigo"] = 0
    if mapping["classificacao"] is None and len(header) >= 7:
        mapping["classificacao"] = 1
    if mapping["descricao"] is None and len(header) >= 7:
        mapping["descricao"] = 2

    return mapping


def _parse_br_float(value_str: str) -> float | None:
    """Parseia valor brasileiro para float.

    "1.234.567,89" → 1234567.89
    "0,00" → 0.0
    "" → None
    """
    s = value_str.strip().replace("**", "")
    if not s:
        return None

    # Remove D/C se grudado
    if s and s[-1] in ("D", "C", "d", "c"):
        s = s[:-1].strip()

    if not s:
        return None

    # Remove sinal
    negative = False
    if s.startswith("-"):
        negative = True
        s = s[1:]
    elif s.startswith("+"):
        s = s[1:]

    # Formato BR: ponto=milhar, vírgula=decimal
    s = s.replace(".", "").replace(",", ".")

    try:
        val = float(s)
        return -val if negative else val
    except ValueError:
        return None


def _map_row(
    csv_row: list[str],
    csv_layout: dict[str, int | None],
) -> list:
    """Mapeia uma linha CSV para a ordem de colunas do Excel.

    Returns:
        Lista com 10 elementos na ordem ALL_COLUMNS.
        Valores numéricos são float (ou None se vazio).
    """
    def _get(key: str) -> str:
        idx = csv_layout[key]
        if idx is None or idx >= len(csv_row):
            return ""
        return csv_row[idx].strip().replace("**", "")

    codigo = _get("codigo")
    classif = _get("classificacao")
    descricao = _get("descricao")
    tipo = _get("tipo")

    sa_str = _get("saldo_anterior")
    nat_sa = _get("nat_sa")
    deb_str = _get("debito")
    cred_str = _get("credito")
    sat_str = _get("saldo_atual")
    nat_sat = _get("nat_sat")

    # Parse numéricos
    sa_val = _parse_br_float(sa_str)
    deb_val = _parse_br_float(deb_str)
    cred_val = _parse_br_float(cred_str)
    sat_val = _parse_br_float(sat_str)

    return [
        codigo,       # 0: Código
        classif,      # 1: Classificação
        descricao,    # 2: Descrição
        sa_val,       # 3: Saldo Anterior (float)
        nat_sa,       # 4: Natureza SA
        deb_val,      # 5: Débito (float)
        cred_val,     # 6: Crédito (float)
        sat_val,      # 7: Saldo Atual (float)
        nat_sat,      # 8: Natureza SAT
        tipo.upper() if tipo else "",  # 9: Tipo
    ]


# ---------------------------------------------------------------------------
# Detecção de período
# ---------------------------------------------------------------------------

_PERIODO_RE = re.compile(
    r"(\d{2})[._\-]?(\d{4})"
    r"|(\d{4})"
)


def detect_periodo(filename: str) -> str:
    """Detecta período do nome do arquivo.

    "Balancete 02.2025.csv" → "02.2025"
    "STV_Balancete 07.2025" → "07.2025"
    "VFR Balancete 112025.csv" → "11.2025"
    "Balancete 2025.csv" → "2025"

    Returns:
        String no formato "MM.YYYY" ou "YYYY" ou "sem_periodo".
    """
    # Usa o nome inteiro (não Path.stem) porque ".2025" pode ser
    # interpretado como extensão pelo pathlib
    name = filename

    # Pattern 1: MM separado por . _ - espaço de YYYY
    m = re.search(r"(\d{2})[._\- ](\d{4})", name)
    if m:
        mm, yyyy = int(m.group(1)), int(m.group(2))
        if 1 <= mm <= 12 and 2000 <= yyyy <= 2099:
            return f"{m.group(1)}_{m.group(2)}"

    # Pattern 2: MMYYYY concatenado (6 dígitos)
    m = re.search(r"(\d{2})(\d{4})", name)
    if m:
        mm, yyyy = int(m.group(1)), int(m.group(2))
        if 1 <= mm <= 12 and 2000 <= yyyy <= 2099:
            return f"{m.group(1)}_{m.group(2)}"

    # Pattern 3: Apenas YYYY
    m = re.search(r"(\d{4})", name)
    if m:
        year = int(m.group(1))
        if 2000 <= year <= 2099:
            return m.group(1)

    return "sem_periodo"


def _sheet_name_from_periodo(periodo: str) -> str:
    """Converte período para nome de aba.

    "02.2025" → "02.2025"
    "2025" → "2025"
    """
    return periodo


def _table_name_from_periodo(periodo: str, version: int = 1) -> str:
    """Converte período para nome de Table.

    "02.2025" → "Tab_02_2025"
    "2025" → "Tab_2025"
    version > 1 → "Tab_02_2025_v2"
    """
    safe = periodo.replace(".", "_")
    name = f"Tab_{safe}"
    if version > 1:
        name += f"_v{version}"
    return name


# ---------------------------------------------------------------------------
# Builder principal
# ---------------------------------------------------------------------------

class BalanceteXlsxBuilder:
    """Constrói Excel profissional a partir de dados CSV unificados.

    Uso:
        builder = BalanceteXlsxBuilder(unified_rows, periodo="02.2025")
        result = builder.build(output_path)
    """

    def __init__(
        self,
        unified_rows: list[list[str]],
        periodo: str | None = None,
        filename: str | None = None,
        sign_config: SignConfig | None = None,
    ):
        """
        Args:
            unified_rows: Linhas CSV (header + dados), já processadas
                          pelo csv_parser (com Natureza SA/SAT split).
            periodo: Período no formato "MM.YYYY" ou "YYYY".
            filename: Nome original do arquivo (para detectar período).
            sign_config: Configuração de sinais D/C.
        """
        if not unified_rows or len(unified_rows) < 2:
            raise ValueError("unified_rows precisa ter header + pelo menos 1 linha.")

        self._raw_header = unified_rows[0]
        self._raw_data = unified_rows[1:]
        self._csv_layout = _detect_csv_layout(self._raw_header)
        self._sign_config = sign_config

        # Detecta período
        if periodo:
            self._periodo = periodo
        elif filename:
            self._periodo = detect_periodo(filename)
        else:
            self._periodo = "sem_periodo"

        # Mapeia dados para formato Excel
        self._rows: list[list] = []
        for raw_row in self._raw_data:
            mapped = _map_row(raw_row, self._csv_layout)
            self._rows.append(mapped)

        # Controle de filtragem por nível de detalhe
        self._force_values: bool = False        # True → agrupadoras usam valor direto
        self._collapsed_parents: set[int] = set()  # Row indices de pais colapsados

        logger.info(
            "XlsxBuilder: %d linhas, período=%s",
            len(self._rows), self._periodo,
        )

    @property
    def periodo(self) -> str:
        return self._periodo

    def detect_signs(self) -> SignDetectionResult:
        """Detecta modo de sinais nos dados."""
        str_rows = self._to_string_rows()
        return detect_sign_mode(
            rows=str_rows,
            sa_col=COL_IDX_SA,
            sat_col=COL_IDX_SAT,
            nat_sa_col=COL_IDX_NAT_SA,
            nat_sat_col=COL_IDX_NAT_SAT,
            classif_col=1,
        )

    def apply_signs(self, config: SignConfig | None = None) -> None:
        """Aplica conversão D/C → +/- nos dados."""
        cfg = config or self._sign_config
        if cfg and cfg.mode == "skip":
            return

        str_rows = self._to_string_rows()
        converted = apply_sign_convention(
            rows=str_rows,
            sa_col=COL_IDX_SA,
            sat_col=COL_IDX_SAT,
            nat_sa_col=COL_IDX_NAT_SA,
            nat_sat_col=COL_IDX_NAT_SAT,
            classif_col=1,
            config=cfg,
        )

        # Re-parseia de volta para float
        for i, str_row in enumerate(converted):
            if i < len(self._rows):
                for col in NUMERIC_COL_INDICES:
                    if col < len(str_row):
                        val = _parse_br_float(str_row[col])
                        self._rows[i][col] = val
                # Limpa natureza se já foi incorporada ao sinal
                if COL_IDX_NAT_SA < len(str_row):
                    self._rows[i][COL_IDX_NAT_SA] = str_row[COL_IDX_NAT_SA]
                if COL_IDX_NAT_SAT < len(str_row):
                    self._rows[i][COL_IDX_NAT_SAT] = str_row[COL_IDX_NAT_SAT]

    def filter_rows(
        self,
        detail_level: str = "completo",
        collapsed_classifs: list[str] | None = None,
    ) -> None:
        """Filtra linhas conforme nível de detalhe.

        Args:
            detail_level: "completo", "agrupadoras" ou "personalizado".
            collapsed_classifs: Classificações a colapsar (modo personalizado).
        """
        if detail_level == "completo":
            return

        if detail_level == "agrupadoras":
            self._rows = [
                row for row in self._rows
                if COL_IDX_TIPO < len(row)
                and str(row[COL_IDX_TIPO]).strip().upper() == "A"
            ]
            self._force_values = True
            logger.info("Filtro agrupadoras: %d linhas mantidas", len(self._rows))
            return

        if detail_level == "personalizado" and collapsed_classifs:
            collapsed_set = set(collapsed_classifs)
            rows_to_keep: list[list] = []
            self._collapsed_parents = set()

            for row in self._rows:
                classif = str(row[1]).strip() if len(row) > 1 else ""

                # Verifica se é descendente de alguma classificação colapsada
                is_descendant = any(
                    classif.startswith(c + ".") for c in collapsed_set
                    if classif != c  # não remover a própria agrupadora
                )

                if is_descendant:
                    continue

                # Se é uma das agrupadoras colapsadas, marca para usar valor direto
                if classif in collapsed_set:
                    self._collapsed_parents.add(len(rows_to_keep))

                rows_to_keep.append(row)

            removed = len(self._rows) - len(rows_to_keep)
            self._rows = rows_to_keep
            logger.info(
                "Filtro personalizado: %d linhas removidas, %d mantidas",
                removed, len(rows_to_keep),
            )

    def build(
        self,
        output_path: Path | str | None = None,
        existing_workbook: Path | str | None = None,
        version: int = 1,
    ) -> Path:
        """Gera o Excel.

        Args:
            output_path: Caminho do arquivo de saída.
            existing_workbook: Se fornecido, adiciona aba a workbook existente.
            version: Versão da aba (1=original, 2+=resubmissão).

        Returns:
            Path do arquivo gerado.
        """
        if existing_workbook and Path(existing_workbook).exists():
            wb = load_workbook(str(existing_workbook))
        else:
            wb = Workbook()
            # Remove aba padrão se vamos criar outra
            if wb.sheetnames == ["Sheet"]:
                del wb["Sheet"]

        sheet_name = _sheet_name_from_periodo(self._periodo)
        if version > 1:
            sheet_name = f"{sheet_name}_v{version}"

        # Se aba já existe, remove (reescrita)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

        ws = wb.create_sheet(title=sheet_name)

        # Hierarquia para SUM formulas
        hierarchy = build_hierarchy(
            self._rows,
            classif_col=1,
            tipo_col=COL_IDX_TIPO,
        )

        # Escreve dados
        self._write_header(ws)
        self._write_data(ws, hierarchy)

        # Formatação
        self._apply_column_widths(ws)
        self._apply_conditional_formatting(ws)
        self._hide_tipo_column(ws)

        # Freeze panes
        ws.freeze_panes = "A2"

        # Excel Table
        self._create_table(ws, version)

        # Salva
        if output_path is None:
            output_path = existing_workbook or "balancete.xlsx"

        out = Path(output_path)
        wb.save(str(out))
        logger.info(
            "XLSX gerado: %s (aba '%s', %d linhas)",
            out, sheet_name, len(self._rows),
        )
        return out

    # -------------------------------------------------------------------
    # Métodos internos
    # -------------------------------------------------------------------

    def _to_string_rows(self) -> list[list[str]]:
        """Converte rows de volta para strings (para sign_logic)."""
        result = []
        for row in self._rows:
            str_row = []
            for i, val in enumerate(row):
                if i in NUMERIC_COL_INDICES and val is not None:
                    # Formata como BR string
                    formatted = f"{abs(val):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                    if val < 0:
                        formatted = f"-{formatted}"
                    str_row.append(formatted)
                elif val is None:
                    str_row.append("")
                else:
                    str_row.append(str(val))
            result.append(str_row)
        return result

    def _write_header(self, ws) -> None:
        """Escreve linha de cabeçalho."""
        for col_idx, col_name in enumerate(ALL_COLUMNS):
            cell = ws.cell(row=1, column=col_idx + 1, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER

    def _write_data(self, ws, hierarchy: dict[int, list[int]]) -> None:
        """Escreve dados com formatação e fórmulas SUM.

        Para agrupadoras com filhos, valida se |SUM(filhos)| ≈ |valor original|
        antes de colocar fórmula. Se não bate, mantém o valor do PDF (ground truth).
        """
        # Pré-valida quais agrupadoras podem usar fórmula SUM
        valid_sums = self._validate_hierarchy_sums(hierarchy)

        for row_idx, row_data in enumerate(self._rows):
            excel_row = row_idx + 2  # +1 header, +1 openpyxl 1-based

            is_agrupadora = (
                COL_IDX_TIPO < len(row_data)
                and str(row_data[COL_IDX_TIPO]).strip().upper() == "A"
            )

            for col_idx, value in enumerate(row_data):
                cell = ws.cell(row=excel_row, column=col_idx + 1)
                cell.border = THIN_BORDER

                # Valor
                if col_idx in NUMERIC_COL_INDICES:
                    use_formula = (
                        is_agrupadora
                        and row_idx in hierarchy
                        and row_idx in valid_sums
                        and not self._force_values
                        and row_idx not in self._collapsed_parents
                    )
                    if use_formula:
                        # Agrupadora validada: fórmula SUM
                        formula = self._build_sum_formula(
                            col_idx, hierarchy[row_idx], excel_row
                        )
                        cell.value = formula
                        # Comment com valor original do Gemini
                        if value is not None:
                            orig = self._format_br_comment(value)
                            cell.comment = Comment(
                                f"Gemini original: {orig}",
                                "Sistema",
                            )
                    elif (is_agrupadora and row_idx in hierarchy
                          and not self._force_values
                          and row_idx not in self._collapsed_parents):
                        # Agrupadora com soma divergente: mantém valor do PDF
                        cell.value = value if value is not None else 0.0
                        # Comment de aviso com a divergência
                        children = hierarchy[row_idx]
                        children_sum = sum(
                            self._rows[c][col_idx]
                            for c in children
                            if col_idx < len(self._rows[c])
                            and self._rows[c][col_idx] is not None
                        )
                        orig_str = self._format_br_comment(value)
                        sum_str = self._format_br_comment(children_sum)
                        cell.comment = Comment(
                            f"Soma filhos: {sum_str}\n"
                            f"Valor PDF: {orig_str}\n"
                            f"Mantido valor original (divergência em módulo).",
                            "Sistema",
                        )
                    else:
                        cell.value = value if value is not None else 0.0

                    cell.number_format = BR_NUMBER_FORMAT
                    cell.alignment = RIGHT_ALIGN
                elif col_idx in (COL_IDX_NAT_SA, COL_IDX_NAT_SAT):
                    cell.value = value if value else ""
                    cell.alignment = CENTER_ALIGN
                elif col_idx == COL_IDX_TIPO:
                    cell.value = value if value else ""
                    cell.alignment = CENTER_ALIGN
                else:
                    cell.value = value if value else ""
                    cell.alignment = LEFT_ALIGN

    @staticmethod
    def _format_br_comment(value: float | None) -> str:
        """Formata valor float para string brasileira (para Comments)."""
        if value is None:
            return "0,00"
        formatted = (
            f"{abs(value):,.2f}"
            .replace(",", "X")
            .replace(".", ",")
            .replace("X", ".")
        )
        return f"-{formatted}" if value < 0 else formatted

    def _build_sum_formula(
        self,
        col_idx: int,
        children: list[int],
        parent_excel_row: int,
    ) -> str:
        """Constrói fórmula SUM para agrupadora.

        Args:
            col_idx: Índice da coluna (0-based).
            children: Índices (0-based em self._rows) dos filhos diretos.
            parent_excel_row: Linha Excel do pai.

        Returns:
            Fórmula Excel tipo "=SUM(D5,D8,D12)".
        """
        col_letter = get_column_letter(col_idx + 1)
        refs = []
        for child_idx in children:
            child_excel_row = child_idx + 2  # +1 header, +1 1-based
            refs.append(f"{col_letter}{child_excel_row}")

        if not refs:
            return 0.0

        return f"=SUM({','.join(refs)})"

    def _validate_hierarchy_sums(
        self, hierarchy: dict[int, list[int]]
    ) -> set[int]:
        """Valida quais agrupadoras têm soma de filhos compatível em módulo.

        Compara |SUM(filhos)| com |valor original| para cada coluna numérica.
        Se pelo menos 2 de 4 colunas batem (tolerância 1%), a fórmula SUM
        é considerada segura. Caso contrário, o valor original do PDF
        deve ser mantido.

        Args:
            hierarchy: Mapeamento pai → filhos diretos.

        Returns:
            Set de row_indices onde a fórmula SUM é válida.
        """
        valid: set[int] = set()
        kept_original = 0

        for parent_idx, children in hierarchy.items():
            parent_row = self._rows[parent_idx]
            matches = 0
            total_checked = 0

            for col_idx in NUMERIC_COL_INDICES:
                parent_val = (
                    parent_row[col_idx]
                    if col_idx < len(parent_row)
                    else None
                )
                if parent_val is None:
                    continue

                # Soma valores originais dos filhos diretos
                children_sum = 0.0
                for child_idx in children:
                    child_row = self._rows[child_idx]
                    child_val = (
                        child_row[col_idx]
                        if col_idx < len(child_row)
                        else None
                    )
                    if child_val is not None:
                        children_sum += child_val

                total_checked += 1

                # Ambos zero → match
                if abs(parent_val) < 0.005 and abs(children_sum) < 0.005:
                    matches += 1
                    continue

                # Um zero e outro não → sem match
                if abs(parent_val) < 0.005 or abs(children_sum) < 0.005:
                    continue

                # Compara em módulo (tolerância 1% ou 0.02)
                tolerance = max(abs(parent_val) * 0.01, 0.02)
                if abs(abs(children_sum) - abs(parent_val)) <= tolerance:
                    matches += 1

            if total_checked < 2:
                # Poucos dados para validar — usa fórmula
                valid.add(parent_idx)
            elif matches >= 2:
                valid.add(parent_idx)
            else:
                kept_original += 1

        if kept_original:
            logger.info(
                "Validação SUM: %d agrupadoras com fórmula, "
                "%d mantidas com valor original (soma não confere em módulo).",
                len(valid), kept_original,
            )
        else:
            logger.info(
                "Validação SUM: todas as %d agrupadoras com fórmula validada.",
                len(valid),
            )

        return valid

    def _apply_column_widths(self, ws) -> None:
        """Define larguras das colunas."""
        for col_idx, width in COL_WIDTHS.items():
            col_letter = get_column_letter(col_idx + 1)
            ws.column_dimensions[col_letter].width = width

    def _apply_conditional_formatting(self, ws) -> None:
        """Aplica formatação condicional baseada na coluna Tipo.

        Tipo=A → negrito + fundo azul claro (toda a linha)
        """
        tipo_col_letter = get_column_letter(COL_IDX_TIPO + 1)  # J
        last_row = len(self._rows) + 1  # +1 header
        last_col_letter = get_column_letter(len(ALL_COLUMNS))

        # Range: A2 até a última coluna/linha
        data_range = f"A2:{last_col_letter}{last_row}"

        # Regra: se $J2="A" → formato agrupadora
        formula = f'${tipo_col_letter}2="A"'

        ws.conditional_formatting.add(
            data_range,
            FormulaRule(
                formula=[formula],
                font=AGRUPADORA_FONT,
                fill=AGRUPADORA_FILL,
            ),
        )

    def _hide_tipo_column(self, ws) -> None:
        """Oculta a coluna Tipo."""
        col_letter = get_column_letter(COL_IDX_TIPO + 1)
        ws.column_dimensions[col_letter].hidden = True

    def _create_table(self, ws, version: int = 1) -> None:
        """Cria Excel Table (ListObject) no range de dados."""
        if not self._rows:
            return

        last_row = len(self._rows) + 1
        last_col_letter = get_column_letter(len(ALL_COLUMNS))
        table_ref = f"A1:{last_col_letter}{last_row}"
        table_name = _table_name_from_periodo(self._periodo, version)

        # Sanitiza nome da tabela (Excel restringe caracteres)
        table_name = re.sub(r"[^A-Za-z0-9_]", "_", table_name)

        table = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleLight9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        logger.info("Table '%s' criada: %s", table_name, table_ref)


# ---------------------------------------------------------------------------
# API pública simplificada
# ---------------------------------------------------------------------------

def build_xlsx(
    unified_rows: list[list[str]],
    filename: str,
    output_dir: Path | str | None = None,
    periodo: str | None = None,
    sign_config: SignConfig | None = None,
    existing_workbook: Path | str | None = None,
    version: int = 1,
) -> tuple[Path, SignDetectionResult | None]:
    """API de alto nível para gerar Excel profissional.

    Args:
        unified_rows: Linhas CSV (header + dados).
        filename: Nome base do arquivo.
        output_dir: Diretório de saída.
        periodo: Período (ou detecta do filename).
        sign_config: Config de sinais (None = auto-detect).
        existing_workbook: Workbook existente para adicionar aba.
        version: Versão (1=original, 2+=resubmissão).

    Returns:
        Tupla (Path do xlsx, SignDetectionResult ou None).
    """
    from src.utils.config import OUTPUT_DIR

    out_dir = Path(output_dir) if output_dir else OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    safe_name = re.sub(r'[<>:"/\\|?*]', "_", filename)

    builder = BalanceteXlsxBuilder(
        unified_rows=unified_rows,
        periodo=periodo,
        filename=filename,
        sign_config=sign_config,
    )

    # Detecção de sinais
    sign_result = builder.detect_signs()

    # Aplica sinais se configurado (auto ou com config explícito)
    if sign_config and sign_config.mode != "skip":
        builder.apply_signs(sign_config)
    elif sign_config is None and sign_result.has_dc and sign_result.matches_convention:
        # Auto: convenção padrão confirmada, aplica
        builder.apply_signs(SignConfig(mode="auto"))

    # Caminho de saída
    if existing_workbook:
        output_path = Path(existing_workbook)
    else:
        output_path = out_dir / f"{safe_name}.xlsx"

    result_path = builder.build(
        output_path=output_path,
        existing_workbook=existing_workbook,
        version=version,
    )

    return result_path, sign_result
