"""Test xlsx_builder + sign_logic — cenários reais de sinais contábeis."""
import tempfile
from pathlib import Path

from src.exporters.xlsx_builder import BalanceteXlsxBuilder, detect_periodo, _parse_br_float
from src.exporters.sign_logic import (
    STANDARD_CONVENTION,
    SignConfig,
    apply_sign_convention,
    detect_sign_mode,
)

print("=" * 60)
print("TEST 1: STANDARD_CONVENTION")
print("=" * 60)

# Verifica que a convencao está correta
assert STANDARD_CONVENTION[1] == {"D": 1, "C": -1}, "Ativo: D=+, C=-"
assert STANDARD_CONVENTION[2] == {"D": -1, "C": 1}, "Passivo: D=-, C=+"
assert STANDARD_CONVENTION[3] == {"D": -1, "C": 1}, "Despesas: D=-, C=+"
assert STANDARD_CONVENTION[4] == {"D": -1, "C": 1}, "Receitas: D=-, C=+"
print("  Grupo 1 (Ativo):    D=+, C=-  OK")
print("  Grupo 2 (Passivo):  D=-, C=+  OK")
print("  Grupo 3 (Despesas): D=-, C=+  OK")
print("  Grupo 4 (Receitas): D=-, C=+  OK")

print()
print("=" * 60)
print("TEST 2: SINAIS — Depreciacao no Ativo (C = negativo)")
print("=" * 60)

# Cenário real: depreciacao acumulada e conta do Ativo (1.x) com natureza C
# Deve ficar NEGATIVA porque no Ativo D=+ e C=-
sign_rows = [
    # [Cod, Classif, Desc, SA, Nat_SA, Deb, Cred, SAT, Nat_SAT, Tipo]
    # Índices: 0=Cod, 1=Classif, 2=Desc, 3=SA, 4=Nat_SA, 5=Deb, 6=Cred, 7=SAT, 8=Nat_SAT, 9=Tipo
    ["1", "1",       "ATIVO",                    "500.000,00", "D", "50.000,00", "30.000,00", "520.000,00", "D", "A"],
    ["2", "1.2",     "NAO CIRCULANTE",           "300.000,00", "D", "20.000,00", "10.000,00", "310.000,00", "D", "A"],
    ["3", "1.2.01",  "IMOBILIZADO",              "400.000,00", "D", "10.000,00", "0,00",      "410.000,00", "D", "D"],
    ["4", "1.2.02",  "DEPRECIACAO ACUMULADA",    "100.000,00", "C", "0,00",      "5.000,00",  "105.000,00", "C", "D"],  # <-- C no Ativo!
]

# Aplica convencao
config = SignConfig(mode="auto")
converted = apply_sign_convention(
    rows=sign_rows,
    sa_col=3,
    sat_col=7,
    nat_sa_col=4,
    nat_sat_col=8,
    classif_col=1,
    config=config,
)

# Verifica depreciacao (row index 3 = "1.2.02")
dep_sa = converted[3][3]    # Saldo Anterior
dep_sat = converted[3][7]   # Saldo Atual

print(f"  ATIVO SA:          {converted[0][3]}  (D -> positivo)")
print(f"  IMOBILIZADO SA:    {converted[2][3]}  (D -> positivo)")
print(f"  DEPRECIACAO SA:    {dep_sa}  (C -> NEGATIVO)")
print(f"  DEPRECIACAO SAT:   {dep_sat}  (C -> NEGATIVO)")

assert dep_sa.startswith("-"), f"Depreciacao SA deveria ser negativa, got: {dep_sa}"
assert dep_sat.startswith("-"), f"Depreciacao SAT deveria ser negativa, got: {dep_sat}"
assert dep_sa == "-100.000,00", f"Expected -100.000,00, got {dep_sa}"
assert dep_sat == "-105.000,00", f"Expected -105.000,00, got {dep_sat}"

# Ativo normal deve ser positivo
ativo_sa = converted[0][3]
assert not ativo_sa.startswith("-"), f"Ativo SA deveria ser positivo, got: {ativo_sa}"
print("  DEPRECIAÇÃO C->NEGATIVO: OK OK")

print()
print("=" * 60)
print("TEST 3: SINAIS — Passivo (C = positivo, D = negativo)")
print("=" * 60)

passivo_rows = [
    ["1", "2",      "PASSIVO",            "200.000,00", "C", "10.000,00", "20.000,00", "210.000,00", "C", "A"],
    ["2", "2.1",    "CIRCULANTE",         "100.000,00", "C", "5.000,00",  "10.000,00", "105.000,00", "C", "A"],
    ["3", "2.1.01", "FORNECEDORES",       "80.000,00",  "C", "3.000,00",  "8.000,00",  "85.000,00",  "C", "D"],
    ["4", "2.1.02", "ADIANT. CLIENTES",   "20.000,00",  "D", "2.000,00",  "2.000,00",  "20.000,00",  "D", "D"],  # D no Passivo!
]

converted_p = apply_sign_convention(
    rows=passivo_rows, sa_col=3, sat_col=7, nat_sa_col=4, nat_sat_col=8,
    classif_col=1, config=config,
)

print(f"  PASSIVO SA:         {converted_p[0][3]}  (C -> positivo)")
print(f"  FORNECEDORES SA:    {converted_p[2][3]}  (C -> positivo)")
print(f"  ADIANT.CLIENTES SA: {converted_p[3][3]}  (D -> NEGATIVO)")

assert converted_p[2][3] == "80.000,00", f"Fornecedores deveria ser +, got: {converted_p[2][3]}"
assert converted_p[3][3] == "-20.000,00", f"Adiant deveria ser -, got: {converted_p[3][3]}"
print("  PASSIVO D->NEGATIVO / C->POSITIVO: OK OK")

print()
print("=" * 60)
print("TEST 4: SINAIS — Despesas (D=-, C=+) e Receitas (D=-, C=+)")
print("=" * 60)

dre_rows = [
    # Despesas (grupo 3): convencao D=-, C=+
    ["1", "3",      "DESPESAS",          "50.000,00", "D", "10.000,00", "2.000,00", "58.000,00", "D", "A"],
    ["2", "3.1",    "DESP OPERACIONAIS", "30.000,00", "D", "8.000,00",  "1.000,00", "37.000,00", "D", "A"],
    ["3", "3.1.01", "SALARIOS",          "20.000,00", "D", "5.000,00",  "0,00",     "25.000,00", "D", "D"],
    ["4", "3.1.02", "REVERSAO PROVISAO", "5.000,00",  "C", "0,00",      "3.000,00", "2.000,00",  "C", "D"],  # C em Despesa
    # Receitas (grupo 4): convencao D=-, C=+
    ["5", "4",      "RECEITAS",          "100.000,00", "C", "5.000,00", "20.000,00", "115.000,00", "C", "A"],
    ["6", "4.1",    "REC OPERACIONAIS",  "80.000,00",  "C", "3.000,00", "15.000,00", "92.000,00",  "C", "A"],
    ["7", "4.1.01", "VENDAS",            "60.000,00",  "C", "2.000,00", "10.000,00", "68.000,00",  "C", "D"],
    ["8", "4.1.02", "DEVOLUCOES",        "10.000,00",  "D", "1.000,00", "0,00",      "11.000,00",  "D", "D"],  # D em Receita = devolucao
]

converted_d = apply_sign_convention(
    rows=dre_rows, sa_col=3, sat_col=7, nat_sa_col=4, nat_sat_col=8,
    classif_col=1, config=config,
)

print(f"  SALARIOS SA:        {converted_d[2][3]}  (Desp D -> NEGATIVO)")
print(f"  REVERSAO SA:        {converted_d[3][3]}  (Desp C -> positivo)")
print(f"  VENDAS SA:          {converted_d[6][3]}  (Rec C -> positivo)")
print(f"  DEVOLUCOES SA:      {converted_d[7][3]}  (Rec D -> NEGATIVO)")

# Despesas: D=-, C=+  (nossa convencao)
assert converted_d[2][3] == "-20.000,00", f"Salarios deveria -, got: {converted_d[2][3]}"
assert converted_d[3][3] == "5.000,00", f"Reversao deveria +, got: {converted_d[3][3]}"
# Receitas: D=-, C=+
assert converted_d[6][3] == "60.000,00", f"Vendas deveria +, got: {converted_d[6][3]}"
assert converted_d[7][3] == "-10.000,00", f"Devolucoes deveria -, got: {converted_d[7][3]}"
print("  DESPESAS D=-/C=+: OK OK")
print("  RECEITAS D=-/C=+: OK OK")

print()
print("=" * 60)
print("TEST 5: detect_periodo")
print("=" * 60)
assert detect_periodo("Balancete 02.2025.pdf") == "02_2025", f"Got: {detect_periodo('Balancete 02.2025.pdf')}"
assert detect_periodo("VFR Balancete 112025.csv") == "11_2025", f"Got: {detect_periodo('VFR Balancete 112025.csv')}"
assert detect_periodo("Balancete 2025.csv") == "2025", f"Got: {detect_periodo('Balancete 2025.csv')}"
print("  detect_periodo: OK OK")

print()
print("=" * 60)
print("TEST 6: Build XLSX com sinais aplicados")
print("=" * 60)

# Dados com Tipo no index 3 (como o header CSV real do sistema)
# Header: Codigo(0) Classificacao(1) Descricao(2) Tipo(3) SA(4) NatSA(5) Deb(6) Cred(7) SAT(8) NatSAT(9)
xlsx_header = [
    "Codigo", "Classificacao", "Descricao", "Tipo",
    "Saldo Anterior", "Natureza SA", "Debito", "Credito",
    "Saldo Atual", "Natureza SAT",
]
xlsx_rows = [
    xlsx_header,
    # ATIVO = NAO CIRCULANTE (único filho direto): SA=300k, Deb=10k, Cred=5k, SAT=305k
    ["1", "1",       "ATIVO",                 "A", "300.000,00", "D", "10.000,00", "5.000,00",  "305.000,00", "D"],
    # NAO CIRCULANTE = IMOB + DEPREC: SA=300k, Deb=10k, Cred=5k, SAT=305k
    ["2", "1.2",     "NAO CIRCULANTE",        "A", "300.000,00", "D", "10.000,00", "5.000,00",  "305.000,00", "D"],
    # IMOBILIZADO: SA=400k D, Deb=10k, Cred=0, SAT=410k D
    ["3", "1.2.01",  "IMOBILIZADO",           "D", "400.000,00", "D", "10.000,00", "0,00",      "410.000,00", "D"],
    # DEPRECIACAO: SA=100k C, Deb=0, Cred=5k, SAT=105k C (sinais invertidos pelo sign_logic)
    ["4", "1.2.02",  "DEPRECIACAO ACUMULADA",  "D", "100.000,00", "C", "0,00",      "5.000,00",  "105.000,00", "C"],
]

builder = BalanceteXlsxBuilder(xlsx_rows, periodo="02.2025", sign_config=SignConfig(mode="auto"))

# Detect
det = builder.detect_signs()
print(f"  has_dc={det.has_dc}, matches_convention={det.matches_convention}")

# Apply
builder.apply_signs()

# Verifica dados internos pos-apply
dep_internal = builder._rows[3]
print(f"  Internal DEPREC SA={dep_internal[3]} SAT={dep_internal[7]}")
assert isinstance(dep_internal[3], float) and dep_internal[3] < 0, f"Internal SA deveria <0, got {dep_internal[3]}"

# Build
tmp = Path(tempfile.mkdtemp()) / "test_signs.xlsx"
result = builder.build(output_path=tmp)
print(f"  XLSX: {result} ({result.stat().st_size} bytes)")

# Verify
from openpyxl import load_workbook

wb = load_workbook(str(result))
ws = wb["02.2025"]

# Col D = SA (col 4), Col H = SAT (col 8)
# Row 2 = ATIVO, Row 3 = NAO CIRCULANTE, Row 4 = IMOBILIZADO, Row 5 = DEPRECIACAO
dep_sa_cell = ws.cell(5, 4)
dep_sat_cell = ws.cell(5, 8)

print(f"  DEPRECIACAO SA cell:  {dep_sa_cell.value}")
print(f"  DEPRECIACAO SAT cell: {dep_sat_cell.value}")

# Depreciacao deve ser negativa no Excel
assert isinstance(dep_sa_cell.value, (int, float)), f"SA deveria ser numerico, got {type(dep_sa_cell.value)}"
assert dep_sa_cell.value < 0, f"Depreciacao SA deveria ser < 0, got {dep_sa_cell.value}"
assert dep_sat_cell.value < 0, f"Depreciacao SAT deveria ser < 0, got {dep_sat_cell.value}"
print(f"  DEPRECIACAO NEGATIVA NO XLSX: OK")

# Imobilizado deve ser positivo
imob_sa = ws.cell(4, 4).value
print(f"  IMOBILIZADO SA cell:  {imob_sa}")
assert isinstance(imob_sa, (int, float)), f"Imob SA deveria ser numerico, got {type(imob_sa)}"
assert imob_sa > 0, f"Imobilizado SA deveria ser > 0, got {imob_sa}"
print(f"  IMOBILIZADO POSITIVO NO XLSX: OK")

# ATIVO e NAO CIRCULANTE agrupadoras devem ter formula SUM (somas consistentes)
ativo_sa = ws.cell(2, 4).value
print(f"  ATIVO SA (agrupadora): {ativo_sa}")
assert isinstance(ativo_sa, str) and ativo_sa.startswith("=SUM"), f"ATIVO SA deveria ser SUM, got {ativo_sa}"
print(f"  ATIVO COM FORMULA SUM: OK")

nc_sa = ws.cell(3, 4).value
print(f"  NAO CIRCULANTE SA (agrupadora): {nc_sa}")
assert isinstance(nc_sa, str) and nc_sa.startswith("=SUM"), f"NAO CIRC SA deveria ser SUM, got {nc_sa}"
print(f"  NAO CIRCULANTE COM FORMULA SUM: OK")

wb.close()

print()
print("=" * 60)
print("TEST 7: Build XLSX — agrupadora com soma divergente (mantém valor original)")
print("=" * 60)

# Cenário: ATIVO diz 500k mas filho único (CIRCULANTE) é 200k → soma NÃO bate
# A validação deve manter o valor original do PDF, sem fórmula SUM
xlsx_rows_div = [
    xlsx_header,
    ["1", "1",      "ATIVO",         "A", "500.000,00", "D", "50.000,00", "30.000,00", "520.000,00", "D"],
    ["2", "1.1",    "CIRCULANTE",    "A", "200.000,00", "D", "20.000,00", "10.000,00", "210.000,00", "D"],
    ["3", "1.1.01", "CAIXA",         "D", "100.000,00", "D", "15.000,00", "5.000,00",  "110.000,00", "D"],
    ["4", "1.1.02", "BANCOS",        "D", "100.000,00", "D", "5.000,00",  "5.000,00",  "100.000,00", "D"],
]

builder_div = BalanceteXlsxBuilder(xlsx_rows_div, periodo="03.2025", sign_config=SignConfig(mode="auto"))
builder_div.apply_signs()

tmp_div = Path(tempfile.mkdtemp()) / "test_divergent.xlsx"
result_div = builder_div.build(output_path=tmp_div)

wb_div = load_workbook(str(result_div))
ws_div = wb_div["03.2025"]

# ATIVO (row 2): soma filhos = 200k ≠ 500k → deve manter valor numérico, NÃO fórmula
ativo_div_sa = ws_div.cell(2, 4).value
print(f"  ATIVO SA (divergente): {ativo_div_sa} (tipo={type(ativo_div_sa).__name__})")
assert isinstance(ativo_div_sa, (int, float)), f"ATIVO divergente deveria ser numerico, got {type(ativo_div_sa)}: {ativo_div_sa}"
assert ativo_div_sa == 500000.0, f"ATIVO deveria manter 500k do PDF, got {ativo_div_sa}"
print(f"  ATIVO MANTEVE VALOR ORIGINAL (divergente): OK")

# CIRCULANTE (row 3): soma filhos = CAIXA(100k) + BANCOS(100k) = 200k → bate → fórmula SUM
circ_sa = ws_div.cell(3, 4).value
print(f"  CIRCULANTE SA: {circ_sa}")
assert isinstance(circ_sa, str) and circ_sa.startswith("=SUM"), f"CIRCULANTE deveria ser SUM, got {circ_sa}"
print(f"  CIRCULANTE COM FORMULA SUM (consistente): OK")

wb_div.close()

print()
print("=" * 60)
print("TEST 8: filter_rows — Somente Agrupadoras")
print("=" * 60)

# Usa os mesmos dados do Test 6 (4 linhas: 2 agrupadoras + 2 detalhe)
builder_agrup = BalanceteXlsxBuilder(xlsx_rows, periodo="04.2025", sign_config=SignConfig(mode="skip"))
assert len(builder_agrup._rows) == 4, f"Deveria ter 4 linhas, got {len(builder_agrup._rows)}"

builder_agrup.filter_rows(detail_level="agrupadoras")
assert len(builder_agrup._rows) == 2, f"Apos filtro deveria ter 2 linhas (so A), got {len(builder_agrup._rows)}"

# Verifica que todas as linhas restantes sao Tipo=A
for i, row in enumerate(builder_agrup._rows):
    assert str(row[9]).strip().upper() == "A", f"Row {i} deveria ser A, got {row[9]}"
print(f"  Filtro agrupadoras: 4 -> {len(builder_agrup._rows)} linhas  OK")

# Build e verifica que nao tem formulas SUM (force_values=True)
tmp_agrup = Path(tempfile.mkdtemp()) / "test_agrup.xlsx"
result_agrup = builder_agrup.build(output_path=tmp_agrup)

wb_agrup = load_workbook(str(result_agrup))
ws_agrup = wb_agrup["04.2025"]

# Row 2 = ATIVO, Row 3 = NAO CIRCULANTE (ambas sem filhos visíveis → valor direto)
ativo_agrup_sa = ws_agrup.cell(2, 4).value
print(f"  ATIVO SA (somente agrup): {ativo_agrup_sa} (tipo={type(ativo_agrup_sa).__name__})")
assert isinstance(ativo_agrup_sa, (int, float)), f"Deveria ser numerico (sem formula), got {type(ativo_agrup_sa)}: {ativo_agrup_sa}"
print(f"  AGRUPADORAS SEM FORMULA SUM: OK")

wb_agrup.close()

print()
print("=" * 60)
print("TEST 9: filter_rows — Personalizado (colapsar grupo)")
print("=" * 60)

# Dados: 1 -> 1.1 -> 1.1.01, 1.1.02 + 1.2 -> 1.2.01
xlsx_rows_custom = [
    xlsx_header,
    ["1", "1",       "ATIVO",                 "A", "500.000,00", "D", "30.000,00", "15.000,00", "515.000,00", "D"],
    ["2", "1.1",     "CIRCULANTE",            "A", "200.000,00", "D", "20.000,00", "10.000,00", "210.000,00", "D"],
    ["3", "1.1.01",  "CAIXA",                 "D", "100.000,00", "D", "15.000,00", "5.000,00",  "110.000,00", "D"],
    ["4", "1.1.02",  "BANCOS",                "D", "100.000,00", "D", "5.000,00",  "5.000,00",  "100.000,00", "D"],
    ["5", "1.2",     "NAO CIRCULANTE",        "A", "300.000,00", "D", "10.000,00", "5.000,00",  "305.000,00", "D"],
    ["6", "1.2.01",  "IMOBILIZADO",           "D", "300.000,00", "D", "10.000,00", "5.000,00",  "305.000,00", "D"],
]

builder_custom = BalanceteXlsxBuilder(xlsx_rows_custom, periodo="05.2025", sign_config=SignConfig(mode="skip"))
assert len(builder_custom._rows) == 6, f"Deveria ter 6 linhas, got {len(builder_custom._rows)}"

# Colapsar "1.1" (Circulante) — remove 1.1.01 e 1.1.02
builder_custom.filter_rows(detail_level="personalizado", collapsed_classifs=["1.1"])

# Devem sobrar: 1, 1.1, 1.2, 1.2.01 (4 linhas)
assert len(builder_custom._rows) == 4, f"Apos colapso deveria ter 4 linhas, got {len(builder_custom._rows)}"

# Verifica classificacoes restantes
remaining = [str(r[1]).strip() for r in builder_custom._rows]
assert remaining == ["1", "1.1", "1.2", "1.2.01"], f"Classificacoes erradas: {remaining}"
print(f"  Colapso 1.1: 6 -> {len(builder_custom._rows)} linhas  OK")
print(f"  Classificacoes restantes: {remaining}  OK")

# Build e verifica
tmp_custom = Path(tempfile.mkdtemp()) / "test_custom.xlsx"
result_custom = builder_custom.build(output_path=tmp_custom)

wb_custom = load_workbook(str(result_custom))
ws_custom = wb_custom["05.2025"]

# Row 3 = 1.1 (colapsada) → deve ter valor direto, nao formula
circ_custom_sa = ws_custom.cell(3, 4).value
print(f"  CIRCULANTE SA (colapsada): {circ_custom_sa} (tipo={type(circ_custom_sa).__name__})")
assert isinstance(circ_custom_sa, (int, float)), f"1.1 colapsada deveria ser numerico, got {type(circ_custom_sa)}: {circ_custom_sa}"
print(f"  1.1 COLAPSADA USA VALOR DIRETO: OK")

# Row 4 = 1.2 (NAO colapsada, filho 1.2.01 visivel) → pode ter formula SUM
nc_custom_sa = ws_custom.cell(4, 4).value
print(f"  NAO CIRCULANTE SA: {nc_custom_sa} (tipo={type(nc_custom_sa).__name__})")
assert isinstance(nc_custom_sa, str) and nc_custom_sa.startswith("=SUM"), f"1.2 deveria ter SUM, got {nc_custom_sa}"
print(f"  1.2 NAO COLAPSADA TEM FORMULA SUM: OK")

wb_custom.close()

print()
print("=" * 60)
print("ALL TESTS PASSED OK")
print("=" * 60)
