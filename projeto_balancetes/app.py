"""Conversor de Balancetes PDF para CSV — FastAPI backend.

Execute: python app.py
Acesse: http://localhost:8000
"""

from __future__ import annotations

import asyncio
import shutil
import tempfile
import time
import uuid
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

# Bootstrap do core (DEVE ser antes dos imports do core)
from controladoria_core.utils.config import configure as _configure
_configure(project_root=Path(__file__).parent)

import fitz  # PyMuPDF — para contar paginas
import uvicorn
from fastapi import FastAPI, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

from controladoria_core.exporters.sign_logic import SignConfig, SignDetectionResult
from controladoria_core.exporters.xlsx_builder import BalanceteXlsxBuilder, build_xlsx, detect_periodo
from controladoria_core.exporters.reference_extractor import (
    extract_reference_from_xlsx,
    list_references,
    load_reference_for_prompt,
    save_reference,
)
from controladoria_core.orchestrator import Orchestrator, OutputFormat
from controladoria_core.utils.config import MODELOS_DISPONIVEIS, config, logger

# ---------------------------------------------------------------------------
# App
# ---------------------------------------------------------------------------
app = FastAPI(title="Conversor de Balancetes")

STATIC_DIR = Path(__file__).parent / "static"
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

# ---------------------------------------------------------------------------
# In-memory job store
# ---------------------------------------------------------------------------

@dataclass
class FileInfo:
    """Info sobre um PDF enviado."""
    name: str
    path: Path
    pages: int = 0
    size: int = 0


@dataclass
class JobProgress:
    """Progresso de um arquivo individual."""
    filename: str
    pages: int = 0
    status: str = "pending"  # pending | processing | done | error
    stage: str = ""  # analyzing | extracting | exporting | classifying
    stage_detail: str = ""  # ex: "Lote 2/3 (páginas 6-10)"
    error: str | None = None
    output_files: list[str] = field(default_factory=list)
    cost: float = 0.0
    time: float = 0.0


@dataclass
class Job:
    """Representa um job de conversao."""
    id: str
    status: str = "uploaded"  # uploaded | converting | done | error
    files: list[FileInfo] = field(default_factory=list)
    progress: list[JobProgress] = field(default_factory=list)
    output_dir: Path | None = None
    total_pages: int = 0
    completed: int = 0
    total: int = 0
    error: str | None = None
    started_at: float = 0.0  # timestamp do inicio
    preview_data: dict[str, list[list[str]]] = field(default_factory=dict)


jobs: dict[str, Job] = {}

# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@app.get("/")
async def index():
    """Serve a pagina principal."""
    return FileResponse(str(STATIC_DIR / "index.html"))


@app.post("/upload")
async def upload(files: list[UploadFile]):
    """Recebe PDFs e retorna job_id com info dos arquivos."""
    if not files:
        raise HTTPException(400, "Nenhum arquivo enviado.")

    job_id = str(uuid.uuid4())[:8]
    tmp_dir = Path(tempfile.mkdtemp(prefix=f"conv_{job_id}_"))
    output_dir = tmp_dir / "output"
    output_dir.mkdir()

    file_infos: list[FileInfo] = []

    for f in files:
        if not f.filename or not f.filename.lower().endswith(".pdf"):
            continue

        # Sanitiza nome do arquivo (remove caracteres problemáticos no Windows)
        safe_filename = f.filename.replace("/", "_").replace("\\", "_")
        dest = tmp_dir / safe_filename
        content = await f.read()

        if len(content) == 0:
            logger.warning("Arquivo vazio ignorado: %s", safe_filename)
            continue

        dest.write_bytes(content)
        logger.info("Upload: %s (%d bytes) -> %s", safe_filename, len(content), dest)

        # Conta paginas
        try:
            doc = fitz.open(str(dest))
            pages = len(doc)
            doc.close()
        except Exception as exc:
            logger.warning("Erro ao abrir PDF %s: %s", safe_filename, exc)
            pages = 0

        file_infos.append(FileInfo(
            name=safe_filename,
            path=dest,
            pages=pages,
            size=len(content),
        ))

    if not file_infos:
        raise HTTPException(400, "Nenhum PDF valido enviado.")

    total_pages = sum(fi.pages for fi in file_infos)

    job = Job(
        id=job_id,
        files=file_infos,
        output_dir=output_dir,
        total_pages=total_pages,
        total=len(file_infos),
    )
    jobs[job_id] = job

    return {
        "job_id": job_id,
        "files": [
            {"name": fi.name, "pages": fi.pages, "size": fi.size}
            for fi in file_infos
        ],
        "total_pages": total_pages,
        "estimated_cost": _estimate_cost(total_pages),
    }


@app.post("/convert/{job_id}")
async def convert(job_id: str, workers: int = 3, reference: str | None = None):
    """Inicia conversao em background.

    Args:
        job_id: ID do job.
        workers: Número de workers paralelos.
        reference: Nome (filename stem) da referência a usar. Se None, usa a mais recente.
    """
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Job nao encontrado.")

    if job.status == "converting":
        raise HTTPException(409, "Conversao ja em andamento.")

    workers = max(1, min(36, workers))

    job.status = "converting"
    job.completed = 0
    job.started_at = time.time()
    job.progress = [
        JobProgress(filename=fi.name, pages=fi.pages) for fi in job.files
    ]

    if reference:
        logger.info("Conversão com referência selecionada: %s", reference)

    # Roda em background
    asyncio.get_event_loop().run_in_executor(
        None, _run_conversion, job, workers, reference
    )

    return {"status": "started", "workers": workers, "reference": reference}


@app.get("/progress/{job_id}")
async def progress_sse(job_id: str):
    """Server-Sent Events com progresso em tempo real."""
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Job nao encontrado.")

    async def event_stream():
        import json

        last_snapshot = ""
        while True:
            elapsed = time.time() - job.started_at if job.started_at > 0 else 0

            data = {
                "status": job.status,
                "completed": job.completed,
                "total": job.total,
                "elapsed": round(elapsed, 1),
                "progress": [
                    {
                        "filename": p.filename,
                        "pages": p.pages,
                        "status": p.status,
                        "stage": p.stage,
                        "stage_detail": p.stage_detail,
                        "error": p.error,
                        "output_files": p.output_files,
                        "cost": p.cost,
                        "time": p.time,
                    }
                    for p in job.progress
                ],
            }

            snapshot = json.dumps(data, ensure_ascii=False)
            # Envia sempre (para atualizar elapsed timer) ou quando dados mudam
            if snapshot != last_snapshot or True:
                last_snapshot = snapshot
                yield f"data: {snapshot}\n\n"

                if job.status in ("done", "error"):
                    break

            await asyncio.sleep(0.3)

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@app.get("/results/{job_id}")
async def results(job_id: str):
    """Lista CSVs e XLSXs gerados."""
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Job nao encontrado.")

    output_files = []
    if job.output_dir and job.output_dir.exists():
        for f in sorted(job.output_dir.iterdir()):
            if f.suffix.lower() in (".csv", ".xlsx"):
                ext = f.suffix.lower().lstrip(".")
                output_files.append({
                    "name": f.name,
                    "size": f.stat().st_size,
                    "type": ext,
                })

    total_cost = sum(p.cost for p in job.progress)
    total_time = sum(p.time for p in job.progress)

    return {
        "status": job.status,
        "files": output_files,
        "total_cost": round(total_cost, 6),
        "total_time": round(total_time, 2),
        "preview_data": job.preview_data,
    }


@app.get("/download/{job_id}/{filename}")
async def download(job_id: str, filename: str):
    """Baixa um CSV ou XLSX individual."""
    job = jobs.get(job_id)
    if not job or not job.output_dir:
        raise HTTPException(404, "Job nao encontrado.")

    file_path = job.output_dir / filename
    if not file_path.exists():
        raise HTTPException(404, "Arquivo nao encontrado.")

    media_types = {
        ".csv": "text/csv",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    media_type = media_types.get(file_path.suffix.lower(), "application/octet-stream")

    return FileResponse(
        str(file_path),
        media_type=media_type,
        filename=filename,
    )


@app.get("/download-all/{job_id}")
async def download_all(job_id: str):
    """Baixa todos os CSVs e XLSXs como ZIP."""
    job = jobs.get(job_id)
    if not job or not job.output_dir:
        raise HTTPException(404, "Job nao encontrado.")

    output_files = [
        f for f in job.output_dir.iterdir()
        if f.suffix.lower() in (".csv", ".xlsx")
    ]
    if not output_files:
        raise HTTPException(404, "Nenhum arquivo gerado.")

    zip_path = job.output_dir.parent / f"balancetes_{job_id}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for out_file in output_files:
            zf.write(out_file, out_file.name)

    return FileResponse(
        str(zip_path),
        media_type="application/zip",
        filename=f"balancetes_{job_id}.zip",
    )


@app.delete("/job/{job_id}/{filename}")
async def remove_file(job_id: str, filename: str):
    """Remove um PDF do job antes da conversao."""
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Job nao encontrado.")

    if job.status == "converting":
        raise HTTPException(409, "Nao pode remover durante conversao.")

    for i, fi in enumerate(job.files):
        if fi.name == filename:
            fi.path.unlink(missing_ok=True)
            job.files.pop(i)
            job.total = len(job.files)
            job.total_pages = sum(f.pages for f in job.files)
            return {
                "removed": filename,
                "total_pages": job.total_pages,
                "estimated_cost": _estimate_cost(job.total_pages),
            }

    raise HTTPException(404, "Arquivo nao encontrado no job.")


# ---------------------------------------------------------------------------
# Excel Profissional — endpoints
# ---------------------------------------------------------------------------

@app.post("/convert-xlsx/{job_id}")
async def convert_xlsx(job_id: str, body: dict | None = None):
    """Gera XLSX profissional consolidado (todas as abas em um único arquivo).

    Body (opcional):
        sign_mode: str — "auto", "skip", ou "ask" (padrão: "auto")
        existing_xlsx: str — nome de XLSX existente no output_dir para adicionar abas
        version: int — versão (1=original, 2+=resubmissão)
        detail_level: str — "completo", "agrupadoras" ou "personalizado" (padrão: "completo")
        collapsed_classifs: list[str] — classificações a colapsar (modo personalizado)
    """
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Job nao encontrado.")

    if not job.preview_data:
        raise HTTPException(400, "Nenhum dado convertido disponível. Execute /convert primeiro.")

    body = body or {}
    sign_mode = body.get("sign_mode", "auto")
    existing_xlsx_name = body.get("existing_xlsx")
    version = body.get("version", 1)
    detail_level = body.get("detail_level", "completo")
    collapsed_classifs = body.get("collapsed_classifs", [])

    sign_config = SignConfig(mode=sign_mode)

    # Arquivo único consolidado
    output_xlsx = job.output_dir / "Balancetes_Profissional.xlsx"

    # Se há XLSX existente para adicionar abas, usa ele como base
    existing_wb = None
    if existing_xlsx_name and job.output_dir:
        candidate = job.output_dir / existing_xlsx_name
        if candidate.exists():
            output_xlsx = candidate
            existing_wb = candidate

    tabs_info = []

    # Ordena por período para abas ficarem em ordem cronológica
    items = sorted(
        job.preview_data.items(),
        key=lambda x: detect_periodo(x[0]),
    )

    for base_name, unified_rows in items:
        if not unified_rows or len(unified_rows) < 2:
            continue

        periodo = detect_periodo(base_name)

        hdr = unified_rows[0] if unified_rows else []
        has_nat = any("natureza" in h.lower() for h in hdr)
        logger.info("[XLSX-SIGN] %s: %d cols, has_nat=%s", base_name, len(hdr), has_nat)

        try:
            builder = BalanceteXlsxBuilder(
                unified_rows=unified_rows,
                periodo=periodo,
                filename=base_name,
                sign_config=sign_config,
            )

            sign_result = builder.detect_signs()
            logger.info(
                "[XLSX-SIGN] %s: has_dc=%s, matches=%s, mode=%s",
                base_name, sign_result.has_dc, sign_result.matches_convention,
                sign_config.mode,
            )

            # Aplica sinais
            if sign_config.mode != "skip":
                builder.apply_signs(sign_config)
                # Verifica resultado
                neg_count = sum(1 for r in builder._rows if isinstance(r[3], (int, float)) and r[3] < 0)
                logger.info("[XLSX-SIGN] %s: APOS apply -> %d valores negativos em SA", base_name, neg_count)
            elif sign_config is None and sign_result.has_dc and sign_result.matches_convention:
                builder.apply_signs(SignConfig(mode="auto"))

            # Filtra por nível de detalhe
            if detail_level != "completo":
                builder.filter_rows(
                    detail_level=detail_level,
                    collapsed_classifs=collapsed_classifs or None,
                )

            # Build — cada iteração adiciona uma aba ao mesmo workbook
            result_path = builder.build(
                output_path=output_xlsx,
                existing_workbook=output_xlsx if output_xlsx.exists() else existing_wb,
                version=version,
            )

            tabs_info.append({
                "periodo": periodo,
                "sign_detection": {
                    "has_dc": sign_result.has_dc,
                    "has_signs": sign_result.has_signs,
                    "matches_convention": sign_result.matches_convention,
                    "needs_user_input": sign_result.needs_user_input,
                    "details": sign_result.details,
                },
            })
        except Exception as exc:
            logger.error("Erro ao gerar aba XLSX para %s: %s", base_name, exc)
            tabs_info.append({
                "periodo": periodo,
                "error": str(exc),
            })

    return {
        "files": [{
            "filename": output_xlsx.name,
            "periodos": [t["periodo"] for t in tabs_info if "error" not in t],
            "tabs_count": len([t for t in tabs_info if "error" not in t]),
            "sign_detection": tabs_info[0].get("sign_detection", {}) if tabs_info else {},
        }],
        "tabs": tabs_info,
    }


@app.post("/detect-signs/{job_id}/{base_name}")
async def detect_signs(job_id: str, base_name: str):
    """Detecta modo de sinais D/C nos dados de um arquivo."""
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Job nao encontrado.")

    unified_rows = job.preview_data.get(base_name)
    if not unified_rows or len(unified_rows) < 2:
        raise HTTPException(404, f"Dados não encontrados para '{base_name}'.")

    try:
        builder = BalanceteXlsxBuilder(unified_rows, filename=base_name)
        result = builder.detect_signs()
        return {
            "has_dc": result.has_dc,
            "has_signs": result.has_signs,
            "matches_convention": result.matches_convention,
            "needs_user_input": result.needs_user_input,
            "details": result.details,
        }
    except Exception as exc:
        raise HTTPException(500, str(exc))


@app.post("/upload-xlsx")
async def upload_xlsx(file: UploadFile, job_id: str | None = None):
    """Upload de XLSX existente para adicionar abas.

    Se job_id fornecido, salva no output_dir do job.
    Retorna info sobre as abas existentes.
    """
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Envie um arquivo .xlsx.")

    content = await file.read()
    if not content:
        raise HTTPException(400, "Arquivo vazio.")

    # Salva no output_dir do job ou em temp
    if job_id:
        job = jobs.get(job_id)
        if not job or not job.output_dir:
            raise HTTPException(404, "Job nao encontrado.")
        dest = job.output_dir / file.filename
    else:
        tmp_dir = Path(tempfile.mkdtemp(prefix="xlsx_"))
        dest = tmp_dir / file.filename

    dest.write_bytes(content)

    # Lê info das abas
    from openpyxl import load_workbook
    try:
        wb = load_workbook(str(dest), read_only=True)
        sheets = wb.sheetnames
        wb.close()
    except Exception as exc:
        raise HTTPException(400, f"Erro ao ler XLSX: {exc}")

    return {
        "filename": file.filename,
        "path": str(dest),
        "sheets": sheets,
    }


@app.post("/resubmit/{job_id}/{base_name}")
async def resubmit(job_id: str, base_name: str, body: dict):
    """Reenvia PDF com prompt de correção ao Gemini.

    Body:
        correction: str — instrução de correção (ex: "A conta 1.1.01 está errada...")
        version: int — próxima versão (default: 2)
    """
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Job nao encontrado.")

    correction = body.get("correction", "")
    if not correction:
        raise HTTPException(400, "Campo 'correction' é obrigatório.")

    version = body.get("version", 2)

    # Encontra o PDF original
    pdf_file = None
    for fi in job.files:
        stem = fi.name.rsplit(".", 1)[0]
        if stem == base_name:
            pdf_file = fi
            break

    if not pdf_file:
        raise HTTPException(404, f"PDF original '{base_name}' não encontrado.")

    # Monta prompt com correção
    correction_prompt = (
        f"ATENÇÃO — CORREÇÃO SOLICITADA PELO USUÁRIO:\n"
        f"{correction}\n\n"
        f"Reprocesse o PDF aplicando a correção acima. "
        f"Mantenha todas as outras contas como estavam."
    )

    # Processa novamente em background
    import threading

    def _resubmit_worker():
        try:
            orch = Orchestrator()
            # Injeta correction no prompt
            agent = orch._gemini_agent
            result = agent.process(
                str(pdf_file.path),
                prompt=correction_prompt,
                financial=True,
            )

            if result.success and result.text:
                from controladoria_core.parsers.csv_parser import save_as_csv
                csv_paths, unified_rows = save_as_csv(
                    result.text,
                    f"{base_name}_v{version}",
                    output_dir=job.output_dir,
                )
                # Guarda dados para /convert-xlsx
                job.preview_data[f"{base_name}_v{version}"] = unified_rows
                logger.info("Resubmissão v%d concluída: %s", version, base_name)
            else:
                logger.error("Resubmissão falhou: %s", result.error)

        except Exception as exc:
            logger.error("Erro na resubmissão: %s", exc)

    thread = threading.Thread(target=_resubmit_worker, daemon=True)
    thread.start()

    return {
        "status": "resubmitting",
        "version": version,
        "base_name": base_name,
    }


# ---------------------------------------------------------------------------
# Referência / RAG — endpoints
# ---------------------------------------------------------------------------

@app.post("/save-reference/{job_id}")
async def save_reference_endpoint(job_id: str, body: dict | None = None):
    """Extrai padrão do XLSX validado e salva como referência para o Gemini.

    O XLSX Profissional já deve ter sido gerado com sinais aplicados.
    O sistema extrai:
    - Hierarquia de agrupadoras (qual soma quais filhos)
    - Convenção de sinais por grupo
    - Plano de contas com tipos A/D

    Body (opcional):
        sheet_name: str — nome da aba específica (default: primeira)
        instructions: str — instruções/problemas encontrados pelo usuário
    """
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Job nao encontrado.")

    if not job.output_dir:
        raise HTTPException(400, "Nenhum output disponível.")

    # Localiza o XLSX profissional
    xlsx_path = job.output_dir / "Balancetes_Profissional.xlsx"
    if not xlsx_path.exists():
        # Procura qualquer XLSX
        xlsx_files = list(job.output_dir.glob("*.xlsx"))
        if not xlsx_files:
            raise HTTPException(400, "Nenhum XLSX encontrado. Gere o XLSX Profissional primeiro.")
        xlsx_path = xlsx_files[0]

    body = body or {}
    sheet_name = body.get("sheet_name")
    instructions = body.get("instructions", "").strip()
    ref_name = body.get("name", "").strip()

    try:
        ref = extract_reference_from_xlsx(xlsx_path, sheet_name=sheet_name)
        txt_path, json_path = save_reference(
            ref, user_instructions=instructions, name=ref_name,
        )

        return {
            "status": "ok",
            "display_name": ref_name or ref.empresa,
            "empresa": ref.empresa,
            "periodo": ref.periodo,
            "total_contas": ref.total_contas,
            "grupos": len(ref.grupos),
            "hierarchy_nodes": len(ref.hierarchy_tree),
            "sign_examples": len(ref.sign_examples),
            "has_instructions": bool(instructions),
            "txt_file": txt_path.name,
            "json_file": json_path.name,
            "preview": ref.to_prompt_text()[:2000],
        }
    except Exception as exc:
        logger.error("Erro ao salvar referência: %s", exc)
        raise HTTPException(500, f"Erro ao extrair referência: {exc}")


@app.post("/upload-reference")
async def upload_reference_endpoint(
    file: UploadFile,
    instructions: str = Form(""),
    name: str = Form(""),
    model: str = Form(""),
):
    """Recebe XLSX corrigido pelo controller e extrai referência.

    O controller pode anexar seu próprio XLSX (corrigido manualmente)
    para que a IA aprenda o padrão correto.

    Form fields:
        file: UploadFile — XLSX corrigido
        instructions: str — instruções/problemas encontrados
        name: str — nome customizado da referência
    """
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Apenas arquivos .xlsx são aceitos.")

    # Salva o XLSX em diretório temporário
    import tempfile

    tmp_dir = Path(tempfile.mkdtemp(prefix="ref_"))
    tmp_path = tmp_dir / file.filename
    content = await file.read()
    tmp_path.write_bytes(content)

    ref_name = name.strip()
    ref_model = model.strip() if model else ""
    if ref_model:
        logger.info("Referência upload com modelo de análise: %s", ref_model)

    try:
        ref = extract_reference_from_xlsx(tmp_path)
        txt_path, json_path = save_reference(
            ref, user_instructions=instructions.strip(), name=ref_name,
        )

        return {
            "status": "ok",
            "source": "upload",
            "filename": file.filename,
            "display_name": ref_name or ref.empresa,
            "empresa": ref.empresa,
            "periodo": ref.periodo,
            "total_contas": ref.total_contas,
            "grupos": len(ref.grupos),
            "hierarchy_nodes": len(ref.hierarchy_tree),
            "sign_examples": len(ref.sign_examples),
            "has_instructions": bool(instructions.strip()),
            "txt_file": txt_path.name,
            "json_file": json_path.name,
            "preview": ref.to_prompt_text()[:2000],
        }
    except Exception as exc:
        logger.error("Erro ao processar XLSX de referência: %s", exc)
        raise HTTPException(500, f"Erro ao extrair referência: {exc}")
    finally:
        # Limpa temporário
        tmp_path.unlink(missing_ok=True)
        tmp_dir.rmdir()


@app.get("/references")
async def get_references():
    """Lista referências disponíveis no knowledge/."""
    refs = list_references()
    has_active = load_reference_for_prompt() is not None
    return {
        "references": refs,
        "has_active": has_active,
    }


@app.get("/references/{filename}")
async def get_reference_detail(filename: str):
    """Retorna o JSON completo de uma referência."""
    from controladoria_core.utils.config import KNOWLEDGE_DIR
    import json as _json

    json_path = KNOWLEDGE_DIR / f"{filename}.json"
    if not json_path.exists():
        raise HTTPException(status_code=404, detail="Referência não encontrada")
    return _json.loads(json_path.read_text(encoding="utf-8"))


@app.delete("/references/{filename}")
async def delete_reference(filename: str):
    """Remove uma referência do knowledge/."""
    from controladoria_core.utils.config import KNOWLEDGE_DIR

    for ext in (".txt", ".json"):
        fpath = KNOWLEDGE_DIR / f"{filename}{ext}"
        if fpath.exists():
            fpath.unlink()

    return {"status": "ok", "removed": filename}


# ---------------------------------------------------------------------------
# Chat — atualizar referências via IA
# ---------------------------------------------------------------------------

@app.post("/chat-reference")
async def chat_reference(body: dict):
    """Chat com IA para ajustar referências existentes.

    Body:
        reference_name: str — filename stem da referência
        message: str — instrução do usuário
        model: str — modelo a usar (default: gemini-2.5-flash)
        history: list — histórico [{role, content}, ...]
    """
    ref_name = body.get("reference_name", "")
    message = body.get("message", "").strip()
    model_id = body.get("model", "gemini-2.5-flash")
    history = body.get("history", [])

    if not ref_name:
        raise HTTPException(400, "Selecione uma referência.")
    if not message:
        raise HTTPException(400, "Digite uma mensagem.")

    # Carrega referência atual
    ref_text = load_reference_for_prompt(reference_name=ref_name)
    if not ref_text:
        raise HTTPException(404, f"Referência '{ref_name}' não encontrada.")

    # Monta prompt de contexto
    system_prompt = (
        "Você é um assistente especializado em contabilidade brasileira. "
        "O usuário tem uma referência de balancete (plano de contas, hierarquia, sinais D/C). "
        "Sua tarefa é ajudar a ajustar essa referência com base nas instruções do usuário.\n\n"
        "REFERÊNCIA ATUAL:\n"
        f"{ref_text}\n\n"
        "INSTRUÇÕES:\n"
        "- Analise a referência acima e responda a pergunta/instrução do usuário.\n"
        "- Se o usuário pedir uma alteração, descreva EXATAMENTE o que deve mudar.\n"
        "- Se precisar atualizar a referência, responda com a seção modificada.\n"
        "- Responda em português.\n"
    )

    # Monta conversa
    conversation_parts = [system_prompt]
    for msg in history[:-1]:  # todas exceto a última (que é a mensagem atual)
        role_label = "Usuário" if msg.get("role") == "user" else "Assistente"
        conversation_parts.append(f"{role_label}: {msg.get('content', '')}")
    conversation_parts.append(f"Usuário: {message}")

    full_prompt = "\n\n".join(conversation_parts)

    try:
        from google import genai
        from controladoria_core.utils.config import GEMINI_API_KEY

        client = genai.Client(api_key=GEMINI_API_KEY)
        response = client.models.generate_content(
            model=model_id,
            contents=[full_prompt],
            config={
                "temperature": 0.3,
                "max_output_tokens": 8192,
            },
        )

        ai_response = response.text or "Sem resposta."

        # TODO: futuramente, parsear a resposta e atualizar os arquivos .txt/.json
        # Por enquanto, o chat é apenas consultivo
        return {
            "response": ai_response,
            "updated": False,
            "model": model_id,
        }

    except Exception as exc:
        logger.error("Erro no chat de referência: %s", exc)
        raise HTTPException(500, f"Erro na comunicação com IA: {exc}")


# ---------------------------------------------------------------------------
# Modelo
# ---------------------------------------------------------------------------

@app.get("/models")
async def get_models():
    """Retorna modelos disponíveis e qual está ativo."""
    return {
        "active": config.gemini_model,
        "models": {
            model_id: info["label"]
            for model_id, info in MODELOS_DISPONIVEIS.items()
        },
    }


@app.post("/set-model")
async def set_model(body: dict):
    """Altera o modelo Gemini ativo."""
    model_id = body.get("model", "")
    if model_id not in MODELOS_DISPONIVEIS:
        raise HTTPException(400, f"Modelo inválido: {model_id}")
    config.gemini_model = model_id
    logger.info("Modelo alterado para: %s", model_id)
    return {"active": model_id}


# ---------------------------------------------------------------------------
# Conversao em background
# ---------------------------------------------------------------------------

def _run_conversion(job: Job, max_workers: int, reference_name: str | None = None) -> None:
    """Executa a conversao de PDFs em threads paralelas."""
    logger.info("Iniciando conversao: %d PDFs, %d workers, ref=%s", job.total, max_workers, reference_name)

    def _process_one(idx: int, fi: FileInfo) -> tuple[int, dict[str, Any]]:
        """Worker: processa um PDF usando Orchestrator."""
        p = job.progress[idx]

        # Stagger: cada worker espera idx * 2s para evitar burst na API
        if idx > 0:
            delay = idx * 2
            p.stage = "queued"
            p.stage_detail = f"Aguardando {delay}s..."
            logger.info("Worker %d: aguardando %ds (stagger)", idx, delay)
            time.sleep(delay)

        p.stage = "processing"
        p.stage_detail = "Processando..."
        start = time.time()

        try:
            file_path = str(fi.path)
            logger.info("Worker %d: processando %s (path=%s, exists=%s)",
                        idx, fi.name, file_path, Path(file_path).exists())

            orch = Orchestrator()
            result = orch.process(
                file_path,
                output_format=OutputFormat.CSV,
                output_dir=job.output_dir,
                reference_name=reference_name,
            )
            elapsed = time.time() - start

            if result.success:
                return idx, {
                    "status": "done",
                    "output_files": [
                        Path(f).name for f in result.output_files
                        if f.endswith((".csv", ".xlsx"))
                    ],
                    "cost": result.estimated_cost,
                    "time": elapsed,
                    "preview_rows": result.details.get("preview_rows", []),
                }
            else:
                return idx, {
                    "status": "error",
                    "error": result.error or "Erro desconhecido",
                    "cost": 0.0,
                    "time": elapsed,
                }
        except Exception as exc:
            elapsed = time.time() - start
            return idx, {
                "status": "error",
                "error": str(exc),
                "cost": 0.0,
                "time": elapsed,
            }

    try:
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            for p in job.progress:
                p.status = "processing"
                p.stage = "processing"
                p.stage_detail = "Processando..."

            future_to_idx = {
                executor.submit(_process_one, i, fi): i
                for i, fi in enumerate(job.files)
            }

            for future in as_completed(future_to_idx):
                idx = future_to_idx[future]
                try:
                    _, info = future.result()
                except Exception as exc:
                    info = {"status": "error", "error": str(exc), "cost": 0.0, "time": 0.0}

                p = job.progress[idx]
                p.status = info["status"]
                p.error = info.get("error")
                p.output_files = info.get("output_files", [])
                p.cost = info.get("cost", 0.0)
                p.time = info.get("time", 0.0)
                p.stage = "done" if info["status"] == "done" else "error"
                p.stage_detail = ""
                job.completed += 1

                # Guarda preview data indexado pelo nome base do PDF
                preview_rows = info.get("preview_rows", [])
                if preview_rows:
                    base_name = job.files[idx].name.rsplit(".", 1)[0]
                    job.preview_data[base_name] = preview_rows

        job.status = "done"
        logger.info("Conversao concluida: %d/%d", job.completed, job.total)

    except Exception as exc:
        job.status = "error"
        job.error = str(exc)
        logger.error("Erro na conversao: %s", exc)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _estimate_cost(total_pages: int) -> float:
    """Estima custo baseado no numero de paginas e modelo ativo.

    ~2000 tokens/pagina input, ~1500 tokens/pagina output (estimativa).
    Pricing varia conforme o modelo selecionado.
    """
    pricing = MODELOS_DISPONIVEIS.get(config.gemini_model, {})
    input_price = pricing.get("input_price", 0.15)
    output_price = pricing.get("output_price", 0.60)
    input_tokens = total_pages * 2000
    output_tokens = total_pages * 1500
    cost = (input_tokens / 1_000_000) * input_price + (output_tokens / 1_000_000) * output_price
    return round(cost, 4)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print()
    print("=" * 50)
    print("  Conversor de Balancetes")
    print("  http://localhost:8000")
    print("=" * 50)
    print()
    uvicorn.run(app, host="0.0.0.0", port=8000)
