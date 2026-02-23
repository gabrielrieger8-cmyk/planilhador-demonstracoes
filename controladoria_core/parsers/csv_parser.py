"""Parser de saída em formato CSV.

Extrai tabelas do texto Markdown, unifica em um único CSV e remove
linhas duplicadas causadas pela sobreposição entre faixas de página.
"""

from __future__ import annotations

import csv
import io
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from controladoria_core.utils.config import OUTPUT_DIR, logger


def save_as_csv(
    text: str,
    filename: str,
    output_dir: str | Path | None = None,
) -> tuple[list[Path], list[list[str]]]:
    """Extrai tabelas do texto, unifica e salva como CSV único.

    Detecta tabelas em formato Markdown, combina todas em uma única
    tabela, remove linhas duplicadas (da sobreposição entre faixas),
    separa D/C dos valores em colunas próprias de natureza,
    e gera um único arquivo CSV.

    Args:
        text: Texto contendo tabelas em Markdown.
        filename: Nome base do arquivo (sem extensão).
        output_dir: Diretório de saída.

    Returns:
        Tupla (lista de Paths gerados, linhas unificadas).
    """
    out_dir = Path(output_dir) if output_dir else OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    safe_name = re.sub(r'[<>:"/\\|?*]', "_", filename)
    tables = extract_markdown_tables(text)

    if not tables:
        logger.warning("Nenhuma tabela encontrada no texto para exportar como CSV.")
        return [], []

    # Unifica todas as tabelas em uma só, removendo duplicatas
    unified = _unify_and_deduplicate(tables)

    if not unified:
        logger.warning("Tabela unificada ficou vazia após deduplicação.")
        return [], []

    # Pós-processamento: valida agrupadoras (insere coluna Tipo se necessário)
    if len(unified) > 1:
        header = unified[0]
        layout = _detect_column_layout(header)
        if not layout["has_tipo"]:
            # Insere coluna "Tipo" na posição 3 (após Descrição) com valor "D" padrão
            header = header[:3] + ["Tipo"] + header[3:]
            unified = [header] + [
                row[:3] + ["D"] + row[3:] for row in unified[1:]
            ]
            layout = _detect_column_layout(header)
            logger.info("Coluna 'Tipo' inserida automaticamente (não veio do Gemini).")
        data_rows = unified[1:]
        data_rows = _postprocess_agrupadoras(data_rows, layout)
        unified = [header] + data_rows

    # Verifica e corrige A/D (Tipo) que vazou para colunas numéricas
    unified = _fix_tipo_in_numeric_columns(unified)

    # Verifica e corrige D/C que ficaram grudados em colunas numéricas erradas
    unified = _fix_dc_in_numeric_columns(unified)

    # Separa D/C dos valores em colunas próprias de natureza
    unified = _split_natureza_columns(unified)

    output_path = out_dir / f"{safe_name}.csv"

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        for row in unified:
            writer.writerow(row)

    logger.info("CSV unificado salvo: %s (%d linhas)", output_path, len(unified))
    return [output_path], unified


def save_as_xlsx(
    unified_rows: list[list[str]],
    filename: str,
    output_dir: str | Path | None = None,
) -> Path:
    """Gera arquivo .xlsx formatado a partir das linhas unificadas.

    Aplica formatação profissional:
    - Header: negrito, fundo azul (#2F5496), texto branco, centralizado
    - Linhas agrupadora (Tipo=A): negrito + fundo cinza claro (#E8E8E8)
    - Colunas numéricas: alinhamento à direita
    - Freeze panes no header, auto-filtro, bordas finas
    - Larguras de coluna otimizadas

    Args:
        unified_rows: Linhas do CSV (header + dados), já processadas.
        filename: Nome base do arquivo (sem extensão).
        output_dir: Diretório de saída.

    Returns:
        Path do arquivo .xlsx gerado.
    """
    out_dir = Path(output_dir) if output_dir else OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    safe_name = re.sub(r'[<>:"/\\|?*]', "_", filename)

    if not unified_rows:
        raise ValueError("unified_rows está vazio — impossível gerar xlsx.")

    header = unified_rows[0]
    layout = _detect_column_layout(header)
    tipo_idx = layout["tipo"] if layout["has_tipo"] else -1

    # Índices das colunas numéricas (para alinhamento à direita)
    numeric_cols = {
        layout["saldo_anterior"],
        layout["debito"],
        layout["credito"],
        layout["saldo_atual"],
    }

    # --- Estilos ---
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    agrupadora_font = Font(name="Calibri", bold=True, size=11)
    agrupadora_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

    normal_font = Font(name="Calibri", size=11)
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    # --- Larguras de coluna ---
    # Nomes típicos: Código, Classificação, Descrição, Tipo, SA, Débito, Crédito, SAT
    col_widths = {}
    for i, h in enumerate(header):
        h_lower = h.strip().lower()
        if "descri" in h_lower:
            col_widths[i] = 40
        elif "classifica" in h_lower:
            col_widths[i] = 14
        elif "tipo" in h_lower:
            col_widths[i] = 6
        elif "c" == h_lower[:1] and ("dig" in h_lower or "ódigo" in h_lower):
            col_widths[i] = 12
        elif "natureza" in h_lower:
            col_widths[i] = 10
        elif i in numeric_cols:
            col_widths[i] = 18
        else:
            col_widths[i] = 14

    # --- Cria workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Balancete"

    for row_idx, row_data in enumerate(unified_rows):
        for col_idx, cell_value in enumerate(row_data):
            cell = ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
            cell.border = thin_border

            if row_idx == 0:
                # Header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            else:
                # Dados — sem formatação direta de agrupadora
                # (negrito+cor serão via formatação condicional)
                cell.font = normal_font

                # Alinhamento por tipo de coluna
                if col_idx in numeric_cols:
                    cell.alignment = right_align
                elif tipo_idx >= 0 and col_idx == tipo_idx:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

    # Larguras
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx + 1)
        ws.column_dimensions[col_letter].width = width

    # Freeze panes (header fixo)
    ws.freeze_panes = "A2"

    # Auto-filtro em todas as colunas
    if unified_rows:
        last_col = get_column_letter(len(header))
        ws.auto_filter.ref = f"A1:{last_col}{len(unified_rows)}"

    # Formatação condicional: Tipo=A → negrito + fundo cinza claro
    if tipo_idx >= 0 and len(unified_rows) > 1:
        tipo_col_letter = get_column_letter(tipo_idx + 1)
        last_col_letter = get_column_letter(len(header))
        last_row = len(unified_rows)
        data_range = f"A2:{last_col_letter}{last_row}"
        formula = f'${tipo_col_letter}2="A"'
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(
                formula=[formula],
                font=agrupadora_font,
                fill=agrupadora_fill,
            ),
        )

    output_path = out_dir / f"{safe_name}.xlsx"
    wb.save(str(output_path))

    logger.info("XLSX formatado salvo: %s (%d linhas)", output_path, len(unified_rows))
    return output_path


def _unify_and_deduplicate(
    tables: list[list[list[str]]],
) -> list[list[str]]:
    """Unifica múltiplas tabelas em uma só e remove duplicatas.

    Estratégia:
    1. Usa o cabeçalho da primeira tabela como cabeçalho único.
    2. Concatena todas as linhas de dados de todas as tabelas.
    3. Remove linhas que são cabeçalhos repetidos.
    4. Remove linhas exatamente duplicadas (mantém a primeira ocorrência).
    5. Remove duplicatas por Código+Classificação (colunas 0+1) para
       capturar linhas quase-iguais de continuações do Gemini.

    Args:
        tables: Lista de tabelas extraídas do Markdown.

    Returns:
        Lista de linhas (cada linha é lista de células) sem duplicatas.
    """
    if not tables:
        return []

    # Detecta o cabeçalho: primeira linha da primeira tabela
    header = tables[0][0] if tables[0] else []
    header_key = _row_key(header)

    # Coleta todas as linhas de dados (exceto cabeçalhos repetidos)
    seen_exact: set[str] = set()
    seen_account: set[str] = set()  # Código+Classificação
    unified: list[list[str]] = []
    duplicates_exact = 0
    duplicates_account = 0

    # Adiciona o cabeçalho
    unified.append(header)
    seen_exact.add(header_key)

    for table in tables:
        for row in table:
            key = _row_key(row)

            # Pula se é um cabeçalho repetido
            if key == header_key:
                continue

            # Pula se é uma linha exatamente duplicada
            if key in seen_exact:
                duplicates_exact += 1
                continue

            # Dedup por Código+Classificação: se as duas primeiras colunas
            # (código da conta e classificação hierárquica) são iguais,
            # é a mesma conta — mantém só a primeira ocorrência
            if len(row) >= 2:
                codigo = row[0].strip().lower()
                classificacao = row[1].strip().lower()
                # Só aplica se parecem ser dados (código numérico ou asterisco)
                if codigo and (codigo[0].isdigit() or codigo.startswith("*")):
                    account_key = f"{codigo}|{classificacao}"
                    if account_key in seen_account:
                        duplicates_account += 1
                        continue
                    seen_account.add(account_key)

            seen_exact.add(key)
            unified.append(row)

    if duplicates_exact or duplicates_account:
        logger.info(
            "Deduplicação CSV: %d exatas + %d por conta removidas",
            duplicates_exact, duplicates_account,
        )

    return unified


def _row_key(row: list[str]) -> str:
    """Gera uma chave única para uma linha, normalizando espaços.

    Args:
        row: Lista de células.

    Returns:
        String normalizada para comparação.
    """
    return "|".join(cell.strip().lower() for cell in row)


def extract_markdown_tables(text: str) -> list[list[list[str]]]:
    """Extrai tabelas de texto em formato Markdown.

    Detecta padrões de tabela Markdown (|col1|col2|...) e retorna
    como listas de linhas/células. Também lida com tabelas dentro
    de blocos de código (```markdown, ```csv, etc.).

    Args:
        text: Texto com tabelas Markdown.

    Returns:
        Lista de tabelas, cada uma sendo lista de linhas (lista de células).
    """
    # Remove blocos de código (```markdown, ```csv, ```, etc.)
    cleaned = re.sub(r"```[a-zA-Z]*\n?", "", text)

    tables: list[list[list[str]]] = []
    current_table: list[list[str]] = []
    in_table = False

    for line in cleaned.split("\n"):
        stripped = line.strip()

        # Detecta linha de tabela Markdown (com | no conteúdo)
        if "|" in stripped:
            # Ignora linhas separadoras (|---|---|) ou (---|---|)
            clean_sep = stripped.strip("|").strip()
            if clean_sep and re.match(r"^[\s\-:|]+$", clean_sep):
                continue

            # Extrai células baseado no formato
            if stripped.startswith("|") and stripped.endswith("|"):
                # Formato padrão: |col1|col2|col3|
                cells = [cell.strip() for cell in stripped.split("|")[1:-1]]
            elif stripped.startswith("|"):
                # Formato: |col1|col2|col3
                cells = [cell.strip() for cell in stripped.split("|")[1:]]
            elif stripped.endswith("|"):
                # Formato: col1|col2|col3|
                cells = [cell.strip() for cell in stripped.split("|")[:-1]]
            else:
                # Formato sem bordas: col1|col2|col3
                cells = [cell.strip() for cell in stripped.split("|")]

            # Filtra: deve ter pelo menos 3 colunas não-vazias para ser tabela
            non_empty = [c for c in cells if c]
            if len(non_empty) >= 3:
                current_table.append(cells)
                in_table = True
            elif in_table:
                # Linha com poucos campos — fim da tabela
                if current_table:
                    tables.append(current_table)
                current_table = []
                in_table = False

        elif in_table:
            # Fim da tabela
            if current_table:
                tables.append(current_table)
            current_table = []
            in_table = False

    # Captura última tabela se o texto terminar com ela
    if current_table:
        tables.append(current_table)

    if not tables:
        logger.debug(
            "Nenhuma tabela Markdown encontrada. Primeiros 500 chars do texto:\n%s",
            text[:500],
        )

    return tables


def _detect_column_layout(header: list[str]) -> dict[str, int]:
    """Detecta layout de colunas do CSV baseado no cabeçalho.

    Suporta dois layouts:
    - 7 colunas (legado): Código, Classificação, Descrição, Saldo Anterior, Débito, Crédito, Saldo Atual
    - 8 colunas (novo):   Código, Classificação, Descrição, Tipo, Saldo Anterior, Débito, Crédito, Saldo Atual

    Returns:
        Dict com índices: {tipo, saldo_anterior, debito, credito, saldo_atual, has_tipo, min_cols}
    """
    n = len(header)
    # Checa se a coluna 3 parece ser "Tipo" (A/D)
    has_tipo = False
    if n >= 8:
        col3 = header[3].strip().lower() if len(header) > 3 else ""
        if "tipo" in col3:
            has_tipo = True

    if has_tipo:
        return {
            "tipo": 3,
            "saldo_anterior": 4,
            "debito": 5,
            "credito": 6,
            "saldo_atual": 7,
            "has_tipo": True,
            "min_cols": 8,
        }
    else:
        return {
            "tipo": -1,
            "saldo_anterior": 3,
            "debito": 4,
            "credito": 5,
            "saldo_atual": 6,
            "has_tipo": False,
            "min_cols": 7,
        }


def _fix_tipo_in_numeric_columns(rows: list[list[str]]) -> list[list[str]]:
    """Detecta e corrige linhas onde A/D (Tipo) vazou para colunas numéricas.

    Quando um batch do Gemini produz colunas desalinhadas, o valor "A" ou "D"
    (indicador de agrupadora/detalhe) pode acabar em uma coluna numérica
    (Saldo Anterior, Débito, Crédito ou Saldo Atual) em vez de ficar na
    coluna Tipo (índice 3 no layout de 8 colunas).

    Padrão detectado: a linha tem o número certo de colunas, mas uma coluna
    numérica contém exatamente "A" ou "D" isolado (não "123,45D" que é
    valor+natureza). Nesse caso, a linha está deslocada e precisa ser
    realinhada.

    Estratégia de correção:
    1. Encontra a posição do "A"/"D" isolado nas colunas numéricas.
    2. Remove essa célula da posição errada.
    3. Insere na posição correta da coluna Tipo (índice 3).
    4. Se a linha ficar com menos colunas, preenche com vazio.

    Args:
        rows: Linhas unificadas (header + dados).

    Returns:
        Linhas com Tipo realinhado nas colunas corretas.
    """
    if not rows or len(rows) < 2:
        return rows

    header = rows[0]
    layout = _detect_column_layout(header)

    # Só faz sentido se o layout tem coluna Tipo
    if not layout["has_tipo"]:
        return rows

    tipo_idx = layout["tipo"]  # 3
    numeric_cols = [
        layout["saldo_anterior"],  # 4
        layout["debito"],          # 5
        layout["credito"],         # 6
        layout["saldo_atual"],     # 7
    ]
    min_cols = layout["min_cols"]  # 8

    # Regex para valor numérico brasileiro (ex: "678.044,93", "0,00", "1.234D")
    numeric_pattern = re.compile(r"^[\d.,]+[DC]?$", re.IGNORECASE)

    result = [header]
    fixes_shifted = 0
    fixes_missing = 0

    for row in rows[1:]:
        new_row = list(row)

        # === Caso 1: A/D isolado em coluna numérica (linha com 8+ colunas) ===
        tipo_found_at = -1
        for col_idx in numeric_cols:
            if col_idx >= len(new_row):
                continue
            val = new_row[col_idx].strip().upper().replace("**", "")
            if val in ("A", "D"):
                tipo_found_at = col_idx
                break

        if tipo_found_at >= 0:
            current_tipo = new_row[tipo_idx].strip().upper() if tipo_idx < len(new_row) else ""
            if current_tipo not in ("A", "D"):
                tipo_value = new_row[tipo_found_at].strip()
                new_row.pop(tipo_found_at)
                new_row.insert(tipo_idx, tipo_value)
                while len(new_row) < min_cols:
                    new_row.append("")
                fixes_shifted += 1

        # === Caso 2: Tipo omitido (linha com 7 colunas quando header tem 8) ===
        elif len(new_row) == min_cols - 1:
            col3_val = new_row[tipo_idx].strip().replace("**", "") if tipo_idx < len(new_row) else ""
            if col3_val and numeric_pattern.match(col3_val):
                new_row.insert(tipo_idx, "")
                fixes_missing += 1

        result.append(new_row)

    total_fixes = fixes_shifted + fixes_missing
    if total_fixes:
        parts = []
        if fixes_shifted:
            parts.append(f"{fixes_shifted} realinhada(s)")
        if fixes_missing:
            parts.append(f"{fixes_missing} com Tipo inserido")
        logger.info(
            "Verificação A/D: %d linha(s) corrigida(s) — %s.",
            total_fixes, ", ".join(parts),
        )

    return result


def _fix_dc_in_numeric_columns(rows: list[list[str]]) -> list[list[str]]:
    """Verifica e corrige D/C que ficaram grudados em colunas numéricas erradas.

    O Gemini às vezes extrai valores como "486.481,14D" nas colunas de
    Débito ou Crédito, onde o D/C não deveria estar (é apenas um número).
    Também pode acontecer em Saldo Anterior e Saldo Atual.

    Esta função percorre TODAS as colunas numéricas (Saldo Anterior,
    Débito, Crédito, Saldo Atual) e separa D/C que estiverem grudados.
    Para Débito e Crédito, o D/C é simplesmente removido (essas colunas
    são valores puros). Para Saldo Anterior e Saldo Atual, o D/C é
    preservado (será tratado por _split_natureza_columns depois).

    Exceção: valores "0", "0,00" etc. não precisam de verificação.

    Args:
        rows: Linhas unificadas (header + dados).

    Returns:
        Linhas com D/C corrigidos nas colunas numéricas.
    """
    if not rows or len(rows) < 2:
        return rows

    header = rows[0]
    layout = _detect_column_layout(header)
    col_debito = layout["debito"]
    col_credito = layout["credito"]

    # Regex para detectar valor brasileiro com D/C grudado
    # Ex: "486.481,14D", "1.796.997,71C", "123,45D"
    dc_pattern = re.compile(r"^([\d.,]+\s*)[DC]$", re.IGNORECASE)

    result = [header]
    fixes = 0

    for row in rows[1:]:
        new_row = list(row)

        for col_idx in (col_debito, col_credito):
            if col_idx >= len(new_row):
                continue

            val = new_row[col_idx].strip().replace("**", "")
            if not val:
                continue

            # Ignora zeros
            cleaned_zero = val.replace(".", "").replace(",", "").replace("0", "").rstrip("DC")
            if not cleaned_zero:
                continue

            # Verifica se tem D ou C no final
            if val.upper().endswith("D") or val.upper().endswith("C"):
                match = dc_pattern.match(val)
                if match:
                    # Remove o D/C — Débito e Crédito são valores puros
                    new_row[col_idx] = val[:-1].strip()
                    fixes += 1

        result.append(new_row)

    if fixes:
        logger.info(
            "Verificação D/C: %d valor(es) corrigido(s) nas colunas Débito/Crédito.",
            fixes,
        )

    return result


def _extract_natureza(value: str) -> tuple[str, str]:
    """Extrai natureza (D/C) de um valor brasileiro.

    Valores como "47.649.092,98D" → ("47.649.092,98", "D")
    Valores sem D/C como "60.917.957,40" → ("60.917.957,40", "")
    Valores zerados como "0,00" → ("0,00", "")

    Args:
        value: Valor com possível sufixo D/C.

    Returns:
        Tupla (valor_sem_natureza, natureza).
    """
    s = value.strip().replace("**", "")
    if not s:
        return "", ""
    if s.endswith("D"):
        return s[:-1], "D"
    if s.endswith("C"):
        return s[:-1], "C"
    return s, ""


def _split_natureza_columns(rows: list[list[str]]) -> list[list[str]]:
    """Separa D/C dos valores de Saldo Anterior e Saldo Atual em colunas próprias.

    Transforma:
        Código | Classificação | Descrição | Tipo | Saldo Anterior | Débito | Crédito | Saldo Atual
    Em:
        Código | Classificação | Descrição | Tipo | Saldo Anterior | Natureza SA | Débito | Crédito | Saldo Atual | Natureza SAT

    Args:
        rows: Linhas unificadas (header + dados).

    Returns:
        Linhas com colunas de natureza adicionadas.
    """
    if not rows:
        return rows

    header = rows[0]
    layout = _detect_column_layout(header)
    col_sa = layout["saldo_anterior"]
    col_sat = layout["saldo_atual"]

    # Verifica se realmente tem D/C nos dados (checa primeiras 10 linhas de dados)
    has_natureza = False
    for row in rows[1:11]:
        if len(row) > col_sat:
            _, nat_sa = _extract_natureza(row[col_sa])
            _, nat_sat = _extract_natureza(row[col_sat])
            if nat_sa or nat_sat:
                has_natureza = True
                break

    if not has_natureza:
        return rows  # Sem D/C, retorna sem modificar

    # Monta novo header com colunas de natureza
    # Inserimos "Natureza SA" DEPOIS de Saldo Anterior,
    # e "Natureza SAT" DEPOIS de Saldo Atual
    # Mas como Saldo Atual vem por último, fazemos na ordem certa
    result = []

    for i, row in enumerate(rows):
        if i == 0:
            # Header: insere colunas de natureza
            new_header = list(row[:col_sa])
            new_header.append(row[col_sa])          # Saldo Anterior
            new_header.append("Natureza SA")         # Nova coluna
            # Débito e Crédito ficam no meio
            for mid_col in range(col_sa + 1, col_sat):
                new_header.append(row[mid_col])
            new_header.append(row[col_sat])          # Saldo Atual
            new_header.append("Natureza SAT")        # Nova coluna
            # Colunas extras depois de Saldo Atual (se houver)
            for extra_col in range(col_sat + 1, len(row)):
                new_header.append(row[extra_col])
            result.append(new_header)
        else:
            # Dados
            if len(row) <= col_sat:
                result.append(row)
                continue

            val_sa, nat_sa = _extract_natureza(row[col_sa] if col_sa < len(row) else "")
            val_sat, nat_sat = _extract_natureza(row[col_sat] if col_sat < len(row) else "")

            new_row = list(row[:col_sa])
            new_row.append(val_sa)       # Saldo Anterior (sem D/C)
            new_row.append(nat_sa)       # Natureza SA
            for mid_col in range(col_sa + 1, col_sat):
                new_row.append(row[mid_col] if mid_col < len(row) else "")
            new_row.append(val_sat)      # Saldo Atual (sem D/C)
            new_row.append(nat_sat)      # Natureza SAT
            for extra_col in range(col_sat + 1, len(row)):
                new_row.append(row[extra_col])
            result.append(new_row)

    return result


def _postprocess_agrupadoras(rows: list[list[str]], layout: dict[str, int]) -> list[list[str]]:
    """Pós-processamento: valida e corrige coluna Tipo usando múltiplas heurísticas.

    Aplica 2 camadas de detecção (qualquer uma pode promover D→A):

    1. HIERARQUIA (prioritária, se classificação disponível): conta cujo
       classificação é prefixo de outra (ex: "3.2" é pai de "3.2.1") → A.

    2. SOMA NUMÉRICA (fallback, funciona sem classificação): se o Saldo Atual
       de uma conta ≈ soma dos Saldos Atuais das linhas consecutivas abaixo,
       essa conta é agrupadora. Checa também Débito e Crédito para confirmar
       (agrupadora totaliza tudo: SA, Deb, Cred e SAT).

    Args:
        rows: Linhas de dados (sem header e sem resumo).
        layout: Layout de colunas detectado.

    Returns:
        Linhas com Tipo corrigido.
    """
    if not layout["has_tipo"] or layout["tipo"] < 0:
        return rows

    tipo_idx = layout["tipo"]
    col_sa = layout["saldo_anterior"]
    col_deb = layout["debito"]
    col_cred = layout["credito"]
    col_sat = layout["saldo_atual"]
    min_cols = layout["min_cols"]

    def _parse_val(val_str: str) -> float:
        """Parseia valor brasileiro para float (ignora D/C)."""
        s = val_str.strip().replace("**", "")
        if not s:
            return 0.0
        if s[-1] in ("D", "C", "d", "c"):
            s = s[:-1]
        s = s.replace(".", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return 0.0

    # --- Camada 1: Hierarquia por classificação ---
    agrupadoras_hierarquia: set[int] = set()
    classificacoes = set()
    for row in rows:
        if len(row) > 1:
            c = row[1].strip()
            if c:
                classificacoes.add(c)

    if classificacoes:
        agrupadoras_class = set()
        for c in classificacoes:
            for other in classificacoes:
                if other != c and other.startswith(c + "."):
                    agrupadoras_class.add(c)
                    break

        for i, row in enumerate(rows):
            if len(row) > 1 and row[1].strip() in agrupadoras_class:
                agrupadoras_hierarquia.add(i)

    # --- Camada 2: Soma numérica (SA, Deb, Cred, SAT ≈ soma dos filhos) ---
    agrupadoras_soma: set[int] = set()

    for i, row in enumerate(rows):
        if len(row) < min_cols:
            continue
        # Já é A? Pula
        if row[tipo_idx].strip().upper() == "A":
            continue
        # Já detectada por hierarquia? Pula
        if i in agrupadoras_hierarquia:
            continue

        pai_sat = _parse_val(row[col_sat])
        pai_deb = _parse_val(row[col_deb])
        pai_cred = _parse_val(row[col_cred])
        pai_sa = _parse_val(row[col_sa])

        # Precisa ter pelo menos um valor não-zero para comparar
        if pai_sat == 0.0 and pai_deb == 0.0 and pai_cred == 0.0 and pai_sa == 0.0:
            continue

        # Soma os filhos consecutivos abaixo
        soma_sat = 0.0
        soma_deb = 0.0
        soma_cred = 0.0
        soma_sa = 0.0
        filhos_count = 0

        for j in range(i + 1, len(rows)):
            filho = rows[j]
            if len(filho) < min_cols:
                continue

            soma_sat += _parse_val(filho[col_sat])
            soma_deb += _parse_val(filho[col_deb])
            soma_cred += _parse_val(filho[col_cred])
            soma_sa += _parse_val(filho[col_sa])
            filhos_count += 1

            # Se Saldo Atual já atingiu o pai, para de somar
            if pai_sat != 0.0 and abs(soma_sat) >= abs(pai_sat) * 0.999:
                break

        if filhos_count < 2:
            continue

        # Verifica se pelo menos 2 das 4 colunas batem (tolerância 1%)
        matches = 0
        for pai_v, soma_v in [
            (pai_sat, soma_sat),
            (pai_deb, soma_deb),
            (pai_cred, soma_cred),
            (pai_sa, soma_sa),
        ]:
            if pai_v == 0.0 and soma_v == 0.0:
                matches += 1  # ambos zero = match
                continue
            if pai_v == 0.0 or soma_v == 0.0:
                continue  # um zero outro não = sem match
            diff = abs(soma_v - pai_v)
            tolerancia = max(abs(pai_v) * 0.01, 0.02)
            if diff <= tolerancia:
                matches += 1

        if matches >= 2:
            agrupadoras_soma.add(i)

    # --- Aplica correções ---
    all_agrupadoras = agrupadoras_hierarquia | agrupadoras_soma
    corrected = 0
    corrected_hierarquia = 0
    corrected_soma = 0

    for i in all_agrupadoras:
        row = rows[i]
        if len(row) <= tipo_idx:
            continue
        if row[tipo_idx].strip().upper() != "A":
            row[tipo_idx] = "A"
            corrected += 1
            if i in agrupadoras_hierarquia:
                corrected_hierarquia += 1
            else:
                corrected_soma += 1

    if corrected:
        parts = []
        if corrected_hierarquia:
            parts.append(f"{corrected_hierarquia} por hierarquia")
        if corrected_soma:
            parts.append(f"{corrected_soma} por soma numérica")
        logger.info(
            "Pós-processamento agrupadoras: %d corrigida(s) → A (%s).",
            corrected, ", ".join(parts),
        )

    return rows


def _parse_brazilian_value(value_str: str) -> tuple[float, str]:
    """Parseia valor em formato brasileiro com sufixo D/C.

    Args:
        value_str: Valor como string (ex: "4.960.556,92D").

    Returns:
        Tupla (valor_float, natureza) — ex: (4960556.92, "D").
    """
    s = value_str.strip().replace("**", "")
    if not s:
        return 0.0, ""

    natureza = ""
    if s.endswith("D"):
        natureza = "D"
        s = s[:-1]
    elif s.endswith("C"):
        natureza = "C"
        s = s[:-1]

    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s), natureza
    except ValueError:
        return 0.0, ""


def _format_brazilian_value(value: float, natureza: str) -> str:
    """Formata valor float de volta para formato brasileiro com D/C.

    Args:
        value: Valor absoluto.
        natureza: "D", "C" ou "".

    Returns:
        String formatada (ex: "334.509,61D").
    """
    formatted = f"{abs(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{formatted}{natureza}" if natureza else formatted


def _valor_com_sinal(valor: float, natureza: str, grupo: int) -> float:
    """Converte valor para formato com sinal baseado na convenção contábil.

    Balanço Patrimonial:
        Ativo (1): D = +valor, C = -valor
        Passivo (2): C = +valor, D = -valor

    DRE (Resultado):
        Custos/Despesas (3): C = +valor, D = -valor
        Receitas (4): C = +valor, D = -valor

    Args:
        valor: Valor absoluto.
        natureza: "D", "C" ou "".
        grupo: Primeiro dígito da classificação.

    Returns:
        Valor com sinal correto.
    """
    if valor == 0.0 or not natureza:
        return valor

    # Ativo: D = positivo, C = negativo
    if grupo == 1:
        return valor if natureza == "D" else -valor

    # Passivo, Custos/Despesas, Receitas: C = positivo, D = negativo
    return valor if natureza == "C" else -valor


def save_synthetic_csv(
    unified_rows: list[list[str]],
    keep_classificacoes: set[str],
    filename: str,
    output_dir: str | Path | None = None,
) -> tuple[Path, list[list[str]]]:
    """Gera CSV sintético filtrando contas e calculando resultado do período.

    Para contas de resultado (grupos 3 e 4), calcula:
        resultado = saldo_atual_com_sinal - saldo_anterior_com_sinal

    Args:
        unified_rows: Linhas do CSV completo (header + dados + resumo).
        keep_classificacoes: Set de classificações a manter.
        filename: Nome base do arquivo.
        output_dir: Diretório de saída.

    Returns:
        Tupla (Path do arquivo CSV sintético, linhas processadas).
    """
    out_dir = Path(output_dir) if output_dir else OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    safe_name = re.sub(r'[<>:"/\\|?*]', "_", filename)

    if not unified_rows:
        raise ValueError("unified_rows está vazio.")

    header = unified_rows[0]
    layout = _detect_column_layout(header)
    col_sa = layout["saldo_anterior"]
    col_sat = layout["saldo_atual"]
    min_cols = layout["min_cols"]

    data_rows = []
    resumo_rows = []

    for row in unified_rows[1:]:
        if len(row) < min_cols:
            continue
        codigo, classificacao = row[0].strip(), row[1].strip()
        # Resumo: Código e Classificação vazios
        if not codigo and not classificacao:
            resumo_rows.append(row)
        else:
            data_rows.append(row)

    # Pós-processamento: valida agrupadoras
    data_rows = _postprocess_agrupadoras(data_rows, layout)

    # Filtra: manter apenas contas cujo classificacao está no set
    filtered = []
    for row in data_rows:
        classificacao = row[1].strip()
        if classificacao in keep_classificacoes:
            filtered.append(row)

    # Transforma contas de resultado (grupos 3 e 4)
    result_rows = []
    for row in filtered:
        classificacao = row[1].strip()
        if not classificacao:
            result_rows.append(row)
            continue

        grupo = 0
        try:
            grupo = int(classificacao[0])
        except (ValueError, IndexError):
            pass

        if grupo in (3, 4):
            # Parseia saldo anterior e saldo atual
            val_ant, nat_ant = _parse_brazilian_value(row[col_sa])
            val_at, nat_at = _parse_brazilian_value(row[col_sat])

            # Converte para valores com sinal
            saldo_ant_signed = _valor_com_sinal(val_ant, nat_ant, grupo)
            saldo_at_signed = _valor_com_sinal(val_at, nat_at, grupo)

            # Resultado do período
            resultado = saldo_at_signed - saldo_ant_signed

            # Converte de volta para D/C (DRE: positivo=C, negativo=D)
            if resultado >= 0:
                nat_resultado = "C"
            else:
                nat_resultado = "D"

            # Monta nova linha: Saldo Anterior=0, Saldo Atual=resultado
            new_row = list(row)
            new_row[col_sa] = "0,00"  # Saldo Anterior zerado
            new_row[col_sat] = _format_brazilian_value(resultado, nat_resultado)
            result_rows.append(new_row)
        else:
            result_rows.append(row)

    # Monta CSV final: header + dados filtrados + resumo
    output_rows = [header] + result_rows + resumo_rows

    output_path = out_dir / f"{safe_name}_sintetico.csv"
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        for row in output_rows:
            writer.writerow(row)

    logger.info(
        "CSV sintético salvo: %s (%d linhas, de %d originais)",
        output_path, len(result_rows), len(data_rows),
    )
    return output_path, output_rows


def _format_signed_brazilian(value: float) -> str:
    """Formata valor float com sinal em formato brasileiro.

    Args:
        value: Valor com sinal (positivo ou negativo).

    Returns:
        String formatada (ex: "4.960.556,92" ou "-24.317,76").
    """
    formatted = f"{abs(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    if value < 0:
        return f"-{formatted}"
    return formatted


def _grupo_from_resumo(descricao: str) -> int:
    """Infere o grupo contábil a partir da descrição da linha de resumo.

    Args:
        descricao: Descrição da conta no resumo.

    Returns:
        Grupo (1-4) ou 0 se não identificado.
    """
    desc = descricao.strip().upper().replace("**", "")
    if desc.startswith("ATIVO") or desc == "CONTAS DEVEDORAS":
        return 1
    if desc.startswith("PASSIVO") or desc.startswith("PATRIMÔNIO") or desc == "CONTAS CREDORAS":
        return 2
    if "CUSTOS" in desc and "DESPESAS" in desc:
        return 3
    if "RECEITAS" in desc:
        return 4
    if "RESULTADO" in desc:
        return 4  # resultados seguem lógica DRE (C=+, D=-)
    return 0


def _convert_row_to_signed(row: list[str], grupo: int, col_sa: int, col_sat: int) -> list[str]:
    """Converte uma linha do CSV de D/C para +/-.

    Converte apenas Saldo Anterior e Saldo Atual.
    Débito e Crédito permanecem como estão.

    Args:
        row: Linha original do CSV.
        grupo: Grupo contábil (1-4).
        col_sa: Índice da coluna Saldo Anterior.
        col_sat: Índice da coluna Saldo Atual.

    Returns:
        Nova linha com valores em formato +/-.
    """
    new_row = list(row)
    if len(new_row) <= col_sat:
        return new_row

    for col_idx in (col_sa, col_sat):
        val, nat = _parse_brazilian_value(new_row[col_idx])
        if val == 0.0 and not nat:
            new_row[col_idx] = "0,00"
        elif grupo > 0:
            signed = _valor_com_sinal(val, nat, grupo)
            new_row[col_idx] = _format_signed_brazilian(signed)
        # Se grupo == 0, mantém original

    return new_row


def _convert_rows_to_signed(rows: list[list[str]]) -> list[list[str]]:
    """Converte linhas de D/C para +/- usando convenção contábil.

    Args:
        rows: Linhas do CSV (header + dados + resumo).

    Returns:
        Novas linhas com valores em formato +/-.
    """
    if not rows:
        return rows

    header = rows[0]
    layout = _detect_column_layout(header)
    col_sa = layout["saldo_anterior"]
    col_sat = layout["saldo_atual"]
    min_cols = layout["min_cols"]

    output_rows = [header]

    for row in rows[1:]:
        if len(row) < min_cols:
            output_rows.append(row)
            continue

        codigo, classificacao = row[0].strip(), row[1].strip()

        if not codigo and not classificacao:
            descricao = row[2].strip()
            grupo = _grupo_from_resumo(descricao)
            output_rows.append(_convert_row_to_signed(row, grupo, col_sa, col_sat))
        elif classificacao:
            try:
                grupo = int(classificacao[0])
            except (ValueError, IndexError):
                grupo = 0
            output_rows.append(_convert_row_to_signed(row, grupo, col_sa, col_sat))
        else:
            output_rows.append(row)

    return output_rows


def save_signed_csv(
    unified_rows: list[list[str]],
    filename: str,
    output_dir: str | Path | None = None,
    suffix: str = "_sinal",
) -> Path:
    """Gera CSV com +/- no lugar de D/C usando convenção contábil.

    Args:
        unified_rows: Linhas do CSV (header + dados + resumo).
        filename: Nome base do arquivo.
        output_dir: Diretório de saída.
        suffix: Sufixo do arquivo (padrão: "_sinal").

    Returns:
        Path do arquivo CSV gerado.
    """
    out_dir = Path(output_dir) if output_dir else OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    safe_name = re.sub(r'[<>:"/\\|?*]', "_", filename)

    if not unified_rows:
        raise ValueError("unified_rows está vazio.")

    output_rows = _convert_rows_to_signed(unified_rows)

    output_path = out_dir / f"{safe_name}{suffix}.csv"
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        for row in output_rows:
            writer.writerow(row)

    logger.info("CSV com sinal salvo: %s (%d linhas)", output_path, len(output_rows) - 1)
    return output_path


def tables_to_csv_string(tables: list[list[list[str]]], delimiter: str = ";") -> str:
    """Converte tabelas extraídas para uma string CSV.

    Args:
        tables: Lista de tabelas (saída de extract_markdown_tables).
        delimiter: Delimitador CSV.

    Returns:
        String formatada como CSV.
    """
    output = io.StringIO()
    writer = csv.writer(output, delimiter=delimiter)

    unified = _unify_and_deduplicate(tables)
    for row in unified:
        writer.writerow(row)

    return output.getvalue()
