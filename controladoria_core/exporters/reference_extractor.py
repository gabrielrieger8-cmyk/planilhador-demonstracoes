"""Extrator de referência para RAG — aprende o padrão do XLSX validado.

Analisa um XLSX profissional já validado pelo usuário e extrai:
1. Hierarquia completa de agrupadoras (qual agrupadora soma quais filhos)
2. Convenção de sinais por grupo contábil (com exemplos concretos)
3. Mapeamento de tipos (A/D) por classificação
4. Formato do plano de contas (níveis, nomenclatura)

Gera um arquivo-texto estruturado em `knowledge/` que é injetado
no prompt do Gemini para extrações futuras.
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from controladoria_core.exporters.hierarchy import build_hierarchy, get_account_group, get_direct_children
from controladoria_core.utils.config import logger


# ---------------------------------------------------------------------------
# Dataclass de referência
# ---------------------------------------------------------------------------

class ReferenceData:
    """Dados extraídos de um XLSX validado."""

    def __init__(self):
        self.empresa: str = ""
        self.periodo: str = ""
        self.total_contas: int = 0
        self.grupos: dict[int, GrupoInfo] = {}
        self.hierarchy_tree: list[HierarchyNode] = []
        self.sign_examples: list[SignExample] = []
        self.plano_contas: list[ContaInfo] = []
        self.user_instructions: str = ""

    def to_prompt_text(self) -> str:
        """Gera texto formatado para injeção no prompt do Gemini."""
        lines = []
        lines.append("=" * 70)
        lines.append("REFERÊNCIA DE PADRÃO CONTÁBIL (aprendido de XLSX validado)")
        lines.append(f"Empresa: {self.empresa} | Período: {self.periodo}")
        lines.append(f"Total de contas: {self.total_contas}")
        lines.append("=" * 70)
        lines.append("")

        # 1. Estrutura de grupos
        lines.append("1. GRUPOS CONTÁBEIS PRESENTES:")
        lines.append("-" * 40)
        for g_num in sorted(self.grupos.keys()):
            g = self.grupos[g_num]
            lines.append(
                f"   Grupo {g_num}: {g.nome} "
                f"({g.total_contas} contas, {g.total_agrupadoras} agrupadoras, "
                f"{g.total_detalhe} detalhe)"
            )
            if g.classificacao_raiz:
                lines.append(f"   Classificação raiz: {g.classificacao_raiz}")
        lines.append("")

        # 2. Hierarquia de agrupadoras — SEÇÃO CRÍTICA
        lines.append("2. HIERARQUIA DE AGRUPADORAS (qual agrupadora soma quais contas):")
        lines.append("-" * 40)
        lines.append("   ATENÇÃO: Cada agrupadora (Tipo=A) deve ser a SOMA de seus filhos diretos.")
        lines.append("   Os filhos são as contas imediatamente abaixo na hierarquia de classificação.")
        lines.append("   Exemplo: se a classificação pai é '01.1', seus filhos diretos são '01.1.XX'")
        lines.append("   onde XX é um único segmento (sem mais pontos).")
        lines.append("")

        for node in self.hierarchy_tree:
            indent = "   " * (node.nivel + 1)
            child_classifs = ", ".join(node.filhos_classif)
            child_tipos = ", ".join(
                f"{c}({'A' if t == 'A' else 'D'})"
                for c, t in zip(node.filhos_classif, node.filhos_tipos)
            )

            lines.append(
                f"{indent}[A] {node.classificacao} — {node.descricao}"
            )
            lines.append(
                f"{indent}    SOMA DE → [{child_tipos}]"
            )

            # Se algum filho é agrupadora, explica a cadeia
            sub_agrupadoras = [
                c for c, t in zip(node.filhos_classif, node.filhos_tipos)
                if t == "A"
            ]
            if sub_agrupadoras:
                lines.append(
                    f"{indent}    ↳ Filhos que também são agrupadoras: {', '.join(sub_agrupadoras)}"
                )
                lines.append(
                    f"{indent}      (esses filhos por sua vez somam SEUS próprios filhos)"
                )
        lines.append("")

        # 3. Convenção de sinais
        lines.append("3. CONVENÇÃO DE SINAIS (D/C → +/-):")
        lines.append("-" * 40)
        lines.append("   Cada valor no PDF vem com sufixo D (Débito) ou C (Crédito).")
        lines.append("   A regra de conversão para +/- depende do GRUPO da conta:")
        lines.append("")

        # Agrupa exemplos por grupo
        examples_by_group: dict[int, list[SignExample]] = {}
        for ex in self.sign_examples:
            examples_by_group.setdefault(ex.grupo, []).append(ex)

        for g_num in sorted(examples_by_group.keys()):
            g_info = self.grupos.get(g_num)
            g_nome = g_info.nome if g_info else f"Grupo {g_num}"
            lines.append(f"   Grupo {g_num} ({g_nome}):")

            exs = examples_by_group[g_num]
            # Mostra regra geral
            d_pos = sum(1 for e in exs if e.natureza == "D" and e.sinal == "+")
            d_neg = sum(1 for e in exs if e.natureza == "D" and e.sinal == "-")
            c_pos = sum(1 for e in exs if e.natureza == "C" and e.sinal == "+")
            c_neg = sum(1 for e in exs if e.natureza == "C" and e.sinal == "-")

            if d_pos > d_neg:
                lines.append(f"     D → positivo (+)  |  C → negativo (-)")
            else:
                lines.append(f"     D → negativo (-)  |  C → positivo (+)")

            # Exemplos concretos (até 3 por grupo)
            shown = 0
            for ex in exs[:3]:
                lines.append(
                    f"     Ex: {ex.classificacao} {ex.descricao[:30]} "
                    f"Nat={ex.natureza} → valor {ex.sinal}{ex.valor_abs}"
                )
                shown += 1
            if len(exs) > 3:
                lines.append(f"     ... e mais {len(exs) - 3} contas")
            lines.append("")

        # 4. Plano de contas resumido (classificações + tipos)
        lines.append("4. PLANO DE CONTAS (classificações e tipos detectados):")
        lines.append("-" * 40)
        lines.append("   Use esta referência para classificar Tipo=A ou Tipo=D:")
        lines.append("")

        for conta in self.plano_contas:
            indent = "   " + "  " * (conta.nivel - 1)
            tipo_marker = "[A]" if conta.tipo == "A" else "[D]"
            lines.append(
                f"{indent}{tipo_marker} {conta.classificacao} — {conta.descricao}"
            )

        # 5. Instruções do usuário (correções e problemas identificados)
        if self.user_instructions:
            lines.append("")
            lines.append("5. INSTRUÇÕES DO CONTROLLER (PRIORIDADE MÁXIMA):")
            lines.append("-" * 40)
            lines.append("   O controller humano revisou a extração e identificou os seguintes")
            lines.append("   problemas/correções. SIGA ESTAS INSTRUÇÕES nas próximas extrações:")
            lines.append("")
            for instruction_line in self.user_instructions.strip().splitlines():
                lines.append(f"   >>> {instruction_line.strip()}")
            lines.append("")
            lines.append("   ATENÇÃO: Estas instruções vêm do controller que validou o balancete.")
            lines.append("   Elas têm PRIORIDADE sobre qualquer heurística automática.")

        lines.append("")
        lines.append("=" * 70)
        lines.append("FIM DA REFERÊNCIA")
        lines.append("=" * 70)

        return "\n".join(lines)

    def to_json(self) -> dict:
        """Serializa para JSON (para persistência)."""
        return {
            "empresa": self.empresa,
            "periodo": self.periodo,
            "total_contas": self.total_contas,
            "created_at": datetime.now().isoformat(),
            "grupos": {
                str(k): {
                    "nome": v.nome,
                    "classificacao_raiz": v.classificacao_raiz,
                    "total_contas": v.total_contas,
                    "total_agrupadoras": v.total_agrupadoras,
                    "total_detalhe": v.total_detalhe,
                }
                for k, v in self.grupos.items()
            },
            "hierarchy": [
                {
                    "classificacao": n.classificacao,
                    "descricao": n.descricao,
                    "nivel": n.nivel,
                    "filhos_classif": n.filhos_classif,
                    "filhos_tipos": n.filhos_tipos,
                }
                for n in self.hierarchy_tree
            ],
            "sign_examples": [
                {
                    "grupo": e.grupo,
                    "classificacao": e.classificacao,
                    "descricao": e.descricao,
                    "natureza": e.natureza,
                    "sinal": e.sinal,
                    "valor_abs": e.valor_abs,
                }
                for e in self.sign_examples
            ],
            "plano_contas": [
                {
                    "classificacao": c.classificacao,
                    "descricao": c.descricao,
                    "tipo": c.tipo,
                    "nivel": c.nivel,
                }
                for c in self.plano_contas
            ],
            "user_instructions": self.user_instructions,
        }


class GrupoInfo:
    """Info de um grupo contábil."""

    def __init__(self):
        self.nome: str = ""
        self.classificacao_raiz: str = ""
        self.total_contas: int = 0
        self.total_agrupadoras: int = 0
        self.total_detalhe: int = 0


class HierarchyNode:
    """Nó na árvore de hierarquia."""

    def __init__(self):
        self.classificacao: str = ""
        self.descricao: str = ""
        self.nivel: int = 0
        self.filhos_classif: list[str] = []
        self.filhos_tipos: list[str] = []


class SignExample:
    """Exemplo concreto de conversão de sinal."""

    def __init__(self):
        self.grupo: int = 0
        self.classificacao: str = ""
        self.descricao: str = ""
        self.natureza: str = ""  # D ou C
        self.sinal: str = ""     # + ou -
        self.valor_abs: str = ""


class ContaInfo:
    """Info de uma conta no plano."""

    def __init__(self):
        self.classificacao: str = ""
        self.descricao: str = ""
        self.tipo: str = ""  # A ou D
        self.nivel: int = 1


# ---------------------------------------------------------------------------
# Extração principal
# ---------------------------------------------------------------------------

# Mapeamento de coluna no Excel (0-based nos dados, 1-based no Excel)
# Segue ALL_COLUMNS do xlsx_builder:
# 0=Código, 1=Classificação, 2=Descrição, 3=SA, 4=NatSA,
# 5=Débito, 6=Crédito, 7=SAT, 8=NatSAT, 9=Tipo

_GRUPO_NOMES = {
    1: "ATIVO",
    2: "PASSIVO",
    3: "CUSTOS E DESPESAS",
    4: "RECEITAS",
    5: "CUSTOS",
    6: "RECEITAS (6)",
}


def extract_reference_from_xlsx(
    xlsx_path: str | Path,
    sheet_name: str | None = None,
) -> ReferenceData:
    """Extrai padrão de referência de um XLSX validado.

    Lê a primeira aba (ou a aba especificada) e analisa:
    - Hierarquia de classificações
    - Convenção de sinais
    - Tipos A/D
    - Plano de contas

    Args:
        xlsx_path: Caminho do XLSX.
        sheet_name: Nome da aba (None = primeira disponível).

    Returns:
        ReferenceData com toda a análise.
    """
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"XLSX não encontrado: {path}")

    wb = load_workbook(str(path), read_only=True, data_only=True)

    # Seleciona aba
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # Pega primeira aba que não seja _Config
        target = None
        for name in wb.sheetnames:
            if not name.startswith("_"):
                target = name
                break
        if not target:
            target = wb.sheetnames[0]
        ws = wb[target]
        sheet_name = target

    logger.info("Extraindo referência de: %s (aba: %s)", path.name, sheet_name)

    # Lê todas as linhas
    rows_raw = []
    for row in ws.iter_rows(values_only=True):
        rows_raw.append([str(cell) if cell is not None else "" for cell in row])

    wb.close()

    if len(rows_raw) < 2:
        raise ValueError("XLSX com menos de 2 linhas (header + dados).")

    header = rows_raw[0]
    data_rows = rows_raw[1:]

    # Detecta índices das colunas pelo header
    col_map = _detect_columns(header)

    ref = ReferenceData()
    ref.empresa = _extract_empresa_from_name(path.name)
    ref.periodo = sheet_name
    ref.total_contas = len(data_rows)

    # Converte para formato interno (lista de listas como _rows)
    internal_rows = []
    for row in data_rows:
        mapped = _map_xlsx_row(row, col_map)
        internal_rows.append(mapped)

    # 1. Analisa grupos
    _analyze_grupos(ref, internal_rows)

    # 2. Constrói hierarquia
    _build_hierarchy_tree(ref, internal_rows)

    # 3. Extrai exemplos de sinais
    _extract_sign_examples(ref, internal_rows)

    # 4. Constrói plano de contas
    _build_plano_contas(ref, internal_rows)

    logger.info(
        "Referência extraída: %d contas, %d grupos, %d nós hierárquicos",
        ref.total_contas, len(ref.grupos), len(ref.hierarchy_tree),
    )

    return ref


def _detect_columns(header: list[str]) -> dict[str, int | None]:
    """Detecta posição das colunas no header do XLSX."""
    mapping = {
        "codigo": None,
        "classificacao": None,
        "descricao": None,
        "saldo_anterior": None,
        "nat_sa": None,
        "debito": None,
        "credito": None,
        "saldo_atual": None,
        "nat_sat": None,
        "tipo": None,
    }

    for i, h in enumerate(header):
        hl = h.strip().lower()
        if "c" == hl[:1] and ("dig" in hl or "ódigo" in hl or "odigo" in hl):
            mapping["codigo"] = i
        elif "classifica" in hl:
            mapping["classificacao"] = i
        elif "descri" in hl:
            mapping["descricao"] = i
        elif hl == "tipo" or hl == "type":
            mapping["tipo"] = i
        elif "saldo" in hl and "ant" in hl:
            mapping["saldo_anterior"] = i
        elif "natureza" in hl and "sa" in hl and "sat" not in hl:
            mapping["nat_sa"] = i
        elif "natureza" in hl and "sat" in hl:
            mapping["nat_sat"] = i
        elif "d" in hl and ("bit" in hl or "ébit" in hl or "ebito" in hl):
            mapping["debito"] = i
        elif "cr" in hl and ("dit" in hl or "édit" in hl or "edito" in hl):
            mapping["credito"] = i
        elif "saldo" in hl and "at" in hl:
            mapping["saldo_atual"] = i

    return mapping


def _map_xlsx_row(row: list[str], col_map: dict[str, int | None]) -> list[str]:
    """Mapeia uma linha XLSX para formato padronizado [10 campos]."""
    def _get(key: str) -> str:
        idx = col_map[key]
        if idx is None or idx >= len(row):
            return ""
        return row[idx].strip()

    return [
        _get("codigo"),           # 0
        _get("classificacao"),    # 1
        _get("descricao"),        # 2
        _get("saldo_anterior"),   # 3
        _get("nat_sa"),           # 4
        _get("debito"),           # 5
        _get("credito"),          # 6
        _get("saldo_atual"),      # 7
        _get("nat_sat"),          # 8
        _get("tipo"),             # 9
    ]


def _extract_empresa_from_name(filename: str) -> str:
    """Extrai nome da empresa do filename."""
    # "STV_Balancete 07.2025.xlsx" → "STV"
    # "Balancetes_Profissional.xlsx" → ""
    name = filename.replace(".xlsx", "").replace(".XLSX", "")
    if "_" in name:
        parts = name.split("_")
        # Se primeiro segmento é curto e parece nome de empresa
        if len(parts[0]) <= 10 and parts[0].isalpha():
            return parts[0]
    return name


def _analyze_grupos(ref: ReferenceData, rows: list[list[str]]) -> None:
    """Analisa grupos contábeis presentes."""
    for row in rows:
        classif = row[1]
        tipo = row[9].upper() if row[9] else ""
        descricao = row[2]
        grupo = get_account_group(classif)

        if grupo == 0:
            continue

        if grupo not in ref.grupos:
            info = GrupoInfo()
            info.nome = _GRUPO_NOMES.get(grupo, f"Grupo {grupo}")
            ref.grupos[grupo] = info

        g = ref.grupos[grupo]
        g.total_contas += 1

        if tipo == "A":
            g.total_agrupadoras += 1
        else:
            g.total_detalhe += 1

        # Classificação raiz = a mais curta do grupo
        if not g.classificacao_raiz or len(classif) < len(g.classificacao_raiz):
            g.classificacao_raiz = classif
            if tipo == "A" and descricao:
                g.nome = descricao.strip()


def _build_hierarchy_tree(ref: ReferenceData, rows: list[list[str]]) -> None:
    """Constrói árvore hierárquica de agrupadoras."""
    # Usa build_hierarchy do módulo hierarchy.py
    hierarchy = build_hierarchy(rows, classif_col=1, tipo_col=9)

    # Coleta todas as entries para get_direct_children
    entries: list[tuple[int, str]] = []
    for i, row in enumerate(rows):
        classif = row[1].strip()
        if classif:
            entries.append((i, classif))

    # Ordena por classificação para apresentação
    sorted_parents = sorted(
        hierarchy.keys(),
        key=lambda idx: rows[idx][1],
    )

    for parent_idx in sorted_parents:
        children_indices = hierarchy[parent_idx]
        parent_row = rows[parent_idx]

        node = HierarchyNode()
        node.classificacao = parent_row[1]
        node.descricao = parent_row[2]
        node.nivel = parent_row[1].count(".")

        for child_idx in children_indices:
            child_row = rows[child_idx]
            node.filhos_classif.append(child_row[1])
            node.filhos_tipos.append(
                child_row[9].upper() if child_row[9] else "D"
            )

        ref.hierarchy_tree.append(node)


def _extract_sign_examples(ref: ReferenceData, rows: list[list[str]]) -> None:
    """Extrai exemplos concretos de sinais por grupo."""
    # Para cada grupo, pega contas de detalhe (D) com valor e natureza
    examples_per_group: dict[int, list[SignExample]] = {}

    for row in rows:
        classif = row[1]
        descricao = row[2]
        tipo = row[9].upper() if row[9] else ""
        nat_sat = row[8].upper() if row[8] else ""
        sat_str = row[7]
        grupo = get_account_group(classif)

        if grupo == 0 or tipo != "D":
            continue

        if nat_sat not in ("D", "C"):
            continue

        # Parse valor
        val_str = sat_str.strip()
        if not val_str or val_str == "0" or val_str == "0.0" or val_str == "None":
            continue

        # Determina sinal do valor
        try:
            val = float(val_str.replace(".", "").replace(",", ".").replace("-", ""))
        except (ValueError, AttributeError):
            continue

        if val == 0:
            continue

        # O sinal no XLSX já reflete a conversão aplicada
        is_negative = val_str.strip().startswith("-")

        ex = SignExample()
        ex.grupo = grupo
        ex.classificacao = classif
        ex.descricao = descricao
        ex.natureza = nat_sat
        ex.sinal = "-" if is_negative else "+"
        ex.valor_abs = f"{val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

        examples_per_group.setdefault(grupo, []).append(ex)

    # Seleciona até 5 exemplos por grupo
    for grupo, exs in examples_per_group.items():
        ref.sign_examples.extend(exs[:5])


def _build_plano_contas(ref: ReferenceData, rows: list[list[str]]) -> None:
    """Constrói plano de contas ordenado."""
    for row in rows:
        classif = row[1].strip()
        descricao = row[2].strip()
        tipo = row[9].upper() if row[9] else "D"

        if not classif:
            continue

        conta = ContaInfo()
        conta.classificacao = classif
        conta.descricao = descricao
        conta.tipo = tipo
        conta.nivel = classif.count(".") + 1

        ref.plano_contas.append(conta)


# ---------------------------------------------------------------------------
# Persistência
# ---------------------------------------------------------------------------

def _get_knowledge_dir() -> Path:
    """Retorna diretório de conhecimento configurado."""
    from controladoria_core.utils.config import KNOWLEDGE_DIR as _kdir
    if _kdir is None:
        raise RuntimeError(
            "Core não configurado. Chame controladoria_core.utils.config.configure() primeiro."
        )
    return _kdir


def save_reference(
    ref: ReferenceData,
    knowledge_dir: Path | None = None,
    user_instructions: str = "",
    name: str = "",
) -> tuple[Path, Path]:
    """Salva referência em formato texto (prompt) e JSON.

    Args:
        ref: Dados extraídos.
        knowledge_dir: Diretório de destino (default: knowledge/).
        user_instructions: Instruções/correções do controller.
        name: Nome customizado definido pelo usuário (usado como filename).

    Returns:
        Tupla (path_texto, path_json).
    """
    if user_instructions:
        ref.user_instructions = user_instructions

    kdir = knowledge_dir or _get_knowledge_dir()
    kdir.mkdir(parents=True, exist_ok=True)

    # Nome base: usa nome customizado ou fallback empresa_periodo
    if name:
        # Sanitiza nome para filesystem
        import re
        base_name = re.sub(r'[^\w\s\-]', '', name).strip().replace(" ", "_")[:40]
        if not base_name:
            base_name = "ref"
    else:
        empresa_safe = ref.empresa.replace(" ", "_")[:20] if ref.empresa else "ref"
        periodo_safe = ref.periodo.replace(".", "_").replace(" ", "_")[:20] if ref.periodo else "geral"
        base_name = f"{empresa_safe}_{periodo_safe}"

    # Guarda display_name no JSON
    json_data = ref.to_json()
    json_data["display_name"] = name or base_name

    # Texto para prompt
    txt_path = kdir / f"{base_name}.txt"
    txt_path.write_text(ref.to_prompt_text(), encoding="utf-8")

    # JSON para persistência e debug
    json_path = kdir / f"{base_name}.json"
    json_path.write_text(
        json.dumps(json_data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    logger.info("Referência salva: %s (.txt + .json) display_name=%s", base_name, name or base_name)
    return txt_path, json_path


def load_reference_for_prompt(
    knowledge_dir: Path | None = None,
    reference_name: str | None = None,
) -> str | None:
    """Carrega referência como texto para injeção no prompt.

    Se `reference_name` é fornecido, carrega a referência específica (filename stem).
    Caso contrário, carrega a mais recente.

    Args:
        knowledge_dir: Diretório de conhecimento (default: knowledge/).
        reference_name: Nome (stem) do arquivo de referência a carregar.

    Returns:
        Texto da referência, ou None se não houver.
    """
    kdir = knowledge_dir or _get_knowledge_dir()
    if not kdir.exists():
        return None

    if reference_name:
        # Carrega referência específica
        txt_path = kdir / f"{reference_name}.txt"
        if txt_path.exists():
            ref_text = txt_path.read_text(encoding="utf-8")
            logger.info("Referência carregada (específica): %s (%d chars)", txt_path.name, len(ref_text))
            return ref_text
        # Fallback: procura JSON com display_name correspondente
        for json_file in kdir.glob("*.json"):
            try:
                data = json.loads(json_file.read_text(encoding="utf-8"))
                if data.get("display_name", "") == reference_name:
                    txt_candidate = json_file.with_suffix(".txt")
                    if txt_candidate.exists():
                        ref_text = txt_candidate.read_text(encoding="utf-8")
                        logger.info("Referência carregada (por display_name): %s (%d chars)", txt_candidate.name, len(ref_text))
                        return ref_text
            except Exception:
                continue
        logger.warning("Referência não encontrada: %s", reference_name)
        return None

    # Sem nome específico: retorna a mais recente
    txt_files = sorted(
        kdir.glob("*.txt"),
        key=lambda f: f.stat().st_mtime,
        reverse=True,
    )

    if not txt_files:
        return None

    ref_text = txt_files[0].read_text(encoding="utf-8")
    logger.info("Referência carregada (mais recente): %s (%d chars)", txt_files[0].name, len(ref_text))
    return ref_text


def list_references(
    knowledge_dir: Path | None = None,
) -> list[dict]:
    """Lista referências disponíveis.

    Returns:
        Lista de dicts com info de cada referência.
    """
    kdir = knowledge_dir or _get_knowledge_dir()
    if not kdir.exists():
        return []

    refs = []
    for json_file in sorted(kdir.glob("*.json"), key=lambda f: f.stat().st_mtime, reverse=True):
        try:
            data = json.loads(json_file.read_text(encoding="utf-8"))
            refs.append({
                "filename": json_file.stem,
                "display_name": data.get("display_name", json_file.stem),
                "empresa": data.get("empresa", ""),
                "periodo": data.get("periodo", ""),
                "total_contas": data.get("total_contas", 0),
                "created_at": data.get("created_at", ""),
                "grupos": len(data.get("grupos", {})),
                "has_instructions": bool(data.get("user_instructions", "")),
            })
        except Exception:
            continue

    return refs
