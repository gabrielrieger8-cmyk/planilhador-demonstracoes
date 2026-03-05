"""Testes unitários para o serviço de exportação Excel/CSV."""

import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.services.exporter import export_excel, export_excel_multi, export_csv


# ---------------------------------------------------------------------------
# Dados de teste
# ---------------------------------------------------------------------------

BALANCETE_DATA = {
    "empresa": "Empresa Teste Ltda",
    "cnpj": "12.345.678/0001-99",
    "periodo": "02/2026",
    "moeda": "BRL",
    "contas": [
        {
            "codigo_conta": "1",
            "descricao": "ATIVO",
            "nivel": 1,
            "natureza": "D",
            "saldo_anterior": 0,
            "debitos": 50000,
            "creditos": 30000,
            "saldo_atual": 120000,
            "is_totalizador": True,
        },
        {
            "codigo_conta": "1.1",
            "descricao": "Ativo Circulante",
            "nivel": 2,
            "natureza": "D",
            "saldo_anterior": 0,
            "debitos": 50000,
            "creditos": 30000,
            "saldo_atual": 80000,
            "is_totalizador": True,
        },
        {
            "codigo_conta": "1.1.01",
            "descricao": "Caixa",
            "nivel": 3,
            "natureza": "D",
            "saldo_anterior": 10000,
            "debitos": 30000,
            "creditos": 20000,
            "saldo_atual": 20000,
            "is_totalizador": False,
        },
        {
            "codigo_conta": "1.1.02",
            "descricao": "Bancos",
            "nivel": 3,
            "natureza": "D",
            "saldo_anterior": 50000,
            "debitos": 20000,
            "creditos": 10000,
            "saldo_atual": 60000,
            "is_totalizador": False,
        },
    ],
    "totais": {
        "total_debitos": 50000,
        "total_creditos": 30000,
    },
}

DRE_DATA = {
    "empresa": "Empresa Teste Ltda",
    "periodo": "01/2026 a 12/2026",
    "linhas": [
        {"descricao": "RECEITA OPERACIONAL BRUTA", "valor": 500000, "nivel": 1, "is_subtotal": False},
        {"descricao": "(-) Deduções da Receita", "valor": -50000, "nivel": 2, "is_subtotal": False},
        {"descricao": "RECEITA OPERACIONAL LÍQUIDA", "valor": 450000, "nivel": 1, "is_subtotal": True},
        {"descricao": "(-) CMV", "valor": -200000, "nivel": 2, "is_subtotal": False},
        {"descricao": "RESULTADO BRUTO", "valor": 250000, "nivel": 1, "is_subtotal": True},
    ],
    "resultado_liquido": 250000,
}

BALANCO_DATA = {
    "empresa": "Empresa Teste Ltda",
    "data_referencia": "31/12/2025",
    "ativo": {
        "circulante": {
            "total": 200000,
            "contas": [
                {"descricao": "Caixa", "valor": 50000, "nivel": 3},
                {"descricao": "Bancos", "valor": 150000, "nivel": 3},
            ],
        },
        "nao_circulante": {
            "total": 300000,
            "contas": [
                {"descricao": "Imobilizado", "valor": 300000, "nivel": 3},
            ],
        },
        "total": 500000,
    },
    "passivo": {
        "circulante": {
            "total": 150000,
            "contas": [
                {"descricao": "Fornecedores", "valor": 150000, "nivel": 3},
            ],
        },
        "nao_circulante": {
            "total": 100000,
            "contas": [
                {"descricao": "Empréstimos LP", "valor": 100000, "nivel": 3},
            ],
        },
        "total": 250000,
    },
    "patrimonio_liquido": {
        "total": 250000,
        "contas": [
            {"descricao": "Capital Social", "valor": 200000, "nivel": 3},
            {"descricao": "Lucros Acumulados", "valor": 50000, "nivel": 3},
        ],
    },
}


# ---------------------------------------------------------------------------
# Testes — Balancete
# ---------------------------------------------------------------------------

class TestBalanceteExporter:
    def test_xlsx_cria_arquivo(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "balancete.xlsx"
            result = export_excel(BALANCETE_DATA, "balancete", path)
            assert result.exists()
            assert result.stat().st_size > 0

    def test_xlsx_conteudo(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "balancete.xlsx"
            export_excel(BALANCETE_DATA, "balancete", path)

            wb = load_workbook(str(path))
            ws = wb.active

            assert "Empresa Teste" in str(ws["A1"].value)
            assert "Balancete" in str(ws["A1"].value)
            assert ws.max_row >= 7

    def test_csv_cria_arquivo(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "balancete.csv"
            result = export_csv(BALANCETE_DATA, "balancete", path)
            assert result.exists()

            content = path.read_text(encoding="utf-8-sig")
            assert "Código" in content
            assert "Caixa" in content
            assert ";" in content


# ---------------------------------------------------------------------------
# Testes — DRE
# ---------------------------------------------------------------------------

class TestDREExporter:
    def test_xlsx_cria_arquivo(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "dre.xlsx"
            result = export_excel(DRE_DATA, "dre", path)
            assert result.exists()

    def test_xlsx_conteudo(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "dre.xlsx"
            export_excel(DRE_DATA, "dre", path)

            wb = load_workbook(str(path))
            ws = wb.active
            assert "DRE" in ws.title

    def test_csv_cria_arquivo(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "dre.csv"
            result = export_csv(DRE_DATA, "dre", path)
            assert result.exists()

            content = path.read_text(encoding="utf-8-sig")
            assert "RECEITA" in content


# ---------------------------------------------------------------------------
# Testes — Balanço Patrimonial
# ---------------------------------------------------------------------------

class TestBalancoExporter:
    def test_xlsx_cria_arquivo(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "balanco.xlsx"
            result = export_excel(BALANCO_DATA, "balanco_patrimonial", path)
            assert result.exists()

    def test_xlsx_validacao_presente(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "balanco.xlsx"
            export_excel(BALANCO_DATA, "balanco_patrimonial", path)

            wb = load_workbook(str(path))
            ws = wb.active

            found_validation = False
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell and "VALIDAÇÃO" in str(cell):
                        found_validation = True
                        break
            assert found_validation

    def test_csv_cria_arquivo(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "balanco.csv"
            result = export_csv(BALANCO_DATA, "balanco_patrimonial", path)
            assert result.exists()

            content = path.read_text(encoding="utf-8-sig")
            assert "Ativo" in content
            assert "Fornecedores" in content


# ---------------------------------------------------------------------------
# Testes — Multi-demonstração
# ---------------------------------------------------------------------------

class TestMultiDemonstracaoExporter:
    def test_multi_abas_criadas(self):
        demonstracoes = [
            {"tipo": "balanco_patrimonial", "periodo": "31/12/2025", "dados": BALANCO_DATA},
            {"tipo": "dre", "periodo": "01/2026 a 12/2026", "dados": DRE_DATA},
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "multi.xlsx"
            result = export_excel_multi(demonstracoes, "Empresa Teste Ltda", path)
            assert result.exists()

            wb = load_workbook(str(path))
            assert len(wb.sheetnames) == 2

    def test_multi_abas_nomes(self):
        demonstracoes = [
            {"tipo": "balancete", "periodo": "02/2026", "dados": BALANCETE_DATA},
            {"tipo": "dre", "periodo": "01/2026 a 12/2026", "dados": DRE_DATA},
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "multi.xlsx"
            export_excel_multi(demonstracoes, "Empresa Teste", path)

            wb = load_workbook(str(path))
            assert len(wb.sheetnames) == 2
            all_names = " ".join(wb.sheetnames)
            assert "Balancete" in all_names
            assert "DRE" in all_names

    def test_titulo_padronizado(self):
        demonstracoes = [
            {"tipo": "dre", "periodo": "01/2026 a 12/2026", "dados": DRE_DATA},
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "titulo.xlsx"
            export_excel_multi(demonstracoes, "Empresa Teste Ltda", path)

            wb = load_workbook(str(path))
            ws = wb.active
            titulo = str(ws["A1"].value)
            assert "Empresa Teste Ltda" in titulo
            assert "DRE" in titulo
            assert "01/2026 a 12/2026" in titulo

    def test_tres_demonstracoes(self):
        demonstracoes = [
            {"tipo": "balancete", "periodo": "12/2025", "dados": BALANCETE_DATA},
            {"tipo": "balanco_patrimonial", "periodo": "31/12/2025", "dados": BALANCO_DATA},
            {"tipo": "dre", "periodo": "01/2025 a 12/2025", "dados": DRE_DATA},
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "tres.xlsx"
            result = export_excel_multi(demonstracoes, "XYZ Ltda", path)
            assert result.exists()

            wb = load_workbook(str(path))
            assert len(wb.sheetnames) == 3


class TestComparativoExporter:
    """Testes para exportação comparativa (multi-período lado a lado)."""

    DRE_P1 = {
        "empresa": "Teste", "periodo": "DEZ/2024",
        "linhas": [
            {"descricao": "RECEITA BRUTA", "valor": 200000, "nivel": 1, "is_subtotal": False},
            {"descricao": "(-) Deduções", "valor": -20000, "nivel": 2, "is_subtotal": False},
            {"descricao": "RECEITA LÍQUIDA", "valor": 180000, "nivel": 1, "is_subtotal": True},
        ],
        "resultado_liquido": 180000,
    }
    DRE_P2 = {
        "empresa": "Teste", "periodo": "DEZ/2025",
        "linhas": [
            {"descricao": "RECEITA BRUTA", "valor": 250000, "nivel": 1, "is_subtotal": False},
            {"descricao": "(-) Deduções", "valor": -25000, "nivel": 2, "is_subtotal": False},
            {"descricao": "RECEITA LÍQUIDA", "valor": 225000, "nivel": 1, "is_subtotal": True},
        ],
        "resultado_liquido": 225000,
    }
    BAL_P1 = {
        "empresa": "Teste", "data_referencia": "DEZ/2024",
        "ativo": {
            "circulante": {"total": 100000, "contas": [
                {"descricao": "Caixa", "valor": 50000, "nivel": 3},
                {"descricao": "Clientes", "valor": 50000, "nivel": 3},
            ]},
            "nao_circulante": {"total": 200000, "contas": [
                {"descricao": "Imobilizado", "valor": 200000, "nivel": 3},
            ]},
            "total": 300000,
        },
        "passivo": {
            "circulante": {"total": 80000, "contas": [
                {"descricao": "Fornecedores", "valor": 80000, "nivel": 3},
            ]},
            "nao_circulante": {"total": 70000, "contas": [
                {"descricao": "Empréstimos LP", "valor": 70000, "nivel": 3},
            ]},
            "total": 150000,
        },
        "patrimonio_liquido": {"total": 150000, "contas": [
            {"descricao": "Capital Social", "valor": 200000, "nivel": 3},
            {"descricao": "Prejuízos Acumulados", "valor": -50000, "nivel": 3},
        ]},
    }
    BAL_P2 = {
        "empresa": "Teste", "data_referencia": "DEZ/2025",
        "ativo": {
            "circulante": {"total": 120000, "contas": [
                {"descricao": "Caixa", "valor": 60000, "nivel": 3},
                {"descricao": "Clientes", "valor": 60000, "nivel": 3},
            ]},
            "nao_circulante": {"total": 230000, "contas": [
                {"descricao": "Imobilizado", "valor": 230000, "nivel": 3},
            ]},
            "total": 350000,
        },
        "passivo": {
            "circulante": {"total": 90000, "contas": [
                {"descricao": "Fornecedores", "valor": 90000, "nivel": 3},
            ]},
            "nao_circulante": {"total": 80000, "contas": [
                {"descricao": "Empréstimos LP", "valor": 80000, "nivel": 3},
            ]},
            "total": 170000,
        },
        "patrimonio_liquido": {"total": 180000, "contas": [
            {"descricao": "Capital Social", "valor": 200000, "nivel": 3},
            {"descricao": "Prejuízos Acumulados", "valor": -20000, "nivel": 3},
        ]},
    }

    def test_dre_multi_periodo_uma_aba(self):
        """DRE com 2 períodos deve gerar 1 aba (não 2)."""
        demos = [
            {"tipo": "dre", "periodo": "DEZ/2024", "dados": self.DRE_P1},
            {"tipo": "dre", "periodo": "DEZ/2025", "dados": self.DRE_P2},
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "dre_comp.xlsx"
            export_excel_multi(demos, "Teste", path)
            wb = load_workbook(str(path))
            assert len(wb.sheetnames) == 1
            assert "DRE" in wb.sheetnames[0]

    def test_dre_multi_periodo_colunas_lado_a_lado(self):
        """Headers devem ter períodos lado a lado."""
        demos = [
            {"tipo": "dre", "periodo": "DEZ/2024", "dados": self.DRE_P1},
            {"tipo": "dre", "periodo": "DEZ/2025", "dados": self.DRE_P2},
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "dre_comp.xlsx"
            export_excel_multi(demos, "Teste", path)
            wb = load_workbook(str(path))
            ws = wb.active
            # Procura header row com os períodos
            found = False
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                vals = [v for v in row if v]
                if "DEZ/2024" in vals and "DEZ/2025" in vals:
                    found = True
                    break
            assert found, "Períodos não encontrados lado a lado nos headers"

    def test_dre_multi_periodo_valores_corretos(self):
        """Valores de cada período devem estar na coluna correta."""
        demos = [
            {"tipo": "dre", "periodo": "DEZ/2024", "dados": self.DRE_P1},
            {"tipo": "dre", "periodo": "DEZ/2025", "dados": self.DRE_P2},
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "dre_comp.xlsx"
            export_excel_multi(demos, "Teste", path)
            wb = load_workbook(str(path))
            ws = wb.active
            # Procura a primeira linha de dados (RECEITA BRUTA)
            # Layout: Desc | Per1 | Per2
            for row in ws.iter_rows(min_row=1, values_only=True):
                if row[0] and "RECEITA BRUTA" in str(row[0]):
                    assert row[1] == 200000, f"P1 esperado 200000, obteve {row[1]}"
                    assert row[2] == 250000, f"P2 esperado 250000, obteve {row[2]}"
                    break

    def test_balanco_multi_periodo_uma_aba(self):
        """Balanço com 2 períodos deve gerar 1 aba (não 2)."""
        demos = [
            {"tipo": "balanco_patrimonial", "periodo": "DEZ/2024", "dados": self.BAL_P1},
            {"tipo": "balanco_patrimonial", "periodo": "DEZ/2025", "dados": self.BAL_P2},
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "bal_comp.xlsx"
            export_excel_multi(demos, "Teste", path)
            wb = load_workbook(str(path))
            assert len(wb.sheetnames) == 1

    def test_balanco_multi_periodo_valores(self):
        """Valores de Balanço devem estar lado a lado por período."""
        demos = [
            {"tipo": "balanco_patrimonial", "periodo": "DEZ/2024", "dados": self.BAL_P1},
            {"tipo": "balanco_patrimonial", "periodo": "DEZ/2025", "dados": self.BAL_P2},
        ]
        no_formulas = {"dre": False, "balanco": False, "balancete": False}
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "bal_comp.xlsx"
            export_excel_multi(demos, "Teste", path, formula_opts=no_formulas)
            wb = load_workbook(str(path))
            ws = wb.active
            # Procura ATIVO total
            for row in ws.iter_rows(min_row=1, values_only=True):
                if row[0] and str(row[0]).strip() == "ATIVO":
                    assert row[1] == 300000, f"Ativo P1 esperado 300000, obteve {row[1]}"
                    assert row[2] == 350000, f"Ativo P2 esperado 350000, obteve {row[2]}"
                    break

    def test_balanco_valores_negativos_preservados(self):
        """Valores negativos (ex: Prejuízos Acumulados) devem ser preservados."""
        demos = [
            {"tipo": "balanco_patrimonial", "periodo": "DEZ/2024", "dados": self.BAL_P1},
            {"tipo": "balanco_patrimonial", "periodo": "DEZ/2025", "dados": self.BAL_P2},
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "bal_comp.xlsx"
            export_excel_multi(demos, "Teste", path)
            wb = load_workbook(str(path))
            ws = wb.active
            for row in ws.iter_rows(min_row=1, values_only=True):
                if row[0] and "Prejuízo" in str(row[0]):
                    assert row[1] == -50000, f"P1 esperado -50000, obteve {row[1]}"
                    assert row[2] == -20000, f"P2 esperado -20000, obteve {row[2]}"
                    break

    def test_mix_tipos_com_comparativo(self):
        """Balancete single + DRE multi-período + Balanço multi-período."""
        demos = [
            {"tipo": "balancete", "periodo": "12/2025", "dados": BALANCETE_DATA},
            {"tipo": "dre", "periodo": "DEZ/2024", "dados": self.DRE_P1},
            {"tipo": "dre", "periodo": "DEZ/2025", "dados": self.DRE_P2},
            {"tipo": "balanco_patrimonial", "periodo": "DEZ/2024", "dados": self.BAL_P1},
            {"tipo": "balanco_patrimonial", "periodo": "DEZ/2025", "dados": self.BAL_P2},
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "mix.xlsx"
            export_excel_multi(demos, "Teste", path)
            wb = load_workbook(str(path))
            # 1 aba balancete + 1 aba DRE comparativo + 1 aba Balanço comparativo = 3
            assert len(wb.sheetnames) == 3
